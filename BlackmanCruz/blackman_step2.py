"""
Blackman Cruz — Step 2
======================
Reads blackman_step1.xlsx, fetches each product detail page,
extracts dimensions / materials / origin / description,
and saves the final Blackman Cruz.xlsx in Julian Chichester format.

Usage:
    python blackman_step2.py
"""

import re
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

BASE_URL   = "https://blackmancruz.com"
STEP1_FILE = Path(__file__).parent / "blackman_step1.xlsx"
OUT_FILE   = Path(__file__).parent / "Blackman Cruz.xlsx"

VENDOR = "Blackman Cruz"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}

FIXED_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Finish",
    "Origin", "Materials", "Price", "Tags",
]


# ─── Dimension parser ──────────────────────────────────────────────────────────

def _parse_number(s: str) -> float | None:
    """Extract the first number from a string, taking the smaller if a range."""
    # Range like "79" max - 41" min" → take smaller
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if not nums:
        return None
    values = [float(n) for n in nums]
    return min(values)


def _inches(raw: str) -> str:
    """
    Convert a raw measurement string to inches (2 dp).
    Handles: 64.5", 64.5 in, 100 cm, 100mm
    """
    raw = raw.strip()
    m_cm = re.search(r"(\d+(?:\.\d+)?)\s*cm", raw, re.IGNORECASE)
    if m_cm:
        return f"{float(m_cm.group(1)) / 2.54:.2f}"
    m_mm = re.search(r"(\d+(?:\.\d+)?)\s*mm", raw, re.IGNORECASE)
    if m_mm:
        return f"{float(m_mm.group(1)) / 25.4:.2f}"
    # Already inches / quoted
    num = _parse_number(raw)
    if num is not None:
        return f"{num:.2f}"
    return ""


def parse_dimensions(dim_text: str) -> dict[str, str]:
    """
    Parse a raw dimension string into W/H/D/Diameter in inches.

    Handles:
      "64.5" H x 21.5" W x 16" D"
      "28" W x 24" D"
      "32" H x 30" Diam"
      "34" H ( not including chain ) x 15" Diam"
      "Arm/Back: 29" H / Seat 17" H 28" W x 24" D"
      "Adjustable Height: 79" max - 41" min Adjustable Width: 60" max - 41" min 3" Depth"
    """
    result = {"Width": "", "Depth": "", "Height": "", "Diameter": ""}

    # Diam pattern  (Diam / Diameter)
    m_dia = re.search(r'([\d.]+)["\s]*(?:Diam(?:eter)?)', dim_text, re.IGNORECASE)
    if m_dia:
        result["Diameter"] = _inches(m_dia.group(1))

    # Height — last/most relevant H occurrence (avoid "Seat 17" H" overriding "Arm 29" H")
    # Prefer "H" not preceded by "Seat" or "Arm"
    h_matches = re.findall(r'([\d.]+(?:\s*-\s*[\d.]+)?)["\s]*H\b', dim_text, re.IGNORECASE)
    # Filter out ones preceded by Seat / Arm
    h_clean = []
    for hm in h_matches:
        idx = dim_text.find(hm)
        prefix = dim_text[max(0, idx - 15):idx].lower()
        if "seat" not in prefix and "arm" not in prefix:
            h_clean.append(hm)
    if h_clean:
        result["Height"] = _inches(h_clean[0])
    elif h_matches:
        result["Height"] = _inches(h_matches[0])

    # Width
    m_w = re.search(r'([\d.]+)["\s]*W\b', dim_text, re.IGNORECASE)
    if m_w:
        result["Width"] = _inches(m_w.group(1))

    # Depth — "D" or "Depth"
    m_d = re.search(r'([\d.]+)["\s]*(?:D\b|Depth)', dim_text, re.IGNORECASE)
    if m_d:
        result["Depth"] = _inches(m_d.group(1))

    # Adjustable variants: take min value
    m_adj_h = re.search(r'Height[:\s]*([\d.]+)["\s]*max\s*-\s*([\d.]+)["\s]*min', dim_text, re.IGNORECASE)
    if m_adj_h:
        result["Height"] = _inches(m_adj_h.group(2))

    m_adj_w = re.search(r'Width[:\s]*([\d.]+)["\s]*max\s*-\s*([\d.]+)["\s]*min', dim_text, re.IGNORECASE)
    if m_adj_w:
        result["Width"] = _inches(m_adj_w.group(2))

    # If Diameter found, clear W and D (round item)
    if result["Diameter"]:
        result["Width"] = ""
        result["Depth"] = ""

    return result


# ─── Detail page scraper ───────────────────────────────────────────────────────

def scrape_detail(url: str) -> dict:
    """Fetch a product detail page and return extracted fields."""
    data: dict[str, str] = {}
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        data["_error"] = str(e)
        return data

    soup = BeautifulSoup(r.text, "html.parser")
    for tag in soup.find_all(["script", "style", "svg", "noscript"]):
        tag.decompose()

    # Title
    h1 = soup.find("h1")
    if h1:
        data["Product Name"] = h1.get_text(strip=True)

    # OG image (highest quality)
    og_img = soup.find("meta", property="og:image")
    if og_img:
        data["Image URL"] = og_img.get("content", "")

    # Description block
    desc_div = soup.find("div", class_=lambda c: c and "product-content__desc" in str(c))
    if desc_div:
        desc_text = desc_div.get_text(" ", strip=True)
        if desc_text:
            data["Description"] = desc_text

    # Key-value spec block: div.flex.flex-col.gap-2.text-xs
    spec_section = soup.find(
        "div",
        class_=lambda c: c and "flex-col" in str(c) and "gap-2" in str(c) and "text-xs" in str(c),
    )
    if spec_section:
        for row in spec_section.find_all(
            "div",
            class_=lambda c: c and "flex" in str(c) and "gap-4" in str(c),
        ):
            h5 = row.find("h5")
            val_div = row.find(
                "div",
                class_=lambda c: not c or ("gap-4" not in str(c) and "flex" not in str(c)),
            )
            if not h5 or not val_div:
                continue
            label = h5.get_text(strip=True)
            value = val_div.get_text(" ", strip=True)

            if label == "Dimensions":
                dims = parse_dimensions(value)
                data.update(dims)
                data["_raw_dimensions"] = value
            elif label == "Materials":
                data["Finish"] = value
                data["Materials"] = value
            elif label == "Origin":
                data["Origin"] = value

    return data


# ─── SKU generation ────────────────────────────────────────────────────────────

def generate_sku(category: str, index: int) -> str:
    """BLA + category initials + 2-digit index."""
    words = re.sub(r"[^A-Za-z\s]", "", category).strip().split()
    if len(words) >= 2:
        c = (words[0][0] + words[1][0]).upper()
    elif words:
        c = words[0][:2].upper()
    else:
        c = "XX"
    return f"BLA{c}{index:02d}"


# ─── Excel helpers ─────────────────────────────────────────────────────────────

def init_workbook() -> openpyxl.Workbook:
    try:
        wb = openpyxl.load_workbook(str(OUT_FILE))
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        for s in list(wb.sheetnames):
            del wb[s]
    return wb


def save_sheet(
    wb: openpyxl.Workbook,
    category_name: str,
    collection_url: str,
    rows: list[dict],
) -> None:
    sheet_name = category_name[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    bold = Font(bold=True)

    ws.cell(1, 1, "Brand").font = bold
    ws.cell(1, 2, VENDOR)
    ws.cell(2, 1, "Category Link").font = bold
    ws.cell(2, 2, collection_url)

    # Collect any extra columns beyond FIXED_COLS
    extra: list[str] = []
    for row in rows:
        for k in row:
            if k not in FIXED_COLS and not k.startswith("_") and k not in extra:
                extra.append(k)
    all_cols = FIXED_COLS + extra

    for col, h in enumerate(all_cols, 1):
        ws.cell(4, col, h).font = bold

    for i, row in enumerate(rows, 1):
        row.setdefault("Index", i)
        row.setdefault("Category", category_name)
        row.setdefault("Manufacturer", VENDOR)
        if not row.get("SKU"):
            row["SKU"] = generate_sku(category_name, i)
        if not row.get("Product Family Id"):
            row["Product Family Id"] = row.get("Product Name", "")

        for col, key in enumerate(all_cols, 1):
            ws.cell(4 + i, col, row.get(key, ""))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A5"


# ─── Main ──────────────────────────────────────────────────────────────────────

def main(demo: bool = False) -> None:
    if not STEP1_FILE.exists():
        print(f"ERROR: {STEP1_FILE} not found. Run blackman_step1.py first.")
        return

    demo_limit = 5 if demo else None
    out_path = Path(__file__).parent / ("Blackman Cruz_demo.xlsx" if demo else "Blackman Cruz.xlsx")

    if demo:
        print("[DEMO MODE] 5 products per category -> Blackman Cruz_demo.xlsx")

    wb_in = openpyxl.load_workbook(str(STEP1_FILE))

    try:
        wb_out = openpyxl.load_workbook(str(out_path))
    except FileNotFoundError:
        wb_out = openpyxl.Workbook()
        for s in list(wb_out.sheetnames):
            del wb_out[s]

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        category_name = sheet_name

        # Read step1 rows (row 1 = header, rows 2+ = data)
        # Columns: #, Category, Product Name, Handle, URL, Image URL, Tags, Price
        rows_in = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[2]:  # no product name
                continue
            rows_in.append({
                "index": row[0],
                "category": row[1],
                "title": row[2],
                "handle": row[3],
                "url": row[4],
                "image_url": row[5],
                "tags": row[6] or "",
                "price": row[7] or "",
            })

        if demo_limit:
            rows_in = rows_in[:demo_limit]

        collection_url = f"{BASE_URL}/collections/{sheet_name.lower().replace(' ', '-').replace('bc-workshop','bc-workshop')}"
        # Fix: use slug from the known mapping
        slug_map = {
            "Lighting": "lighting", "Seating": "seating", "Tables": "tables",
            "Art": "art", "Case Goods": "case-goods", "Garden": "garden",
            "Objects": "objects", "Mirrors": "mirrors", "BC Workshop": "bc-workshop",
        }
        slug = slug_map.get(category_name, category_name.lower().replace(" ", "-"))
        collection_url = f"{BASE_URL}/collections/{slug}"

        print(f"\n{'='*60}")
        print(f"[{category_name}] {len(rows_in)} products")
        print(f"{'='*60}")

        output_rows: list[dict] = []

        for i, p_in in enumerate(rows_in, 1):
            url = p_in["url"]
            safe_title = p_in['title'][:60].encode('ascii', 'replace').decode('ascii')
            print(f"  [{i}/{len(rows_in)}] {safe_title} ...", end=" ", flush=True)

            detail = scrape_detail(url)

            if "_error" in detail:
                print(f"ERROR: {detail['_error']}")
            else:
                print("OK")

            row: dict = {}
            row["Product Name"] = detail.get("Product Name") or p_in["title"]
            row["Source"]       = url
            row["Image URL"]    = detail.get("Image URL") or p_in["image_url"]
            row["Description"]  = detail.get("Description", "")
            row["Width"]        = detail.get("Width", "")
            row["Depth"]        = detail.get("Depth", "")
            row["Height"]       = detail.get("Height", "")
            row["Diameter"]     = detail.get("Diameter", "")
            row["Finish"]       = detail.get("Finish", "")
            row["Origin"]       = detail.get("Origin", "")
            row["Materials"]    = detail.get("Materials", "")
            row["Price"]        = p_in["price"]
            row["Tags"]         = p_in["tags"]

            output_rows.append(row)
            time.sleep(0.4)

        save_sheet(wb_out, category_name, collection_url, output_rows)
        wb_out.save(str(out_path))
        print(f"  Saved sheet '{category_name}' ({len(output_rows)} products)")

    wb_out.save(str(out_path))
    print(f"\nDone -> {out_path}")


if __name__ == "__main__":
    import sys
    main(demo="--demo" in sys.argv)
