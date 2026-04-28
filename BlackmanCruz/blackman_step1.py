"""
Blackman Cruz — Step 1
======================
Categories and URLs read from SD_Web Scraping - Status Tracker.xlsx
(sheet: "Blackman Cruz"). For each category, scrapes all product handles
from the sub-collection HTML pages, then fetches product details via the
Shopify product JSON API.

Saves to BlackmanCruz/blackman_step1.xlsx (one sheet per category).

Usage:
    python blackman_step1.py
"""

import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
TRACKER_FILE = SCRIPT_DIR.parent / "SD_Web Scraping - Status Tracker.xlsx"
OUT_FILE     = SCRIPT_DIR / "blackman_step1.xlsx"
BASE_URL     = "https://blackmancruz.com"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}


# ── Read categories from Status Tracker ───────────────────────────────────────

def read_categories_from_tracker() -> list[dict]:
    """
    Parse 'Blackman Cruz' sheet in Status Tracker.
    Returns list of {category, urls: [str, ...]}
    Only includes categories that have at least one URL.
    """
    wb = openpyxl.load_workbook(str(TRACKER_FILE))
    ws = wb["Blackman Cruz"]

    categories = []
    current_cat = None
    current_urls = []

    for r in range(1, ws.max_row + 1):
        col3 = ws.cell(r, 3).value  # 'Category' or 'Link' or None
        col4 = ws.cell(r, 4).value  # category name or URL or None

        if str(col3).strip() == "Category" and col4:
            # Save previous category if it had URLs
            if current_cat and current_urls:
                categories.append({"category": current_cat, "urls": current_urls})
            current_cat = str(col4).strip()
            current_urls = []

        elif str(col3).strip() == "Link" and col4:
            url = str(col4).strip()
            if url.startswith("http"):
                current_urls.append(url)

        elif col4 and str(col4).strip().startswith("http"):
            # Additional URL row (no label in col3)
            current_urls.append(str(col4).strip())

    # Save last category
    if current_cat and current_urls:
        categories.append({"category": current_cat, "urls": current_urls})

    wb.close()
    return categories


# ── Scrape product handles from collection page ────────────────────────────────

def scrape_handles_from_page(url: str) -> list[str]:
    """
    Fetch a collection/sub-collection HTML page and extract unique product handles.
    Handles pagination via ?page=N.
    """
    handles: list[str] = []
    page = 1

    while True:
        page_url = f"{url}?page={page}" if page > 1 else url
        try:
            r = requests.get(page_url, headers=HEADERS, timeout=30)
            if r.status_code != 200:
                break
        except Exception as e:
            print(f"    fetch error: {e}")
            break

        soup = BeautifulSoup(r.text, "html.parser")
        found_on_page = []

        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "/products/" in href:
                # Extract handle — normalize to /products/{handle}
                handle = href.split("/products/")[-1].split("?")[0].strip("/")
                if handle and handle not in handles and handle not in found_on_page:
                    found_on_page.append(handle)

        if not found_on_page:
            break

        handles.extend(found_on_page)
        print(f"    page {page}: +{len(found_on_page)} handles ({len(handles)} total)")

        # Check for next page link
        next_link = soup.find("a", href=lambda h: h and f"page={page + 1}" in str(h))
        if not next_link:
            break
        page += 1
        time.sleep(0.4)

    return handles


# ── Fetch product details via Shopify JSON API ─────────────────────────────────

def fetch_product_json(handle: str) -> dict | None:
    """GET /products/{handle}.json and return the product dict."""
    url = f"{BASE_URL}/products/{handle}.json"
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        if r.status_code == 200:
            return r.json().get("product", {})
    except Exception:
        pass
    return None


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if not TRACKER_FILE.exists():
        print(f"ERROR: Status Tracker not found at {TRACKER_FILE}")
        return

    print("Reading categories from Status Tracker...")
    categories = read_categories_from_tracker()
    print(f"Found {len(categories)} categories with URLs\n")
    for cat in categories:
        print(f"  {cat['category']}: {len(cat['urls'])} URL(s)")

    wb = openpyxl.Workbook()
    for s in list(wb.sheetnames):
        del wb[s]

    bold = Font(bold=True)
    col_headers = ["#", "Category", "Product Name", "Handle", "URL",
                   "Image URL", "Tags", "Product Type", "Price", "Collection URLs"]

    for cat_info in categories:
        category_name = cat_info["category"]
        urls = cat_info["urls"]

        print(f"\n{'='*60}")
        print(f"[{category_name}] ({len(urls)} URL(s))")
        print(f"{'='*60}")

        # Collect unique handles across all URLs for this category
        all_handles: list[str] = []
        seen: set[str] = set()

        for url in urls:
            print(f"  Scraping: {url}")
            handles = scrape_handles_from_page(url)
            new = [h for h in handles if h not in seen]
            all_handles.extend(new)
            seen.update(new)
            time.sleep(0.5)

        print(f"  Total unique products: {len(all_handles)}")

        # Create sheet
        sheet_name = category_name[:31]
        ws = wb.create_sheet(sheet_name)
        for col, h in enumerate(col_headers, 1):
            ws.cell(1, col, h).font = bold

        # Fetch product JSON for each handle
        rows = []
        for i, handle in enumerate(all_handles, 1):
            print(f"  [{i}/{len(all_handles)}] {handle[:55]}", end=" ", flush=True)
            p = fetch_product_json(handle)
            if p:
                title = p.get("title", "")
                tags = ", ".join(p.get("tags", []))
                product_type = p.get("product_type", "")
                images = p.get("images", [])
                img_src = images[0].get("src", "") if images else ""
                variants = p.get("variants", [])
                price = variants[0].get("price", "") if variants else ""
                print(f"OK - {title[:40]}")
            else:
                title = ""
                tags = ""
                product_type = ""
                img_src = ""
                price = ""
                print("FAIL")

            product_url = f"{BASE_URL}/products/{handle}"
            collection_urls = ", ".join(urls)

            row = [i, category_name, title, handle, product_url,
                   img_src, tags, product_type, price, collection_urls]
            rows.append(row)
            time.sleep(0.3)

        # Write rows to sheet
        for i, row in enumerate(rows, 1):
            for col, val in enumerate(row, 1):
                ws.cell(i + 1, col, val)

        # Auto width
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"

        wb.save(str(OUT_FILE))
        print(f"  Saved '{sheet_name}' ({len(rows)} products)")
        time.sleep(0.3)

    wb.save(str(OUT_FILE))
    print(f"\nDone -> {OUT_FILE}")


if __name__ == "__main__":
    main()
