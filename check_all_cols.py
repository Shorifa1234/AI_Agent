import openpyxl
from pathlib import Path

FILES = {
    "Worlds Away":       r"d:\Automation\SKU_AGENT\WorldsAway\Worlds Away.xlsx",
    "Vanguard":          r"d:\Automation\SKU_AGENT\Vanguard\Vanguard.xlsx",
    "Blackman Cruz":     r"d:\Automation\SKU_AGENT\BlackmanCruz\Blackman Cruz.xlsx",
    "Crystorama":        r"d:\Automation\SKU_AGENT\Crystorama\Crystorama.xlsx",
    "CrystoramaLights":  r"d:\Automation\SKU_AGENT\CrystoramaLights\CrystoramaLights_demo.xlsx",
    "Brownstone":        r"d:\Automation\SKU_AGENT\Brownstone\Brownstone.xlsx",
    "Wesley Hall":       r"d:\Automation\SKU_AGENT\Wesley Hall\Wesley Hall.xlsx",
    "Blackman Cruz demo":r"d:\Automation\SKU_AGENT\BlackmanCruz\Blackman Cruz_demo.xlsx",
    "Gabby demo":        r"d:\Automation\SKU_AGENT\Gabby\Gabby_demo.xlsx",
    "Paul Ferrante":     r"d:\Automation\SKU_AGENT\Paul Ferrante\Paul Ferrante.xlsx",
    "Invisible":         r"d:\Automation\SKU_AGENT\Invisible Collection\Invisible Collection.xlsx",
    "Julian Chichester": r"d:\Automation\SKU_AGENT\Julian Chichester.xlsx",
}

all_cols = set()

for vendor, path in FILES.items():
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[0]
        # Try row 4 first
        h4 = [ws.cell(4,c).value for c in range(1, ws.max_column+1) if ws.cell(4,c).value]
        h1 = [ws.cell(1,c).value for c in range(1, ws.max_column+1) if ws.cell(1,c).value]
        headers = h4 if len(h4) > 3 else h1
        print(f"\n{vendor} [{len(headers)} cols]:")
        for h in headers:
            print(f"  - {h}")
            all_cols.add(str(h))
    except Exception as e:
        print(f"\n{vendor}: SKIP - {e}")

print("\n" + "="*60)
print(f"ALL UNIQUE COLUMNS ACROSS ALL VENDORS ({len(all_cols)}):")
for c in sorted(all_cols):
    print(f"  {c}")
