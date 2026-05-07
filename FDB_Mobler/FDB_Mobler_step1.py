"""
FDB Mobler - Step 1: List Page Scraper
Collects all product URLs from all category URLs
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
import time

# Headers for requests
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

# Category → list of URLs to scrape
CATEGORIES = {
    'Coffee & Cocktail Tables': ['https://www.fdbmobler.com/collections/coffee-tables'],
    'Side & End Tables':        ['https://www.fdbmobler.com/collections/side-tables'],
    'Dining Tables':            ['https://www.fdbmobler.com/collections/dining-tables'],
    'Consoles':                 ['https://www.fdbmobler.com/collections/sideboard'],
    'Desks':                    ['https://www.fdbmobler.com/collections/desks'],
    'Bookcases':                ['https://www.fdbmobler.com/collections/shelves',
                                 'https://www.fdbmobler.com/collections/book-cases',
                                 'https://www.fdbmobler.com/collections/shelves-1'],
    'Cabinets':                 ['https://www.fdbmobler.com/collections/display-cabinets'],
    'Dining Chairs':            ['https://www.fdbmobler.com/collections/dining-table-chairs-1'],
    'Bar Stools':               ['https://www.fdbmobler.com/collections/bar-stools',
                                 'https://www.fdbmobler.com/collections/stools'],
    'Sofas & Loveseats':        ['https://www.fdbmobler.com/collections/module-sofas',
                                 'https://www.fdbmobler.com/collections/2-person-sofa',
                                 'https://www.fdbmobler.com/collections/2-5-person-sofa',
                                 'https://www.fdbmobler.com/collections/cushions-for-sofas'],
    'Lounge Chairs':            ['https://www.fdbmobler.com/collections/armchairs',
                                 'https://www.fdbmobler.com/collections/cushions-for-easy-chairs'],
    'Benches':                  ['https://www.fdbmobler.com/collections/benches'],
    'Pendants':                 ['https://www.fdbmobler.com/collections/pedants'],
    'Sconces':                  ['https://www.fdbmobler.com/collections/wall-lamps'],
    'Table Lamps':              ['https://www.fdbmobler.com/collections/table-lamps'],
    'Floor Lamps':              ['https://www.fdbmobler.com/collections/floor-lamps'],
    'Mirrors':                  ['https://www.fdbmobler.com/collections/mirrors'],
    'Pillows & Throws':         ['https://www.fdbmobler.com/collections/cushions'],
    'Vases':                    ['https://www.fdbmobler.com/collections/vases'],
    'Boxes':                    ['https://www.fdbmobler.com/collections/boxes',
                                 'https://www.fdbmobler.com/collections/apple-boxes'],
    'Outdoor Seating':          ['https://www.fdbmobler.com/collections/garden-chairs',
                                 'https://www.fdbmobler.com/collections/lounge-furniture',
                                 'https://www.fdbmobler.com/collections/garden-benches'],
    'Outdoor Tables':           ['https://www.fdbmobler.com/collections/garden-tables'],
    'Outdoor Accessories':      ['https://www.fdbmobler.com/collections/accessories-for-the-garden'],
}

def scrape_page(url):
    """Fetch a page with retry, return soup or None"""
    for attempt in range(3):
        try:
            response = requests.get(url, headers=HEADERS, timeout=15)
            if response.status_code == 200:
                return BeautifulSoup(response.content, 'html.parser')
            print(f"  Status {response.status_code}, retrying...")
            time.sleep(3)
        except Exception as e:
            print(f"  ERROR: {e}")
            time.sleep(3)
    return None

def scrape_single_url(category_name, start_url):
    """Scrape all pages of one collection URL, return set of product URLs"""
    product_urls = set()
    current_url = start_url
    page_num = 1

    while current_url:
        soup = scrape_page(current_url)
        if not soup:
            print(f"    Failed to load page {page_num} of {start_url}")
            break

        for card in soup.find_all('product-card'):
            fig = card.find('div', class_='product-card__figure')
            if fig:
                a = fig.find('a', href=True)
                if a and '/products/' in a['href']:
                    href = a['href'].split('?')[0].split('#')[0]
                    full_url = 'https://www.fdbmobler.com' + href if href.startswith('/') else href
                    product_urls.add(full_url)

        next_link = soup.find('a', rel='next')
        if next_link and next_link.get('href'):
            href = next_link['href']
            current_url = 'https://www.fdbmobler.com' + href if href.startswith('/') else href
            page_num += 1
            time.sleep(0.5)
        else:
            break

    return product_urls, page_num

def scrape_category_products(category_name, category_urls):
    """Scrape all product URLs from all URLs of a category (deduped)"""
    print(f"Scraping: {category_name} ({len(category_urls)} URL{'s' if len(category_urls) > 1 else ''})...")

    all_urls = set()
    for url in category_urls:
        urls, pages = scrape_single_url(category_name, url)
        print(f"  {url.split('/')[-1]}: {len(urls)} products ({pages} page{'s' if pages > 1 else ''})")
        all_urls.update(urls)
        time.sleep(0.5)

    products = [{'category': category_name, 'product_url': url, 'image_url': '', 'product_name': '', 'sku': ''} for url in all_urls]
    print(f"  Total: {len(products)} unique products")
    return products

def save_to_excel(all_products, output_file):
    """Save all scraped data to Excel"""
    print(f"\nSaving to {output_file}...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group products by category
    categories = {}
    for product in all_products:
        cat = product['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(product)

    # Create one sheet per category
    for category_name, products in categories.items():
        # Create safe sheet name (max 31 chars)
        sheet_name = category_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        ws['A1'] = 'Category'
        ws['B1'] = 'Product URL'
        ws['C1'] = 'Image URL'
        ws['D1'] = 'Product Name'
        ws['E1'] = 'SKU'

        # Make header bold
        for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
            ws[col].font = Font(bold=True)

        # Write data rows
        for idx, product in enumerate(products, 2):
            ws[f'A{idx}'] = product['category']
            ws[f'B{idx}'] = product['product_url']
            ws[f'C{idx}'] = product['image_url']
            ws[f'D{idx}'] = product['product_name']
            ws[f'E{idx}'] = product['sku']

    wb.save(output_file)
    print(f"[OK] Saved {len(all_products)} products to {output_file}")

def main():
    print("=" * 80)
    print("FDB Mobler - Step 1: List Page Scraper")
    print("=" * 80)
    print()

    # Scrape each category
    all_products = []
    total = len(CATEGORIES)
    for idx, (cat_name, cat_urls) in enumerate(CATEGORIES.items(), 1):
        print(f"[{idx}/{total}] ", end='')
        products = scrape_category_products(cat_name, cat_urls)
        all_products.extend(products)
        time.sleep(1)

    print()
    print("=" * 80)
    print(f"SUMMARY:")
    print(f"  Categories scraped: {total}")
    print(f"  Total products: {len(all_products)}")
    print("=" * 80)

    # Save to Excel
    save_to_excel(all_products, 'FDB_Mobler_step1.xlsx')
    print("\n[OK] Step 1 complete!")

if __name__ == '__main__':
    main()
