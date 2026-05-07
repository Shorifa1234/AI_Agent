"""
FDB Mobler - Step 2: Detail Page Scraper
Reads product URLs from Step 1 Excel, extracts full details, generates SKUs, creates final output
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
import json
import re
import time
from collections import defaultdict

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

# SKU generation: FDBXX##
VENDOR_CODE = "FDB"

# Category codes mapping
CATEGORY_CODES = {
    'Coffee & Cocktail Tables': 'CO',
    'Side & End Tables': 'SI',
    'Dining Tables': 'DI',
    'Consoles': 'CN',
    'Desks': 'DE',
    'Bookcases': 'BO',
    'Cabinets': 'CA',
    'Dining Chairs': 'DC',
    'Bar Stools': 'BS',
    'Sofas & Loveseats': 'SO',
    'Lounge Chairs': 'LO',
    'Benches': 'BE',
    'Pendants': 'PE',
    'Sconces': 'SC',
    'Table Lamps': 'TL',
    'Floor Lamps': 'FL',
    'Mirrors': 'MI',
    'Pillows & Throws': 'PI',
    'Vases': 'VA',
    'Boxes': 'BX',
    'Outdoor Seating': 'OS',
    'Outdoor Tables': 'OT',
    'Outdoor Accessories': 'OA',
}

def extract_product_details(url):
    """Extract full product details from detail page"""
    try:
        response = requests.get(url, headers=HEADERS, timeout=15)

        if response.status_code != 200:
            print(f"    ERROR: Status {response.status_code}")
            return None

        soup = BeautifulSoup(response.content, 'html.parser')

        # Try to extract JSON-LD data
        json_ld_scripts = soup.find_all('script', {'type': 'application/ld+json'})
        product_data = {}

        for script in json_ld_scripts:
            try:
                data = json.loads(script.string)
                if data.get('@type') == 'Product':
                    product_data = data
                    break
            except:
                pass

        # If no JSON-LD, try HTML extraction
        if not product_data:
            # Try to get title from h1
            title_tag = soup.find('h1')
            if not title_tag:
                print(f"    WARNING: No product data found")
                return None

            product_data['name'] = title_tag.get_text(strip=True)

            # Try to get description
            desc_div = soup.find('div', {'class': lambda x: x and 'description' in str(x).lower()})
            if desc_div:
                product_data['description'] = desc_div.get_text(strip=True)

            # Try to get image
            img_tag = soup.find('img', {'src': lambda x: x and 'products/' in str(x)})
            if img_tag:
                product_data['image'] = img_tag.get('src', '')

        # Extract data
        name = product_data.get('name', '')
        brand = name  # Product Family Id = Product Name

        # Description from section-stack__intro > .prose (not feature-chart prose)
        description = ''
        intro = soup.find('div', class_='section-stack__intro')
        if intro:
            prose_div = intro.find('div', class_='prose')
            if prose_div:
                for tag in prose_div.find_all(['h1', 'h2', 'h3', 'h4']):
                    tag.decompose()
                for tag in prose_div.find_all('div', class_='feature-chart__certifications'):
                    tag.decompose()
                description = prose_div.get_text(separator=' ', strip=True)
        if not description:
            description = product_data.get('description', '')
        for prefix in ['Info about the product', 'Info om produktet']:
            if description.startswith(prefix):
                description = description[len(prefix):].strip()

        # Extract image URL
        image_url = ''
        if 'image' in product_data:
            images = product_data['image']
            if isinstance(images, list):
                if isinstance(images[0], dict):
                    image_url = images[0].get('url', images[0].get('contentUrl', ''))
                else:
                    image_url = images[0]
            elif isinstance(images, dict):
                image_url = images.get('url', images.get('contentUrl', ''))
            else:
                image_url = str(images)

        # Fallback 1: og:image meta tag
        if not image_url:
            og_img = soup.find('meta', property='og:image')
            if og_img:
                image_url = og_img.get('content', '')

        # Fallback 2: CDN image from HTML — skip variant-picker/swatch thumbnails
        if not image_url:
            for img in soup.find_all('img', {'src': lambda x: x and 'cdn/shop' in str(x)}):
                if img.find_parent(class_=lambda x: x and ('thumbnail-swatch' in str(x) or 'variant-picker' in str(x))):
                    continue
                image_url = img.get('src', '')
                break

        # Make sure image URL is absolute
        if image_url and image_url.startswith('//'):
            image_url = 'https:' + image_url
        elif image_url and not image_url.startswith('http'):
            image_url = 'https://www.fdbmobler.com' + image_url

        # Extract dimensions from feature-chart (Measurements section)
        width = depth = height = diameter = ''
        finish_parts = []

        chart_rows = soup.find_all('div', class_='feature-chart__table-row')
        spec = {}
        for row in chart_rows:
            heading = row.find('div', class_='feature-chart__heading')
            value = row.find('div', class_='feature-chart__value')
            if heading and value:
                spec[heading.get_text(strip=True).lower()] = value.get_text(strip=True)

        def cm_to_in(val):
            try:
                return str(round(float(val) / 2.54, 2))
            except:
                return ''

        # Danish: Bredde=Width, Dybde=Depth, Højde=Height, Diameter=Diameter
        # Prefer "samlet" (assembled) values
        for key in ['bredde samlet', 'bredde']:
            if key in spec:
                width = cm_to_in(spec[key]); break
        for key in ['dybde samlet', 'dybde']:
            if key in spec:
                depth = cm_to_in(spec[key]); break
        for key in ['højde samlet', 'hojde samlet', 'højde', 'hojde']:
            if key in spec:
                height = cm_to_in(spec[key]); break
        for key in ['diameter samlet', 'diameter']:
            if key in spec:
                diameter = cm_to_in(spec[key]); break

        # Fallback: diameter from product title (Ø55)
        if not diameter:
            diam_match = re.search(r'[ØOø](\d+)', name)
            if diam_match:
                diameter = cm_to_in(diam_match.group(1))

        # Finish from feature-chart material fields
        for key in ['frame material', 'frame color', 'frame surface treatment',
                    'seat material', 'seat color', 'upholstery material']:
            if key in spec and spec[key]:
                finish_parts.append(spec[key])
        finish = ', '.join(finish_parts) if finish_parts else ''

        # Lighting spec fields as separate values
        def get_spec(key):
            return spec.get(key, '')

        lightbulb    = get_spec('lightbulb')
        ip_class     = get_spec('ip class')
        voltage      = get_spec('electrical voltage (volts)')
        ceiling_cup  = get_spec('ceiling cup')
        cable_len    = get_spec('ledningslængde') or get_spec('ledningslaengde')
        lamp_shade   = get_spec('lamp shade')
        cable_color  = get_spec('cable color')
        launch_year  = get_spec('launch year')
        ean          = get_spec('ean-nummer')
        seat_height        = get_spec('seat height')
        seat_width         = get_spec('seat width')
        seat_depth         = get_spec('seat depth')
        arm_height         = get_spec('arm height')
        arm_width          = get_spec('arm width')
        warranty           = get_spec('warranty')
        upholstery_comp    = get_spec('polsterkomposition') or get_spec('upholstery composition')
        upholstery_color   = get_spec('polsterfarve') or get_spec('upholstery color')
        fill_material      = get_spec('fyld i tekstiler') or get_spec('fill material')
        martindale         = get_spec('martindale')
        pilling            = get_spec('fnugdannelse (pilling)') or get_spec('pilling')
        light_fastness     = get_spec('lysægthed') or get_spec('lysaegthed') or get_spec('light fastness')
        additional_info    = get_spec('yderligere information') or get_spec('additional information')

        return {
            'name': name,
            'brand': brand,
            'description': description,
            'image_url': image_url,
            'width': width,
            'depth': depth,
            'height': height,
            'diameter': diameter,
            'finish': finish,
            'lightbulb': lightbulb,
            'ip_class': ip_class,
            'voltage': voltage,
            'ceiling_cup': ceiling_cup,
            'cable_length': cable_len,
            'lamp_shade': lamp_shade,
            'cable_color': cable_color,
            'launch_year': launch_year,
            'ean': ean,
            'seat_height': seat_height,
            'seat_width': seat_width,
            'seat_depth': seat_depth,
            'arm_height': arm_height,
            'arm_width': arm_width,
            'warranty': warranty,
            'upholstery_comp': upholstery_comp,
            'upholstery_color': upholstery_color,
            'fill_material': fill_material,
            'martindale': martindale,
            'pilling': pilling,
            'light_fastness': light_fastness,
            'additional_info': additional_info,
        }

    except Exception as e:
        print(f"    ERROR: {e}")
        return None

def read_step1_data(input_file):
    """Read all product URLs from Step 1 Excel"""
    print(f"Reading {input_file}...")

    wb = openpyxl.load_workbook(input_file)
    products_by_category = defaultdict(list)

    for sheet in wb.worksheets:
        category = sheet.title

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:  # Product URL in column B
                products_by_category[category].append({
                    'category': row[0] if row[0] else category,
                    'product_url': row[1],
                })

    total = sum(len(prods) for prods in products_by_category.values())
    print(f"  Found {total} products across {len(products_by_category)} categories\n")

    return products_by_category

def generate_sku(category, sequence):
    """Generate SKU in format FDBXX##"""
    category_code = CATEGORY_CODES.get(category, 'XX')
    return f"{VENDOR_CODE}{category_code}{sequence:02d}"

def save_final_output(products_by_category, output_file):
    """Save to final Excel in standard format"""
    print(f"\nSaving to {output_file}...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for category, products in products_by_category.items():
        # Create safe sheet name
        sheet_name = category[:31]
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: Brand
        ws['A1'] = 'Brand'
        ws['B1'] = 'FDB Mobler'
        ws['A1'].font = Font(bold=True)

        # Row 2: Category Link (use first product's URL to derive category URL)
        ws['A2'] = 'Category Link'
        if products:
            # Derive category link from first product (not perfect but reasonable)
            ws['B2'] = ''  # Will leave empty for now
        ws['A2'].font = Font(bold=True)

        # Row 3: Empty

        # Row 4: Headers
        headers = ['Index', 'Category', 'Manufacturer', 'Source', 'Image URL', 'Product Name',
                   'SKU', 'Product Family Id', 'Description', 'Width', 'Depth', 'Height',
                   'Diameter', 'Finish', 'Lightbulb', 'IP Class', 'Voltage',
                   'Ceiling Cup', 'Cable Length', 'Lamp Shade', 'Cable Color',
                   'Launch Year', 'EAN', 'Seat Height', 'Seat Width', 'Seat Depth',
                   'Arm Height', 'Arm Width', 'Warranty',
                   'Upholstery Composition', 'Upholstery Color', 'Fill Material',
                   'Martindale', 'Pilling', 'Light Fastness', 'Additional Info']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        # Row 5+: Data
        for idx, product in enumerate(products, 1):
            row = 4 + idx

            ws.cell(row, 1).value = idx  # Index
            ws.cell(row, 2).value = product['category']  # Category
            ws.cell(row, 3).value = 'FDB Mobler'  # Manufacturer
            ws.cell(row, 4).value = product['product_url']  # Source
            ws.cell(row, 5).value = product.get('image_url', '')  # Image URL
            ws.cell(row, 6).value = product.get('name', '')  # Product Name
            ws.cell(row, 7).value = product.get('sku', '')  # SKU
            ws.cell(row, 8).value = product.get('brand', '')  # Product Family Id
            ws.cell(row, 9).value = product.get('description', '')  # Description
            ws.cell(row, 10).value = product.get('width', '')  # Width
            ws.cell(row, 11).value = product.get('depth', '')  # Depth
            ws.cell(row, 12).value = product.get('height', '')  # Height
            ws.cell(row, 13).value = product.get('diameter', '')  # Diameter
            ws.cell(row, 14).value = product.get('finish', '')
            ws.cell(row, 15).value = product.get('lightbulb', '')
            ws.cell(row, 16).value = product.get('ip_class', '')
            ws.cell(row, 17).value = product.get('voltage', '')
            ws.cell(row, 18).value = product.get('ceiling_cup', '')
            ws.cell(row, 19).value = product.get('cable_length', '')
            ws.cell(row, 20).value = product.get('lamp_shade', '')
            ws.cell(row, 21).value = product.get('cable_color', '')
            ws.cell(row, 22).value = product.get('launch_year', '')
            ws.cell(row, 23).value = product.get('ean', '')
            ws.cell(row, 24).value = product.get('seat_height', '')
            ws.cell(row, 25).value = product.get('seat_width', '')
            ws.cell(row, 26).value = product.get('seat_depth', '')
            ws.cell(row, 27).value = product.get('arm_height', '')
            ws.cell(row, 28).value = product.get('arm_width', '')
            ws.cell(row, 29).value = product.get('warranty', '')
            ws.cell(row, 30).value = product.get('upholstery_comp', '')
            ws.cell(row, 31).value = product.get('upholstery_color', '')
            ws.cell(row, 32).value = product.get('fill_material', '')
            ws.cell(row, 33).value = product.get('martindale', '')
            ws.cell(row, 34).value = product.get('pilling', '')
            ws.cell(row, 35).value = product.get('light_fastness', '')
            ws.cell(row, 36).value = product.get('additional_info', '')

    wb.save(output_file)

    total = sum(len(prods) for prods in products_by_category.values())
    print(f"[OK] Saved {total} products to {output_file}")

def main():
    print("=" * 80)
    print("FDB Mobler - Step 2: Detail Page Scraper")
    print("=" * 80)
    print()

    # Read Step 1 data
    products_by_category = read_step1_data('FDB_Mobler_step1.xlsx')

    # Process each category
    enriched_products = defaultdict(list)
    total_processed = 0
    total_count = sum(len(prods) for prods in products_by_category.values())

    for category, products in products_by_category.items():
        print(f"\nProcessing: {category} ({len(products)} products)")

        for idx, product in enumerate(products, 1):
            total_processed += 1
            print(f"  [{total_processed}/{total_count}] {product['product_url'][:60]}...")

            # Extract details
            details = extract_product_details(product['product_url'])

            if details:
                # Generate SKU
                sku = generate_sku(category, idx)

                # Merge data
                enriched = {**product, **details, 'sku': sku}
                enriched_products[category].append(enriched)

            time.sleep(0.3)  # Rate limiting

    print("\n" + "=" * 80)
    print("SUMMARY:")
    print(f"  Total products processed: {total_processed}")
    print(f"  Total products enriched: {sum(len(p) for p in enriched_products.values())}")
    print("=" * 80)

    # Save final output
    save_final_output(enriched_products, 'FDB Mobler.xlsx')
    print("\n[OK] Step 2 complete!")

if __name__ == '__main__':
    main()
