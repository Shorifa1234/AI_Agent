import openpyxl

wb = openpyxl.load_workbook('SD_Web Scraping - Status Tracker.xlsx')
ws = wb['FDB Mobler']

categories = []

for i in range(1, ws.max_row + 1):
    row = list(ws[i])

    # Check if this row has "Category" in column C (index 2)
    if len(row) > 3 and row[2].value == 'Category' and row[3].value:
        cat_name = row[3].value

        # Next row should have "Link" in column C
        if i + 1 <= ws.max_row:
            link_row = list(ws[i + 1])
            if len(link_row) > 3 and link_row[2].value == 'Link' and link_row[3].value:
                cat_link = link_row[3].value
                categories.append({
                    'category': cat_name,
                    'link': cat_link
                })

print('FDB Mobler Categories:\n')
print('=' * 100)
for idx, c in enumerate(categories, 1):
    print(f'{idx:2}. {c["category"]:<40} -> {c["link"]}')

print('=' * 100)
print(f'\nTotal: {len(categories)} categories found')
