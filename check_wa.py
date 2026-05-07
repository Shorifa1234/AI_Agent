import openpyxl

wb = openpyxl.load_workbook(r"d:\Automation\SKU_AGENT\WorldsAway\Worlds Away.xlsx")

for sheet_name in wb.sheetnames[:3]:  # first 3 sheets
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"Row1: {ws.cell(1,1).value} | {ws.cell(1,2).value}")
    print(f"Row2: {ws.cell(2,1).value} | {ws.cell(2,2).value}")

    # Headers row 4
    headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1) if ws.cell(4, c).value]
    print(f"Headers ({len(headers)}): {headers}")

    # First data row
    row5 = {ws.cell(4, c).value: ws.cell(5, c).value for c in range(1, len(headers) + 1)}
    print(f"\nFirst product:")
    for k, v in row5.items():
        if v:
            print(f"  {k}: {v}")
