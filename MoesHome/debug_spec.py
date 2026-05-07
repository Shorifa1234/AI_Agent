import openpyxl

wb = openpyxl.load_workbook("MoesHome_demo.xlsx")
ws = wb["Lounge Chairs"]
headers = [cell.value for cell in ws[4]]
print("Last 5 headers:", headers[-5:])

# Check tearsheet value for first data row
row = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
tearsheet_col = headers.index("Tearsheet")
source_col = headers.index("Source")
print(f"Source:    {row[source_col]}")
print(f"Tearsheet: {row[tearsheet_col]}")
