"""
Moe's Home - Time Log Calculator
Code Development time + Code Implementation time.
"""

import openpyxl
from datetime import datetime

DEV_FIRST_CAT_MIN      = 90   # 1st category = 1 hr 30 min
DEV_OTHER_CAT_MIN_LOW  = 10   # remaining categories low
DEV_OTHER_CAT_MIN_HIGH = 15   # remaining categories high
SEC_PER_PRODUCT        = 0.5  # Code Implementation: per product URL


def fmt_min(minutes):
    h = int(minutes) // 60
    m = int(minutes) % 60
    return f"{h}h {m:02d}m" if h > 0 else f"{m}m"


def fmt_sec(seconds):
    if seconds >= 3600:
        h = int(seconds) // 3600
        m = (int(seconds) % 3600) // 60
        s = int(seconds) % 60
        return f"{h}h {m:02d}m {s:02d}s"
    elif seconds >= 60:
        m = int(seconds) // 60
        s = int(seconds) % 60
        return f"{m}m {s:02d}s"
    return f"{seconds:.1f}s"


def main():
    wb = openpyxl.load_workbook("MoesHome.xlsx")

    categories = []
    total_products = 0
    for sh in wb.sheetnames:
        ws = wb[sh]
        count = max(ws.max_row - 4, 0)
        categories.append((sh, count))
        total_products += count

    n_cats = len(categories)

    # ── Code Development Time ──────────────────────────────────────────────────
    dev_low  = DEV_FIRST_CAT_MIN + (n_cats - 1) * DEV_OTHER_CAT_MIN_LOW
    dev_high = DEV_FIRST_CAT_MIN + (n_cats - 1) * DEV_OTHER_CAT_MIN_HIGH

    # ── Code Implementation Time ───────────────────────────────────────────────
    impl_sec = total_products * SEC_PER_PRODUCT

    # ── Grand Total ───────────────────────────────────────────────────────────
    grand_low  = dev_low  * 60 + impl_sec
    grand_high = dev_high * 60 + impl_sec

    lines = []
    lines.append("=" * 62)
    lines.append("  MOE'S HOME — TIME LOG")
    lines.append(f"  Generated : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 62)

    # ── Section 1: Code Development ───────────────────────────────────────────
    lines.append("")
    lines.append("CODE DEVELOPMENT TIME")
    lines.append("  (1st category = 1h 30m  |  remaining = 10-15 min each)")
    lines.append("-" * 62)
    lines.append(f"  {'#':<4} {'Category':<34} {'Time':>10}")
    lines.append(f"  {'-'*4} {'-'*34} {'-'*10}")

    for i, (cat, _) in enumerate(categories):
        num = f"{i+1}."
        t   = "1h 30m" if i == 0 else "10 - 15 min"
        lines.append(f"  {num:<4} {cat:<34} {t:>10}")

    lines.append(f"  {'-'*4} {'-'*34} {'-'*10}")
    lines.append(f"  {'':4} {'TOTAL Code Development Time':<34} "
                 f"{(fmt_min(dev_low) + ' – ' + fmt_min(dev_high)):>10}")

    # ── Section 2: Code Implementation ────────────────────────────────────────
    lines.append("")
    lines.append("CODE IMPLEMENTATION TIME")
    lines.append("  (Product URLs × 0.5 sec)")
    lines.append("-" * 62)
    lines.append(f"  {'#':<4} {'Category':<26} {'URLs':>6}  {'× 0.5s':>6}  {'Time':>10}")
    lines.append(f"  {'-'*4} {'-'*26} {'-'*6}  {'-'*6}  {'-'*10}")

    for i, (cat, count) in enumerate(categories):
        cat_sec = count * SEC_PER_PRODUCT
        lines.append(f"  {i+1:<4} {cat:<26} {count:>6}  {'':>6}  {fmt_sec(cat_sec):>10}")

    lines.append(f"  {'-'*4} {'-'*26} {'-'*6}  {'-'*6}  {'-'*10}")
    lines.append(f"  {'':4} {'TOTAL Products':<26} {total_products:>6}  "
                 f"{'':>6}  {fmt_sec(impl_sec):>10}")

    # ── Grand Total ───────────────────────────────────────────────────────────
    lines.append("")
    lines.append("=" * 62)
    lines.append("  GRAND TOTAL  (Code Development + Code Implementation)")
    lines.append(f"  Low  estimate : {fmt_sec(grand_low)}")
    lines.append(f"  High estimate : {fmt_sec(grand_high)}")
    lines.append("=" * 62)

    log_text = "\n".join(lines)
    print(log_text)

    with open("moes_time_log.txt", "w", encoding="utf-8") as f:
        f.write(log_text)
    print("\nSaved: moes_time_log.txt")


if __name__ == "__main__":
    main()
