"""
Moe's Home - Step 2
Reads moes_step1.xlsx, fetches specification + warehouse data per SKU,
and writes final MoesHome.xlsx in Julian Chichester format.
"""

import sys
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

DEMO = "--demo" in sys.argv
GQL_URL = "https://mcprod.moeshome.com/graphql"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Content-Type": "application/json",
    "Store": "default",
}

SPEC_QUERY = """
query GetSpec($sku: String!) {
  specification(sku: $sku) {
    items { label value }
  }
}
"""

EXTRAS_QUERY = """
query GetExtras($sku: String!) {
  products(filter: {sku: {eq: $sku}}) {
    items {
      care_instructions
      assembly_requirements
    }
  }
}
"""

WAREHOUSE_QUERY = """
query GetWarehouse($sku: String!) {
  warehouses(sku: $sku) {
    warehouse_code
    in_stock_qty
    information {
      estimate {
        inbound { eta }
      }
    }
  }
}
"""

COLUMNS = [
    "Index", "Category", "Manufacturer", "Source",
    "Image URL", "Product Name", "SKU",
    "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter",
    "Weight", "Clearance to Floor",
    "Shipping Method", "Carton Dimensions",
    "Stock on Hand", "Next Arrival",
    "Care Instructions", "Hospitality Approved",
    "Country Of Origin", "Materials",
    "Seat Height", "Seat Depth", "Seat Width",
    "Arm Width", "Leg Height", "Arm Depth", "Arm Height", "Back Height",
    "Length Between Legs", "Width Between Legs",
    "Assembly Requirements", "Minimum Advertised Price (MAP) Policy",
    "Tearsheet",
]


def gql(query, variables):
    for attempt in range(3):
        try:
            resp = requests.post(GQL_URL, json={"query": query, "variables": variables},
                                 headers=HEADERS, timeout=45)
            data = resp.json()
            if "errors" in data:
                return None
            return data["data"]
        except Exception as e:
            if attempt < 2:  # Only print if not the last attempt
                print(f"  Request error (attempt {attempt+1}): {e}")
            time.sleep(2)
    return None


def get_spec(sku):
    data = gql(SPEC_QUERY, {"sku": sku})
    if not data:
        return {}
    items = data.get("specification", {}).get("items") or []
    return {i["label"]: i["value"] for i in items}


def get_extras(sku):
    data = gql(EXTRAS_QUERY, {"sku": sku})
    if not data:
        return {}
    items = (data.get("products") or {}).get("items") or []
    if not items:
        return {}
    return items[0]


def get_warehouse(sku):
    data = gql(WAREHOUSE_QUERY, {"sku": sku})
    if not data:
        return 0, ""
    warehouses = data.get("warehouses") or []
    total_stock = sum(w.get("in_stock_qty", 0) for w in warehouses)
    # Collect all ETAs and return the earliest (first in list, sorted by date string)
    etas = []
    for wh in warehouses:
        for est in (wh.get("information") or {}).get("estimate") or []:
            eta = (est.get("inbound") or {}).get("eta", "")
            if eta:
                etas.append(eta)
    next_arrival = etas[0] if etas else ""
    return total_stock, next_arrival


def parse_dimensions(dim_str):
    if not dim_str:
        return {}
    s = re.sub(r"[^\x00-\x7F]", " ", dim_str)
    result = {}
    dia = re.search(r"([\d.]+)\s*Dia", s, re.IGNORECASE)
    w = re.search(r"([\d.]+)\s*W\b", s, re.IGNORECASE)
    d = re.search(r"([\d.]+)\s*D\b", s, re.IGNORECASE)
    h = re.search(r"([\d.]+)\s*H\b", s, re.IGNORECASE)
    if dia:
        result["Diameter"] = dia.group(1)
    if w:
        result["Width"] = w.group(1)
    if d:
        result["Depth"] = d.group(1)
    if h:
        result["Height"] = h.group(1)
    return result


def format_weight(val):
    if not val:
        return ""
    try:
        f = float(val)
        return "" if f == 0.0 else f"{f:.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(val)


def style_sheet(ws, url):
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)

    ws["A1"] = "Brand"
    ws["B1"] = "Moe's Home"
    ws["A2"] = "Category Link"
    ws["B2"] = url

    for cell in [ws["A1"], ws["B1"], ws["A2"], ws["B2"]]:
        cell.font = Font(bold=True)

    ws.append([])  # row 3 blank

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")



def main():
    mode = "DEMO" if DEMO else "FULL"
    print(f"=== Moe's Home Step 2 [{mode}] ===", flush=True)

    suffix = "_demo" if DEMO else ""
    step1_path = f"moes_step1{suffix}.xlsx"
    out_path = f"MoesHome{suffix}.xlsx"

    # Auto-save interval (save every N sheets)
    autosave_interval = 5

    try:
        wb_in = openpyxl.load_workbook(step1_path)
    except FileNotFoundError:
        print(f"ERROR: {step1_path} not found. Run moes_step1.py first.")
        sys.exit(1)

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    total_sheets = len(wb_in.sheetnames)
    total_products = sum(len(list(wb_in[s].iter_rows(min_row=2, values_only=True))) for s in wb_in.sheetnames)
    processed_products = 0
    start_time = time.time()

    for sheet_idx, sheet_name in enumerate(wb_in.sheetnames, start=1):
        ws_in = wb_in[sheet_name]
        rows = list(ws_in.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        # step1 columns: Category(0) Code(1) Primary URL(2) SKU(3) Name(4)
        #   Product URL(5) Image URL(6) Description(7)
        #   Shipping Method(8) Box W(9) Box H(10) Box D(11) Box Wt(12)
        cat_name = rows[0][0] if rows[0][0] else sheet_name
        primary_url = rows[0][2] if rows[0][2] else ""

        print(f"\n[Sheet {sheet_idx}/{total_sheets}] [{sheet_name}] {len(rows)} products", flush=True)

        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        style_sheet(ws_out, primary_url)

        for row_idx, row in enumerate(rows, start=1):
            processed_products += 1
            sku = row[3] if len(row) > 3 else None
            name = row[4] if len(row) > 4 else ""
            prod_url = row[5] if len(row) > 5 else ""
            img_url = row[6] if len(row) > 6 else ""
            desc = row[7] if len(row) > 7 else ""
            ship_method = row[8] if len(row) > 8 else ""
            box_w = row[9] if len(row) > 9 else ""
            box_h = row[10] if len(row) > 10 else ""
            box_d = row[11] if len(row) > 11 else ""
            box_wt = row[12] if len(row) > 12 else ""

            if not sku:
                continue

            # Calculate progress
            percent = (processed_products / total_products) * 100
            elapsed = time.time() - start_time
            if processed_products > 10:  # Only estimate after a few products
                avg_time = elapsed / processed_products
                remaining = (total_products - processed_products) * avg_time
                eta_min = int(remaining / 60)
                print(f"  [{row_idx}/{len(rows)}] {sku} | Progress: {processed_products}/{total_products} ({percent:.1f}%) | ETA: ~{eta_min}min", flush=True)
            else:
                print(f"  [{row_idx}/{len(rows)}] {sku} - {name}", flush=True)

            spec = get_spec(sku)
            extras = get_extras(sku)
            stock, next_arrival = get_warehouse(sku)
            time.sleep(0.1)  # Reduced delay for faster processing

            dim_str = spec.get("General Dimensions", "")
            dims = parse_dimensions(dim_str)

            weight = format_weight(spec.get("Weight", ""))
            clearance = spec.get("Clearance to Floor", "")

            # Carton Dimensions: W x H x D (skip if all zero)
            carton_parts = [x for x in [box_w, box_h, box_d] if x]
            carton_dims = " x ".join(str(x) for x in carton_parts) if carton_parts else ""

            care         = extras.get("care_instructions", "")
            hospitality  = ""   # not available in public API
            country      = spec.get("Country Of Origin", "")
            materials    = spec.get("Product Materials", "")
            seat_height  = spec.get("Seat Height", "")
            seat_depth   = spec.get("Seat Depth", "")
            seat_width   = spec.get("Seat Width", "")
            arm_width    = spec.get("Arm Width", "")
            leg_height   = spec.get("Leg Height", "")
            arm_depth    = spec.get("Arm Depth", "")
            arm_height   = spec.get("Arm Height", "")
            back_height  = spec.get("Back Height", "")
            len_legs     = spec.get("Length Between Legs", "")
            wid_legs     = spec.get("Width Between Legs", "")
            assembly     = extras.get("assembly_requirements", "")
            map_policy   = ""   # not available in public API

            data_row = [
                row_idx,
                cat_name,
                "Moe's Home",
                prod_url,
                img_url,
                name,
                sku,
                name,
                desc,
                dims.get("Width", ""),
                dims.get("Depth", ""),
                dims.get("Height", ""),
                dims.get("Diameter", ""),
                weight,
                clearance,
                ship_method,
                carton_dims,
                stock if stock else "",
                next_arrival,
                care,
                hospitality,
                country,
                materials,
                seat_height,
                seat_depth,
                seat_width,
                arm_width,
                leg_height,
                arm_depth,
                arm_height,
                back_height,
                len_legs,
                wid_legs,
                assembly,
                map_policy,
                prod_url,
            ]
            ws_out.append(data_row)

        # Auto-width columns
        for col in ws_out.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    val = str(cell.value) if cell.value else ""
                    if len(val) > max_len:
                        max_len = len(val)
                except Exception:
                    pass
            ws_out.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # Auto-save every N sheets
        if sheet_idx % autosave_interval == 0:
            wb_out.save(out_path)
            print(f"  [Auto-saved after sheet {sheet_idx}]", flush=True)

    wb_out.save(out_path)
    total_time = time.time() - start_time
    print(f"\n{'='*60}")
    print(f"✓ Completed! Processed {total_products} products in {int(total_time/60)}m {int(total_time%60)}s")
    print(f"✓ Saved: {out_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
