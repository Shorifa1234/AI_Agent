# crylights_step1.py
# crystoramalightinglights.com -- Product list collector (Step 1)
# Paginates each category URL, collects product links, names, prices, images.
# Output: crylights_step1.xlsx
#
# deps: pip install requests openpyxl beautifulsoup4

from __future__ import annotations
import os
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# ------------------------------------------------------------------------------
# CONFIG
# ------------------------------------------------------------------------------
BASE_URL    = "https://crystoramalightinglights.com"
OUTPUT_FILE = "crylights_step1.xlsx"
PAGE_SIZE    = 24    # default items per page
DELAY_SEC    = 1.5   # polite delay between page requests
CAT_DELAY    = 30    # seconds to wait between categories
RETRY_WAIT   = 90    # seconds to wait on 503 before retrying
MAX_RETRIES  = 3     # max retries per request

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# (Category name, [list of category URLs to combine])
CATEGORIES = [
    ("Lighting", [
        "https://crystoramalightinglights.com/cry/category/lighting/outdoor-lighting/?showproducts=true",
    ]),
    ("Chandeliers", [
        "https://crystoramalightinglights.com/cry/category/lighting/ceiling-lights/chandeliers/",
        "https://crystoramalightinglights.com/cry/category/lighting/ceiling-lights/mini-chandeliers/",
    ]),
    ("Pendants", [
        "https://crystoramalightinglights.com/cry/category/lighting/ceiling-lights/pendants/",
        "https://crystoramalightinglights.com/cry/category/lighting/ceiling-lights/mini-pendants/",
    ]),
    ("Sconces", [
        "https://crystoramalightinglights.com/cry/category/lighting/wall-lights/wall-sconces/",
        "https://crystoramalightinglights.com/cry/category/lighting/wall-lights/bathroom-vanity-lights/",
        "https://crystoramalightinglights.com/cry/category/lighting/wall-lights/swing-arm-%2F-wall-lamps/",
    ]),
    ("Flush Mount", [
        "https://crystoramalightinglights.com/cry/category/lighting/ceiling-lights/flush-mounts/",
        "https://crystoramalightinglights.com/cry/category/lighting/ceiling-lights/semi-flush-mounts/",
    ]),
    ("Mirrors", [
        "https://crystoramalightinglights.com/cry/category/furniture-and-decor/?showproducts=true",
    ]),
]
# ------------------------------------------------------------------------------


def _get_with_retry(url: str) -> requests.Response | None:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=45)
            if resp.status_code == 503:
                print(f"    [503] Rate limited. Waiting {RETRY_WAIT}s before retry {attempt}/{MAX_RETRIES}...")
                time.sleep(RETRY_WAIT)
                continue
            resp.raise_for_status()
            return resp
        except Exception as e:
            if attempt == MAX_RETRIES:
                print(f"    [WARN] {url}: {e}")
                return None
            print(f"    [ERR] Attempt {attempt} failed: {e}. Retrying in 30s...")
            time.sleep(30)
    return None


def _parse_cards(soup: BeautifulSoup) -> list[dict]:
    products = []
    for card in soup.select(".product-tile"):
        link = card.select_one("a.link[href*='/product/']")
        if not link:
            continue
        href = link.get("href", "")
        product_url = href if href.startswith("http") else BASE_URL + href
        name = link.get_text(strip=True)

        price_el = card.select_one("span.value")
        price = ""
        if price_el:
            price = price_el.get("content", "") or re.sub(r"[^\d.]", "", price_el.get_text(strip=True))

        img = card.select_one("img[itemprop='image']")
        image_url = ""
        if img:
            src = img.get("src", "")
            image_url = src.split("?")[0] if src else ""

        products.append({
            "Product Name": name,
            "Product URL":  product_url,
            "Image URL":    image_url,
            "Price":        price,
        })
    return products


def fetch_category(category_name: str, urls: list) -> list:
    seen = set()
    results = []

    for base_url in urls:
        start = 0
        while True:
            sep      = "&" if "?" in base_url else "?"
            full_url = f"{base_url}{sep}start={start}" if start > 0 else base_url
            resp = _get_with_retry(full_url)
            if resp is None:
                break

            soup  = BeautifulSoup(resp.text, "html.parser")
            cards = _parse_cards(soup)
            if not cards:
                break

            new = 0
            for p in cards:
                if p["Product URL"] not in seen:
                    seen.add(p["Product URL"])
                    p["Category"] = category_name
                    results.append(p)
                    new += 1

            print(f"    page start={start}: +{new} new  (total {len(results)})")

            if len(cards) < PAGE_SIZE:
                break
            start += PAGE_SIZE
            time.sleep(DELAY_SEC)

    return results


def load_existing(filepath: str) -> tuple[list, set]:
    # Returns (existing_rows, done_categories)
    if not os.path.exists(filepath):
        return [], set()
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        cats = set()
        for r in range(2, ws.max_row + 1):
            row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
            if any(v for v in row.values()):
                rows.append(row)
                if row.get("Category"):
                    cats.add(row["Category"])
        wb.close()
        return rows, cats
    except Exception:
        return [], set()


def save_to_excel(all_rows: list) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Products"

    if not all_rows:
        wb.save(OUTPUT_FILE)
        print("No products found.")
        return

    bold = Font(bold=True)
    cols = ["Category", "Product Name", "Product URL", "Image URL", "Price"]

    for c, h in enumerate(cols, 1):
        ws.cell(1, c, h).font = bold

    for r, row in enumerate(all_rows, 2):
        for c, key in enumerate(cols, 1):
            ws.cell(r, c, row.get(key, ""))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 80)

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"\nSaved {len(all_rows)} products to {OUTPUT_FILE}")


def main() -> None:
    print("=" * 60)
    print("CrystoramaLights Step 1 -- Product List Collection")
    print("=" * 60)

    existing_rows, done_cats = load_existing(OUTPUT_FILE)
    if done_cats:
        print(f"Resume: {len(done_cats)} categories already done: {', '.join(sorted(done_cats))}")

    all_rows = list(existing_rows)
    for category_name, urls in CATEGORIES:
        if category_name in done_cats:
            print(f"\n[{category_name}] -- already collected, skipping")
            continue
        print(f"\n[{category_name}] ({len(urls)} URL(s))")
        rows = fetch_category(category_name, urls)
        all_rows.extend(rows)
        print(f"  -> {len(rows)} products")
        print(f"  Waiting {CAT_DELAY}s before next category...")
        time.sleep(CAT_DELAY)

    print(f"\nTotal: {len(all_rows)} products across {len(CATEGORIES)} categories")
    save_to_excel(all_rows)


if __name__ == "__main__":
    main()
