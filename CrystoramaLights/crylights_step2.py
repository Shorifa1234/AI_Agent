# crylights_step2.py
# crystoramalightinglights.com -- Detail page scraper + final Excel (Step 2)
# Reads crylights_step1.xlsx, visits each product URL, extracts all fields,
# saves CrystoramaLights.xlsx in Julian Chichester format (one sheet per category).
#
# deps: pip install requests openpyxl beautifulsoup4

from __future__ import annotations
import json
import os
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# ------------------------------------------------------------------------------
# CONFIG
# ------------------------------------------------------------------------------
INPUT_FILE   = "crylights_step1.xlsx"
OUTPUT_FILE  = "CrystoramaLights.xlsx"
VENDOR_NAME  = "Crystorama"

DELAY_SEC    = 0.5
SAVE_EVERY_N = 50

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

OUTPUT_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Base SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Length", "Weight",
    "Extension", "Canopy", "Maximum Adjustable Height",
    "Outside Length", "Outside Depth", "Outside Height",
    "Inside Length", "Inside Depth", "Inside Height", "Seat Height", "Arm Height",
    "Finish", "Finish Sample Code", "Color", "Collection",
    "Materials", "Material", "Origin", "Country of Origin",
    "Body Fabric", "Welt Fabric",
    "Price", "List Price", "Availability", "Shipping", "Shipping Method",
    "Style", "Product Type", "Features", "Tags",
    "Total Bulbs", "Wattage", "Dimmable", "Lamping Type",
    "Socket", "Voltage", "Shape", "Glass Features",
    "Install Position", "UL Ratings", "Prop 65", "Title 20", "Warranty", "UPC",
    "All SKUs",
]
# ------------------------------------------------------------------------------


_SPEC_KEY_MAP = {
    "Shipping Method":            "Shipping Method",
    "UPC":                        "UPC",
    "Lamping Type":               "Lamping Type",
    "Lamping Features":           "Lamping Type",
    "Socket":                     "Socket",
    "Glass Features":             "Glass Features",
    "Material":                   "Material",
    "Shape":                      "Shape",
    "Extension":                  "Extension",
    "Height":                     "Height",
    "Width":                      "Width",
    "Length":                     "Length",
    "Depth":                      "Depth",
    "Diameter":                   "Diameter",
    "Maximum Adjustable Height":  "Maximum Adjustable Height",
    "Weight":                     "Weight",
    "Country of Origin":          "Country of Origin",
    "Install Position":           "Install Position",
    "Prop 65":                    "Prop 65",
    "Title 20":                   "Title 20",
    "UL Ratings":                 "UL Ratings",
    "Warranty":                   "Warranty",
    "Voltage":                    "Voltage",
    "Number of Lights":           "Total Bulbs",
    "Wattage":                    "Wattage",
    "Dimmable":                   "Dimmable",
    "Finish":                     "Finish",
    "Collection":                 "Collection",
    "Style":                      "Style",
    "Canopy Width":               "Canopy",
    "Backplate/Canopy Width":     "Canopy",
    "Backplate/Canopy Extension": "Extension",
}


def _parse_spec_pairs(soup: BeautifulSoup) -> dict:
    """Parse each .spec-item element into {label: value}."""
    pairs = {}
    for si in soup.select(".spec-item"):
        label_el = si.select_one(".spec-label, .label, dt")
        value_el = si.select_one(".spec-value, .value, dd")
        if label_el and value_el:
            key = label_el.get_text(strip=True).rstrip(":")
            val = value_el.get_text(" ", strip=True)
        else:
            text = si.get_text(" ", strip=True)
            m = re.match(r"^([^:]+?):\s*(.+)$", text, re.S)
            if not m:
                continue
            key, val = m.group(1).strip(), m.group(2).strip()
        if key and val:
            pairs[key] = val
    return pairs


def _base_sku(sku: str) -> str:
    parts = sku.rsplit("-", 1)
    if len(parts) == 2 and re.fullmatch(r"[A-Z0-9]{1,4}", parts[1]):
        return parts[0]
    return sku


def extract_detail(html: str) -> dict:
    data = {}
    soup = BeautifulSoup(html, "html.parser")

    # --- 1. JSON-LD (SKU, name, description, image, price, availability) ---
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            ld = json.loads(script.string or "")
            if ld.get("@type") == "Product":
                data["SKU"]          = ld.get("sku") or ld.get("mpn", "")
                data["Product Name"] = ld.get("name", "")
                data["Description"]  = ld.get("description", "")
                imgs = ld.get("image", [])
                if isinstance(imgs, list) and imgs:
                    data["Image URL"] = imgs[0].split("?")[0]
                elif isinstance(imgs, str) and imgs:
                    data["Image URL"] = imgs.split("?")[0]
                offers = ld.get("offers", {})
                if offers:
                    data["Price"] = str(offers.get("price", ""))
                    avail = offers.get("availability", "")
                    if "OutOfStock" in avail:
                        data["Availability"] = "Out of Stock"
                    elif "InStock" in avail:
                        data["Availability"] = "In Stock"
                data.setdefault("Product Type", ld.get("category", ""))
                break
        except Exception:
            pass

    # --- 2. og:image fallback ---
    if not data.get("Image URL"):
        og_img = soup.select_one('meta[property="og:image"]')
        if og_img:
            data["Image URL"] = og_img.get("content", "").split("?")[0]

    # --- 3. SKU fallback ---
    if not data.get("SKU"):
        sku_el = soup.select_one("[class*=sku]")
        if sku_el:
            m = re.search(r"SKU:\s*(\S+)", sku_el.get_text(strip=True))
            if m:
                data["SKU"] = m.group(1)

    # --- 4. Description fallback ---
    if not data.get("Description"):
        desc_el = soup.select_one(".product-overview")
        if desc_el:
            for h in desc_el.find_all("h3"):
                h.decompose()
            data["Description"] = desc_el.get_text(" ", strip=True)

    # --- 5. Collection / Color / Finish / Style from .product-overview-attrs ---
    attrs_div = soup.select_one(".product-overview-attrs")
    if attrs_div:
        for col_div in attrs_div.select("div"):
            h3 = col_div.find("h3")
            a  = col_div.find("a")
            if h3 and a:
                label = h3.get_text(strip=True)
                value = a.get_text(strip=True)
                label_map = {
                    "Collection": "Collection", "Finish": "Finish",
                    "Color": "Color", "Style": "Style", "Category": "Product Type",
                }
                if label in label_map:
                    data.setdefault(label_map[label], value)

    # --- 6. All spec-item key:value pairs ---
    spec_pairs = _parse_spec_pairs(soup)
    for page_key, out_col in _SPEC_KEY_MAP.items():
        if out_col not in data and page_key in spec_pairs:
            data[out_col] = spec_pairs[page_key]

    # "Lamping Features: 3 light" -> Total Bulbs
    if "Total Bulbs" not in data:
        for key, val in spec_pairs.items():
            if re.search(r"lamping\s+features", key, re.I):
                m = re.search(r"(\d+)\s+light", val, re.I)
                if m:
                    data["Total Bulbs"] = m.group(1)
                break

    # Drop meaningless "0 light" lamping entries (mirrors etc.)
    for f in ("Lamping Type", "Total Bulbs"):
        val = str(data.get(f, ""))
        if re.match(r"^0\b", val.strip()):
            data.pop(f, None)

    # Ensure Wattage has "W" suffix
    if data.get("Wattage"):
        raw = str(data["Wattage"])
        if not raw.upper().endswith("W"):
            data["Wattage"] = re.sub(r"[^\d.]", "", raw) + "W"

    # --- 7. Parse "Dimensions: 7.5"W x 22.25"H x 7.5"D" as fallback ---
    dims_raw = spec_pairs.get("Dimensions", "")
    if dims_raw:
        m = re.search(
            r'([\d.]+)["\s]*W\s*[xX×]\s*([\d.]+)["\s]*H(?:\s*[xX×]\s*([\d.]+)["\s]*D)?',
            dims_raw, re.I,
        )
        if m:
            data.setdefault("Width",  m.group(1))
            data.setdefault("Height", m.group(2))
            if m.group(3):
                data.setdefault("Depth", m.group(3))

    # --- 8. Fallback: regex on concatenated spec text ---
    spec_text = " ".join(si.get_text(" ", strip=True) for si in soup.select(".spec-item"))
    if not spec_text:
        specs_el  = soup.select_one(".product-specifications")
        spec_text = specs_el.get_text(" ", strip=True) if specs_el else ""

    if not data.get("Finish"):
        m = re.search(r"Finish\s*:\s*([A-Za-z /&,\-]+?)(?:\s+[A-Z][a-z]+\s*:|$)", spec_text, re.I)
        if m:
            data["Finish"] = m.group(1).strip().rstrip(",")

    if not data.get("Total Bulbs"):
        m = re.search(r"Lamping\s+Features\s*:\s*(\d+)\s+light", spec_text, re.I)
        if m:
            data["Total Bulbs"] = m.group(1)

    if not data.get("Wattage"):
        m = re.search(r"(\d+)[-\s]*watt", spec_text, re.I)
        if m:
            data["Wattage"] = m.group(1) + "W"

    if not data.get("Dimmable"):
        m = re.search(r"Dimmable\s*:\s*(Yes|No)", spec_text, re.I)
        if m:
            data["Dimmable"] = m.group(1)

    # --- 9. Price fallback from DOM ---
    if not data.get("Price"):
        price_el = soup.select_one(".price .value")
        if price_el:
            data["Price"] = price_el.get("content", "") or re.sub(r"[^\d.]", "", price_el.get_text(strip=True))

    return data


def fetch_detail(url: str) -> dict:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=45)
        resp.raise_for_status()
        return extract_detail(resp.text)
    except Exception as e:
        print(f"    [WARN] {url}: {e}")
        return {}


def load_step1(filepath: str) -> list:
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"'{filepath}' not found -- run crylights_step1.py first.")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        if any(v for v in row.values()):
            rows.append(row)
    wb.close()
    return rows


def _generate_sku(category: str, index: int) -> str:
    words = re.sub(r"[^A-Za-z\s]", "", category).strip().split()
    if len(words) >= 2:
        c = (words[0][0] + words[1][0]).upper()
    elif words:
        c = words[0][:2].upper()
    else:
        c = "XX"
    return f"CRY{c}{index:02d}"


def _write_excel(enriched_by_cat: dict, out_path: str) -> None:
    wb = openpyxl.Workbook()
    for name in list(wb.sheetnames):
        del wb[name]
    bold = Font(bold=True)

    for category, products in enriched_by_cat.items():
        ws = wb.create_sheet(category[:31])

        ws.cell(1, 1, "Brand").font = bold
        ws.cell(1, 2, VENDOR_NAME)

        cat_url = products[0].get("Category URL", "") if products else ""
        ws.cell(2, 1, "Category Link").font = bold
        ws.cell(2, 2, cat_url)

        extra_cols = []
        for p in products:
            for k in p:
                if k not in OUTPUT_COLS and k not in extra_cols:
                    extra_cols.append(k)
        all_cols = OUTPUT_COLS + extra_cols

        for col, header in enumerate(all_cols, 1):
            ws.cell(4, col, header).font = bold

        for idx, prod in enumerate(products, 1):
            prod["Index"]            = idx
            prod["Manufacturer"]     = VENDOR_NAME
            prod["Category"]         = category
            prod["Source"]           = prod.get("Product URL", "")
            sku = prod.get("SKU", "")
            if not sku:
                prod["SKU"] = _generate_sku(category, idx)
            else:
                prod.setdefault("Base SKU", _base_sku(sku))
            prod.setdefault("Product Family Id", prod.get("Product Name", ""))

            for col, key in enumerate(all_cols, 1):
                ws.cell(4 + idx, col, prod.get(key, ""))

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 55)

        ws.freeze_panes = "A5"

    wb.save(out_path)


def save_progress(enriched_by_cat: dict) -> None:
    tmp = OUTPUT_FILE.replace(".xlsx", "_tmp.xlsx")
    _write_excel(enriched_by_cat, tmp)
    try:
        if os.path.exists(OUTPUT_FILE):
            os.remove(OUTPUT_FILE)
        os.rename(tmp, OUTPUT_FILE)
    except PermissionError:
        alt = OUTPUT_FILE.replace(".xlsx", "_saved.xlsx")
        if os.path.exists(alt):
            os.remove(alt)
        os.rename(tmp, alt)
        print(f"  [WARN] '{OUTPUT_FILE}' open in Excel -- saved to '{alt}'")


def load_done_urls(out_path: str) -> set:
    done = set()
    if not os.path.exists(out_path):
        return done
    try:
        wb = openpyxl.load_workbook(out_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            src_col = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(4, c).value == "Source":
                    src_col = c
                    break
            if src_col is None:
                continue
            for r in range(5, ws.max_row + 1):
                v = ws.cell(r, src_col).value
                if v:
                    done.add(str(v).strip())
        wb.close()
    except Exception:
        pass
    return done


def load_existing_output(out_path: str, categories: list) -> dict:
    """Load previously scraped products from output Excel into enriched_by_cat."""
    result = {cat: [] for cat in categories}
    if not os.path.exists(out_path):
        return result
    try:
        wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
        cat_by_sheet = {cat[:31]: cat for cat in categories}
        for sheet_name in wb.sheetnames:
            if sheet_name not in cat_by_sheet:
                continue
            cat = cat_by_sheet[sheet_name]
            ws = wb[sheet_name]
            all_rows = list(ws.rows)
            if len(all_rows) < 5:
                continue
            headers = [cell.value for cell in all_rows[3]]  # row 4
            for row_cells in all_rows[4:]:                  # row 5+
                row = {h: cell.value for h, cell in zip(headers, row_cells) if h}
                if any(v for v in row.values()):
                    if row.get("Source") and not row.get("Product URL"):
                        row["Product URL"] = row["Source"]
                    result[cat].append(row)
        wb.close()
    except Exception as e:
        print(f"  [WARN] load_existing_output: {e}")
    return result


def main() -> None:
    print("=" * 60)
    print("CrystoramaLights Step 2 -- Detail Scraper")
    print(f"Input : {INPUT_FILE}")
    print(f"Output: {OUTPUT_FILE}")
    print("=" * 60)

    rows = load_step1(INPUT_FILE)
    print(f"Loaded {len(rows)} products from {INPUT_FILE}")

    done_urls = load_done_urls(OUTPUT_FILE)
    if done_urls:
        print(f"Resume: {len(done_urls)} already done -- skipping")

    categories_order = []
    by_category = {}
    for row in rows:
        cat = row.get("Category", "Unknown")
        if cat not in by_category:
            by_category[cat] = []
            categories_order.append(cat)
        by_category[cat].append(row)

    print(f"Categories: {len(categories_order)}")

    enriched_by_cat = load_existing_output(OUTPUT_FILE, categories_order)
    existing_count = sum(len(v) for v in enriched_by_cat.values())
    if existing_count:
        print(f"Loaded {existing_count} previously scraped products from {OUTPUT_FILE}")
    processed = 0
    skipped   = 0
    total     = len(rows)

    for cat in categories_order:
        print(f"\n[{cat}] -- {len(by_category[cat])} products")

        for row in by_category[cat]:
            url  = row.get("Product URL", "")
            name = row.get("Product Name", url)

            if url in done_urls:
                skipped += 1
                continue

            processed += 1
            print(f"  [{processed}/{total - skipped}] {name[:65]}")

            detail = fetch_detail(url) if url else {}
            row.update(detail)
            enriched_by_cat[cat].append(row)

            if processed % SAVE_EVERY_N == 0:
                save_progress(enriched_by_cat)
                print(f"  [auto-save] {processed} written to {OUTPUT_FILE}")

            time.sleep(DELAY_SEC)

    save_progress(enriched_by_cat)

    total_saved = sum(len(v) for v in enriched_by_cat.values())
    print(f"\n{'=' * 60}")
    print(f"Done! {total_saved} products saved to {OUTPUT_FILE}")
    print(f"Sheets: {', '.join(categories_order)}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
