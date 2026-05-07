"""
Lee Industries -- Step 2
========================
Reads Lee_step1.xlsx (one sheet per category).
Visits each product/fabric detail page with requests + BeautifulSoup.
Saves to Lee/Lee.xlsx (or Lee_demo.xlsx) in Julian Chichester format.

Usage:
    python lee_step2.py
    python lee_step2.py --demo      # 5 products per category
"""

import re
import sys
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

SCRIPT_DIR = Path(__file__).parent
STEP1_FILE = SCRIPT_DIR / "Lee_step1.xlsx"
VENDOR     = "Lee Industries"
BASE_URL   = "https://www.leeindustries.com"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
}
SLEEP = 0.5

FIXED_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Finish",
]

# Extra columns per category type
SEATING_EXTRA = [
    "Weight", "Seat Height", "Arm Height", "Seat Depth",
    "Inside Width", "Inside Depth", "Inside Height",
    "COM Available", "COM", "COL", "COT", "Cushion", "Tearsheet",
]
BEDS_EXTRA    = ["Weight", "Inside Width", "Inside Depth", "Inside Height", "Color", "Length", "Tearsheet"]
FABRIC_EXTRA  = ["Grade", "Content", "Cut Direction", "Repeat", "Cleaning Code", "Rub Count", "Rub Notes", "Tearsheet"]
LEATHER_EXTRA = ["Grade", "Content", "Pattern", "Construction", "Repeat", "CFA", "Cleaning Code", "Rub Count", "Rub Notes", "Tearsheet"]


# ── SKU generation ──────────────────────────────────────────────────────────────

CATEGORY_CODES = {
    "Fabric":            "FA",
    "Beds & Headboards": "BH",
    "Dining Chairs":     "DI",
    "Bar Stools":        "BS",
    "Sofas & Loveseats": "SL",
    "Sectionals":        "SE",
    "Lounge Chairs":     "LC",
    "Ottomans":          "OT",
    "Desk Chairs":       "DE",
    "Pillows & Throws":  "PT",
    "Leather":           "LE",
}


def generate_sku(category: str, index: int) -> str:
    code = CATEGORY_CODES.get(category, category[:2].upper())
    return f"LEE{code}{index:02d}"


def is_fabric_url(url: str) -> bool:
    return "/Upholstery/" in url or "/Options/" in url


# ── Detail page parsers ─────────────────────────────────────────────────────────

def parse_specs(soup) -> dict:
    """Extract all .column-key / .column-value pairs."""
    specs = {}
    for row_div in soup.find_all("div", class_="row"):
        keys = row_div.find_all("div", class_="column-key")
        vals = row_div.find_all("div", class_="column-value")
        for k, v in zip(keys, vals):
            key = k.get_text(strip=True).rstrip(":")
            val = v.get_text(strip=True)
            if key:
                specs[key] = val
    return specs


def parse_overall_dims(overall: str) -> tuple[str, str, str]:
    """Parse 'W65 D33 H35' -> (width, depth, height). Returns '' for zero or missing."""
    w = re.search(r"W([\d.]+)", overall, re.I)
    d = re.search(r"D([\d.]+)", overall, re.I)
    h = re.search(r"H([\d.]+)", overall, re.I)
    def nonzero(m):
        return m.group(1) if m and float(m.group(1)) != 0 else ""
    return nonzero(w), nonzero(d), nonzero(h)


def scrape_product_detail(url: str, fallback_name: str, fallback_image: str) -> dict:
    """Scrape a /Product/Detail/ page."""
    data: dict = {}
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        data["_error"] = str(e)
        return data

    soup = BeautifulSoup(r.text, "html.parser")

    # Product name: h2.hide-for-small preferred, fallback to h1
    h2 = soup.find("h2", class_="hide-for-small")
    h1 = soup.find("h1")
    name = (h2 and h2.get_text(strip=True)) or (h1 and h1.get_text(strip=True)) or fallback_name
    data["Product Name"] = name

    # High-res image from a#image-download, then any /hires/ link, then thumbnail
    hires = soup.find("a", id="image-download")
    if hires and hires.get("href"):
        data["Image URL"] = hires["href"]
    else:
        hires_link = soup.find("a", href=lambda h: h and "/azure/product/hires/" in h)
        if hires_link:
            data["Image URL"] = hires_link["href"]
        else:
            for img in soup.find_all("img"):
                src = img.get("src") or img.get("data-src") or ""
                if "/azure/product/thumbnail/" in src:
                    data["Image URL"] = src.split("?")[0]
                    break
    if not data.get("Image URL"):
        data["Image URL"] = fallback_image

    # Specs
    specs = parse_specs(soup)

    overall = specs.get("OVERALL", "")
    if overall:
        w, d, h = parse_overall_dims(overall)
        data["Width"]  = w
        data["Depth"]  = d
        data["Height"] = h

    def nz(val):
        return val if val and val.strip() not in ("0", "0.0") else ""

    data["Seat Height"] = nz(specs.get("SEAT HEIGHT", ""))
    data["Arm Height"]  = nz(specs.get("ARM HEIGHT", ""))
    data["Weight"]      = nz(specs.get("WEIGHT", ""))

    # Inside dimensions from INSIDE: W19 D18.50 H22.50
    inside = specs.get("INSIDE", "")
    if inside:
        iw = re.search(r"W([\d.]+)", inside, re.I)
        id_ = re.search(r"D([\d.]+)", inside, re.I)
        ih = re.search(r"H([\d.]+)", inside, re.I)
        def nz2(m):
            return m.group(1) if m and float(m.group(1)) != 0 else ""
        data["Inside Width"]  = nz2(iw)
        data["Inside Depth"]  = nz2(id_)
        data["Inside Height"] = nz2(ih)
        data["Seat Depth"]    = nz2(id_)

    # COM from spec keys (number only, no unit)
    com_yds = specs.get("COM PLAIN YARDAGE", "") or specs.get("COM YARDAGE", "")
    if com_yds:
        data["COM"] = com_yds

    # Tearsheet URL
    tearsheet = soup.find("a", id="btnPrintTearsheet")
    if tearsheet and tearsheet.get("href"):
        href = tearsheet["href"]
        data["Tearsheet"] = BASE_URL + href if href.startswith("/") else href

    # Finish: shown fabric name
    shown_fabric = soup.find("div", id="shownFabricName")
    if not shown_fabric:
        shown_fabric = soup.find("div", class_="showin-text")
    if shown_fabric:
        data["Finish"] = shown_fabric.get_text(strip=True)

    # Description + Cushion: from STANDARD WITH section
    for detail_div in soup.find_all("div", class_="detail-info"):
        h4 = detail_div.find("h4")
        if h4 and "STANDARD" in h4.get_text(strip=True).upper():
            product_info = detail_div.find("div", class_="product-detail-info")
            if product_info:
                text = product_info.get_text(separator="; ", strip=True)
            else:
                # Pillows: text directly in detail-info (no product-detail-info wrapper)
                h4.extract()
                text = detail_div.get_text(separator="; ", strip=True)
            if text:
                data["Description"] = text
                m = re.search(r"Cushion\s*[-–]\s*(.+?)(?:;|$)", text, re.I)
                if m:
                    data["Cushion"] = m.group(1).strip()
            break

    return data


def scrape_fabric_detail(url: str, fallback_name: str, fallback_image: str) -> dict:
    """Scrape a /Upholstery/Detail/ page (fabric or leather)."""
    data: dict = {}
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        data["_error"] = str(e)
        return data

    soup = BeautifulSoup(r.text, "html.parser")
    body = r.text

    # Product name: h2.hide-for-small (fabric/leather/pillows pages)
    h2 = soup.find("h2", class_="hide-for-small")
    name = h2.get_text(strip=True) if h2 and h2.get_text(strip=True) else ""
    if not name:
        title_tag = soup.find("title")
        if title_tag:
            title_text = title_tag.get_text(strip=True)
            name = title_text.split("|")[0].strip()
    data["Product Name"] = name or fallback_name

    # High-res image link (most reliable for fabric/leather)
    hires_link = soup.find("a", href=lambda h: h and "/hires/" in h)
    if hires_link:
        data["Image URL"] = hires_link["href"]
    else:
        # Fall back to thumbnail in page, then listing image
        for img in soup.find_all("img"):
            src = img.get("src") or img.get("data-src") or ""
            if "/azure/fabric-leather/thumbnail/" in src:
                data["Image URL"] = src.split("?")[0]
                break
        if not data.get("Image URL"):
            # Pillows: product thumbnail
            for img in soup.find_all("img"):
                src = img.get("src") or img.get("data-src") or ""
                if "/azure/product/thumbnail/" in src:
                    data["Image URL"] = src.split("?")[0]
                    break
    if not data.get("Image URL"):
        data["Image URL"] = fallback_image

    # Specs
    specs = parse_specs(soup)

    data["Grade"]         = specs.get("GRADE", "")
    data["Content"]       = specs.get("CONTENT", "")
    data["Cut Direction"] = specs.get("CUT DIRECTION", "")
    data["Cleaning Code"] = specs.get("CLEANING CODE", "")
    data["Repeat"]        = specs.get("REPEAT", "")
    data["Rub Count"]     = specs.get("RUB COUNT", "")
    data["Rub Notes"]     = specs.get("RUB NOTES", "")
    data["Pattern"]       = specs.get("PATTERN", "")
    data["Construction"]  = specs.get("CONSTRUCTION", "")
    data["CFA"]           = specs.get("CFA", "")

    # Description: NOTES from specs, or from STANDARD WITH / NOTES detail-info sections
    notes = specs.get("NOTES", "")
    if notes:
        data["Description"] = notes
    else:
        parts = []
        for detail_div in soup.find_all("div", class_="detail-info"):
            h4 = detail_div.find("h4")
            label = h4.get_text(strip=True).upper() if h4 else ""
            if label in ("STANDARD WITH", "NOTES"):
                product_info = detail_div.find("div", class_="product-detail-info")
                if product_info:
                    parts.append(product_info.get_text(separator="; ", strip=True))
                else:
                    if h4: h4.extract()
                    text = detail_div.get_text(separator="; ", strip=True)
                    if text:
                        parts.append(text)
        if parts:
            data["Description"] = " | ".join(parts)

    # Tearsheet URL
    tearsheet = soup.find("a", id="btnPrintTearsheet")
    if tearsheet and tearsheet.get("href"):
        href = tearsheet["href"]
        data["Tearsheet"] = BASE_URL + href if href.startswith("/") else href

    return data


# ── Excel helpers ───────────────────────────────────────────────────────────────

def get_extra_cols(category: str) -> list[str]:
    cat_lower = category.lower()
    if cat_lower == "leather":
        return LEATHER_EXTRA
    if cat_lower in ("fabric",):
        return FABRIC_EXTRA
    if cat_lower in ("beds & headboards",):
        return BEDS_EXTRA
    # All seating + accessories
    return SEATING_EXTRA


def save_sheet(wb, category: str, category_url: str, rows: list[dict]) -> None:
    sheet_name = category[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    bold = Font(bold=True)

    ws.cell(1, 1, "Brand").font          = bold
    ws.cell(1, 2, VENDOR)
    ws.cell(2, 1, "Category Link").font  = bold
    ws.cell(2, 2, category_url)

    extra_cols = get_extra_cols(category)
    all_cols = FIXED_COLS + [c for c in extra_cols if c not in FIXED_COLS]

    for col, h in enumerate(all_cols, 1):
        ws.cell(4, col, h).font = bold

    for i, row in enumerate(rows, 1):
        row.setdefault("Index",             i)
        row.setdefault("Category",          category)
        row.setdefault("Manufacturer",      VENDOR)
        row.setdefault("Product Family Id", row.get("Product Name", ""))
        row.setdefault("SKU",               generate_sku(category, i))

        for col, key in enumerate(all_cols, 1):
            ws.cell(4 + i, col, row.get(key, ""))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A5"


# ── Main ───────────────────────────────────────────────────────────────────────

def main(demo: bool = False) -> None:
    if not STEP1_FILE.exists():
        print(f"ERROR: {STEP1_FILE} not found. Run lee_step1.py first.")
        return

    demo_limit = 5 if demo else None
    out_path   = SCRIPT_DIR / ("Lee_demo.xlsx" if demo else "Lee Industries.xlsx")

    if demo:
        print("[DEMO MODE] 5 products per category -> Lee_demo.xlsx\n")

    wb_in = openpyxl.load_workbook(str(STEP1_FILE))

    try:
        wb_out = openpyxl.load_workbook(str(out_path))
    except FileNotFoundError:
        wb_out = openpyxl.Workbook()
        for s in list(wb_out.sheetnames):
            del wb_out[s]

    # Step1 columns: #, Category, Product Name, SKU, Product URL, Image URL, Category URL
    for sheet_name in wb_in.sheetnames:
        ws_in    = wb_in[sheet_name]
        category = sheet_name

        products_in: list[dict] = []
        category_url = ""

        for row in ws_in.iter_rows(min_row=2, values_only=True):
            url = str(row[4] or "").strip()
            if not url:
                continue
            if not category_url:
                category_url = str(row[6] or "")
            products_in.append({
                "name":  str(row[2] or ""),
                "sku":   str(row[3] or ""),
                "url":   url,
                "image": str(row[5] or ""),
            })

        if not products_in:
            print(f"\nSkipping '{category}' -- no products.")
            continue

        if demo_limit:
            products_in = products_in[:demo_limit]

        print(f"\n{'='*60}")
        print(f"[{category}]  {len(products_in)} products")
        print(f"{'='*60}")

        output_rows: list[dict] = []

        for i, p in enumerate(products_in, 1):
            safe = p["name"][:55].encode("ascii", "replace").decode()
            print(f"  [{i}/{len(products_in)}] {safe} ...", end=" ", flush=True)

            if is_fabric_url(p["url"]):
                detail = scrape_fabric_detail(p["url"], p["name"], p["image"])
            else:
                detail = scrape_product_detail(p["url"], p["name"], p["image"])

            if "_error" in detail:
                print(f"ERROR: {detail['_error']}")
            else:
                print("OK")

            row_data: dict = {
                "Product Name":     detail.get("Product Name")  or p["name"],
                "Source":           p["url"],
                "Image URL":        detail.get("Image URL")     or p["image"],
                "SKU":              p["sku"] or generate_sku(category, i),
                "Width":            detail.get("Width",         ""),
                "Depth":            detail.get("Depth",         ""),
                "Height":           detail.get("Height",        ""),
                "Diameter":         detail.get("Diameter",      ""),
                "Finish":           detail.get("Finish",        ""),
                "Description":      detail.get("Description",   ""),
                "Weight":           detail.get("Weight",        ""),
                "Inside Width":     detail.get("Inside Width",  ""),
                "Inside Depth":     detail.get("Inside Depth",  ""),
                "Inside Height":    detail.get("Inside Height", ""),
                "Seat Height":      detail.get("Seat Height",   ""),
                "Arm Height":       detail.get("Arm Height",    ""),
                "Seat Depth":       detail.get("Seat Depth",    ""),
                "COM":              detail.get("COM",           ""),
                "COL":              detail.get("COL",           ""),
                "COT":              detail.get("COT",           ""),
                "Grade":            detail.get("Grade",         ""),
                "Content":          detail.get("Content",       ""),
                "Cut Direction":    detail.get("Cut Direction", ""),
                "Cleaning Code":    detail.get("Cleaning Code", ""),
                "Repeat":           detail.get("Repeat",        ""),
                "Rub Count":        detail.get("Rub Count",     ""),
                "Rub Notes":        detail.get("Rub Notes",     ""),
                "Pattern":          detail.get("Pattern",       ""),
                "Construction":     detail.get("Construction",  ""),
                "CFA":              detail.get("CFA",           ""),
                "Cushion":          detail.get("Cushion",       ""),
                "Tearsheet":        detail.get("Tearsheet",     ""),
                "Color":            detail.get("Color",         ""),
                "Length":           detail.get("Length",        ""),
            }
            output_rows.append(row_data)
            time.sleep(SLEEP)

        save_sheet(wb_out, category, category_url, output_rows)
        wb_out.save(str(out_path))
        print(f"  Saved '{category}' ({len(output_rows)} products)")

    wb_out.save(str(out_path))
    print(f"\nDone -> {out_path}")


if __name__ == "__main__":
    main(demo="--demo" in sys.argv)
