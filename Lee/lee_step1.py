"""
Lee Industries -- Step 1
========================
Reads 'Lee' sheet from Status Tracker.
Scrapes all category listing pages with requests + BeautifulSoup.
Saves to Lee/Lee_step1.xlsx -- one sheet per category.

Usage:
    python lee_step1.py
"""

import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

SCRIPT_DIR   = Path(__file__).parent
TRACKER_FILE = SCRIPT_DIR.parent / "SD_Web Scraping - Status Tracker.xlsx"
OUT_FILE     = SCRIPT_DIR / "Lee_step1.xlsx"
BASE_URL     = "https://www.leeindustries.com"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
}
SLEEP = 0.5


def read_categories() -> list[dict]:
    wb = openpyxl.load_workbook(str(TRACKER_FILE))
    ws = wb["Lee"]

    categories: list[dict] = []
    current_cat = None
    current_urls: list[str] = []

    for r in range(1, ws.max_row + 1):
        c3 = str(ws.cell(r, 3).value or "").strip()
        c4 = str(ws.cell(r, 4).value or "").strip()

        if c3 == "Category" and c4:
            if current_cat and current_urls:
                # skip duplicate category names (Fabric appears twice)
                if not any(c["category"] == current_cat for c in categories):
                    categories.append({
                        "category":  current_cat,
                        "urls":      current_urls,
                        "first_url": current_urls[0],
                    })
            current_cat  = c4
            current_urls = []

        elif c3.startswith("Link") and c4.startswith("http"):
            if c4 not in current_urls:
                current_urls.append(c4)

    if current_cat and current_urls:
        if not any(c["category"] == current_cat for c in categories):
            categories.append({
                "category":  current_cat,
                "urls":      current_urls,
                "first_url": current_urls[0],
            })

    wb.close()
    return categories


def is_fabric_url(url: str) -> bool:
    return "/Upholstery/" in url or "/Options/" in url


def scrape_product_listing(url: str) -> list[dict]:
    """Scrape a /Product/Category/ page -- returns list of product dicts."""
    products: list[dict] = []
    seen: set[str] = set()

    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        print(f"    Fetch error: {e}")
        return products

    soup = BeautifulSoup(r.text, "html.parser")

    for item in soup.find_all("gallery-item"):
        link_el = item.find("a", class_="gallery-image")
        if not link_el:
            continue
        href = link_el.get("href", "").strip()
        if not href or href in seen:
            continue
        seen.add(href)

        product_url = BASE_URL + href if href.startswith("/") else href

        img_el = item.find("img")
        img_src = ""
        if img_el:
            img_src = img_el.get("src") or img_el.get("data-src") or ""
            img_src = img_src.split("?")[0]

        sku_el   = item.find("a", class_="sku")
        name     = ""
        sku_text = ""
        if sku_el:
            spans = sku_el.find_all("span")
            if spans:
                sku_text = spans[0].get_text(strip=True)
                name     = " ".join(s.get_text(strip=True) for s in spans)

        products.append({
            "Product Name": name or sku_text or href.split("/")[-1],
            "SKU":          sku_text or href.split("/")[-1],
            "Product URL":  product_url,
            "Image URL":    img_src,
        })

    return products


def scrape_fabric_listing(url: str) -> list[dict]:
    """Scrape a /Upholstery/ or /Options/ page -- fabric/leather/pillows."""
    products: list[dict] = []
    seen: set[str] = set()

    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        print(f"    Fetch error: {e}")
        return products

    soup = BeautifulSoup(r.text, "html.parser")

    # Fabric / Leather: a.fabric-gallery-image inside gallery-item
    for item in soup.find_all("gallery-item"):
        link_el = item.find("a", class_="fabric-gallery-image")
        if not link_el:
            # Pillows may use regular gallery-image
            link_el = item.find("a", class_="gallery-image")
        if not link_el:
            continue
        href = link_el.get("href", "").strip()
        if not href or href in seen:
            continue
        seen.add(href)

        product_url = BASE_URL + href if href.startswith("/") else href

        img_el  = item.find("img")
        img_src = ""
        if img_el:
            img_src = img_el.get("src") or img_el.get("data-src") or ""
            img_src = img_src.split("?")[0]

        name     = link_el.get("data-name", "").strip()
        sku_code = ""
        if not name:
            # Try a.sku spans (used by Pillows page): first span = SKU code
            sku_el = item.find("a", class_="sku")
            if sku_el:
                spans = sku_el.find_all("span")
                if spans:
                    sku_code = spans[0].get_text(strip=True)
                    name     = " ".join(s.get_text(strip=True) for s in spans)
        if not name:
            slug = href.rstrip("/").split("/")[-2] if href.count("/") >= 2 else href.split("/")[-1]
            name = slug.replace("-", " ").title()

        products.append({
            "Product Name": name,
            "SKU":          sku_code or name,
            "Product URL":  product_url,
            "Image URL":    img_src,
        })

    return products


def main() -> None:
    if not TRACKER_FILE.exists():
        print(f"ERROR: Status Tracker not found: {TRACKER_FILE}")
        return

    print("Reading Lee categories from Status Tracker...")
    categories = read_categories()
    print(f"Found {len(categories)} categories:\n")
    for c in categories:
        print(f"  {c['category']}  ({len(c['urls'])} URL(s))")

    wb = openpyxl.Workbook()
    for s in list(wb.sheetnames):
        del wb[s]

    bold    = Font(bold=True)
    col_hdrs = ["#", "Category", "Product Name", "SKU", "Product URL", "Image URL", "Category URL"]

    for cat in categories:
        category  = cat["category"]
        urls      = cat["urls"]
        first_url = cat["first_url"]

        print(f"\n{'='*60}")
        print(f"[{category}]  ({len(urls)} URL(s))")

        all_products: list[dict] = []
        seen_urls: set[str] = set()

        for url in urls:
            print(f"  Scraping: {url}")
            if is_fabric_url(url):
                prods = scrape_fabric_listing(url)
            else:
                prods = scrape_product_listing(url)

            for p in prods:
                if p["Product URL"] not in seen_urls:
                    all_products.append(p)
                    seen_urls.add(p["Product URL"])

            print(f"  +{len(prods)} products from this URL")
            time.sleep(SLEEP)

        print(f"  Total unique: {len(all_products)}")

        sheet_name = category[:31]
        ws = wb.create_sheet(sheet_name)
        for col, h in enumerate(col_hdrs, 1):
            ws.cell(1, col, h).font = bold

        for i, p in enumerate(all_products, 1):
            ws.cell(i + 1, 1, i)
            ws.cell(i + 1, 2, category)
            ws.cell(i + 1, 3, p["Product Name"])
            ws.cell(i + 1, 4, p["SKU"])
            ws.cell(i + 1, 5, p["Product URL"])
            ws.cell(i + 1, 6, p["Image URL"])
            ws.cell(i + 1, 7, first_url)

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"

        wb.save(str(OUT_FILE))
        print(f"  Saved sheet '{sheet_name}'")

    wb.save(str(OUT_FILE))
    print(f"\nDone -> {OUT_FILE}")


if __name__ == "__main__":
    main()
