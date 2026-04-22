#!/usr/bin/env python3
"""
SKU Scraping Agent
==================
Usage:
    python agent.py
    python agent.py "Julian Chichester"

The agent reads the Status Tracker to find all category URLs for a vendor,
scrapes each category page + product detail pages using Claude's intelligence,
and saves the output to {VendorName}.xlsx in Julian Chichester format.

Requirements:
    pip install anthropic requests beautifulsoup4 openpyxl selenium chromedriver-autoinstaller
"""

from __future__ import annotations
import os
import re
import sys
import json
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup, Comment
import anthropic

# ─── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
STATUS_TRACKER = SCRIPT_DIR / "SD_Web Scraping - Status Tracker.xlsx"

# ─── Config ───────────────────────────────────────────────────────────────────
MODEL = "claude-sonnet-4-6"
MAX_HTML_CHARS = 30000      # Max cleaned HTML chars sent to Claude per page
MAX_ITERATIONS = 300        # Safety limit on agent loop
CHROMEDRIVER_PATH = "C:/chromedriver.exe"

# ─── Anthropic client ─────────────────────────────────────────────────────────
client = anthropic.Anthropic()


# ══════════════════════════════════════════════════════════════════════════════
# TOOL IMPLEMENTATIONS
# ══════════════════════════════════════════════════════════════════════════════

def tool_get_vendor_categories(vendor_name: str) -> str:
    """Read Status Tracker and return all categories + URLs for a vendor."""
    try:
        wb = openpyxl.load_workbook(str(STATUS_TRACKER))
        ws = wb.active

        results = []
        vendor_lower = vendor_name.lower().strip()

        for r in range(2, ws.max_row + 1):
            cell_vendor = ws.cell(r, 2).value
            if not cell_vendor:
                continue
            if vendor_lower not in str(cell_vendor).lower().strip():
                continue

            category  = ws.cell(r, 1).value
            url       = ws.cell(r, 4).value   # Sample Link column
            done_flag = ws.cell(r, 7).value   # Scraping Status

            if not url or not str(url).strip().startswith("http"):
                continue

            results.append({
                "category": str(category).strip() if category else "Unknown",
                "url":      str(url).strip(),
                "already_scraped": bool(done_flag),
            })

        wb.close()

        if not results:
            # Try to suggest close matches
            wb2 = openpyxl.load_workbook(str(STATUS_TRACKER))
            ws2 = wb2.active
            vendors = set()
            for r in range(2, ws2.max_row + 1):
                v = ws2.cell(r, 2).value
                if v:
                    vendors.add(str(v).strip())
            wb2.close()
            matches = [v for v in vendors if vendor_lower[:4] in v.lower()][:10]
            return json.dumps({
                "error": f"No categories found for '{vendor_name}'.",
                "suggestions": matches,
            })

        return json.dumps({
            "vendor": vendor_name,
            "total_categories": len(results),
            "categories": results,
        })

    except Exception as e:
        return json.dumps({"error": f"tool_get_vendor_categories failed: {e}"})


# ─── HTML cleaning helpers ─────────────────────────────────────────────────────

def _semantic_summary(soup: BeautifulSoup, base_url: str, max_chars: int) -> str:
    """
    Extract a compact semantic representation:
    links, images, headings, product-like paragraphs.
    """
    lines: list[str] = []

    for tag in soup.find_all(["a", "img", "h1", "h2", "h3", "h4", "p", "li",
                               "td", "th", "dt", "dd", "span"]):
        try:
            if tag.name == "a":
                href = (tag.get("href") or "").strip()
                text = tag.get_text(" ", strip=True)
                if href and text and len(text) < 300:
                    lines.append(f'<a href="{href}">{text}</a>')

            elif tag.name == "img":
                src = (
                    tag.get("src") or tag.get("data-src") or
                    tag.get("data-lazy-src") or tag.get("data-original") or ""
                ).strip()
                alt = (tag.get("alt") or "").strip()
                if src:
                    lines.append(f'<img src="{src}" alt="{alt}">')

            else:
                text = tag.get_text(" ", strip=True)
                if text and 4 < len(text) < 600:
                    lines.append(f"<{tag.name}>{text}</{tag.name}>")
        except Exception:
            pass

    result = "\n".join(lines)
    return result[:max_chars]


def clean_html(html: str, base_url: str = "", max_chars: int = MAX_HTML_CHARS) -> str:
    """Remove noise, try to isolate product area, then compact if still large."""
    soup = BeautifulSoup(html, "html.parser")

    # Strip noise tags
    for tag in soup.find_all(["script", "style", "svg", "noscript",
                               "iframe", "meta", "link", "head"]):
        tag.decompose()
    for cmt in soup.find_all(string=lambda t: isinstance(t, Comment)):
        cmt.extract()

    # Try to find the main product area
    product_area_selectors = [
        "[class*='product-grid']", "[class*='product-list']",
        "[class*='products']",    "[class*='catalog']",
        "ul.products",            "ol.products",
        "div.products",           "main",
        "#content",               ".content",
    ]
    for sel in product_area_selectors:
        try:
            node = soup.select_one(sel)
            if node:
                candidate = str(node)
                if len(candidate) <= max_chars:
                    return candidate
        except Exception:
            pass

    # Full soup still fits?
    full = str(soup)
    if len(full) <= max_chars:
        return full

    # Fall back to semantic summary
    return _semantic_summary(soup, base_url, max_chars)


def tool_fetch_page_html(url: str,
                         use_selenium: bool = False,
                         wait_seconds: float = 3.0) -> str:
    """Fetch a page and return cleaned HTML for Claude to analyse."""
    try:
        raw = _selenium_fetch(url, wait_seconds) if use_selenium else _requests_fetch(url)
        cleaned = clean_html(raw, base_url=url)
        return json.dumps({
            "url":         url,
            "char_count":  len(cleaned),
            "html":        cleaned,
        })
    except Exception as e:
        return json.dumps({"error": str(e), "url": url})


def _requests_fetch(url: str) -> str:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    resp = requests.get(url, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.text


def _selenium_fetch(url: str, wait: float = 3.0) -> str:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options

    opts = Options()
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,1000")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])

    # Try chromedriver-autoinstaller first, then fall back to fixed path
    try:
        import chromedriver_autoinstaller as cda
        cda.install()
        drv = webdriver.Chrome(options=opts)
    except Exception:
        from selenium.webdriver.chrome.service import Service
        drv = webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=opts)

    drv.set_page_load_timeout(60)
    try:
        drv.get(url)
        time.sleep(wait)
        # Trigger lazy-load
        drv.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1.5)
        drv.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)
        return drv.page_source
    finally:
        try:
            drv.quit()
        except Exception:
            pass


# ─── Excel helpers ─────────────────────────────────────────────────────────────

def _generate_sku(vendor_name: str, category: str, index: int) -> str:
    """JULNI01 style SKU."""
    v = re.sub(r"[^A-Za-z]", "", vendor_name).upper()[:3]
    words = re.sub(r"[^A-Za-z\s]", "", category).strip().split()
    if len(words) >= 2:
        c = (words[0][0] + words[1][0]).upper()
    elif words:
        c = words[0][:2].upper()
    else:
        c = "XX"
    return f"{v}{c}{index:02d}"


# Fixed column order that mirrors Julian Chichester.xlsx
FIXED_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Finish",
]


def tool_save_products(
    vendor_name: str,
    category_name: str,
    category_url: str,
    products: list[dict],
    output_file: str,
) -> str:
    """Save scraped products to Excel in Julian Chichester multi-sheet format."""
    try:
        out_path = SCRIPT_DIR / output_file

        try:
            wb = openpyxl.load_workbook(str(out_path))
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            for name in list(wb.sheetnames):
                del wb[name]

        sheet_name = category_name[:31]
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        bold = Font(bold=True)

        # Row 1 – Brand
        ws.cell(1, 1, "Brand").font = bold
        ws.cell(1, 2, vendor_name)

        # Row 2 – Category Link
        ws.cell(2, 1, "Category Link").font = bold
        ws.cell(2, 2, category_url)

        # Row 3 – empty separator

        # Collect extra columns that some products have
        extra_cols: list[str] = []
        for p in products:
            for k in p:
                if k not in FIXED_COLS and k not in extra_cols:
                    extra_cols.append(k)

        all_cols = FIXED_COLS + extra_cols

        # Row 4 – Headers
        for col, header in enumerate(all_cols, 1):
            ws.cell(4, col, header).font = bold

        # Rows 5+ – Data
        for idx, product in enumerate(products, 1):
            product.setdefault("Index", idx)
            product.setdefault("Category", category_name)
            product.setdefault("Manufacturer", vendor_name)
            if not product.get("SKU"):
                product["SKU"] = _generate_sku(vendor_name, category_name, idx)
            if not product.get("Product Family Id"):
                product["Product Family Id"] = product.get("Product Name", "")

            for col, key in enumerate(all_cols, 1):
                ws.cell(4 + idx, col, product.get(key, ""))

        # Auto column widths
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max_len + 2, 60)

        ws.freeze_panes = "A5"

        wb.save(str(out_path))
        return json.dumps({
            "success": True,
            "message": f"Saved {len(products)} products → '{output_file}' sheet '{sheet_name}'",
            "rows": len(products),
        })

    except Exception as e:
        return json.dumps({"error": f"save_products failed: {e}"})


# ══════════════════════════════════════════════════════════════════════════════
# TOOL DISPATCH
# ══════════════════════════════════════════════════════════════════════════════

def execute_tool(name: str, inputs: dict) -> str:
    if name == "get_vendor_categories":
        return tool_get_vendor_categories(inputs["vendor_name"])

    elif name == "fetch_page_html":
        return tool_fetch_page_html(
            url=inputs["url"],
            use_selenium=inputs.get("use_selenium", False),
            wait_seconds=float(inputs.get("wait_seconds", 3.0)),
        )

    elif name == "save_products":
        return tool_save_products(
            vendor_name=inputs["vendor_name"],
            category_name=inputs["category_name"],
            category_url=inputs["category_url"],
            products=inputs["products"],
            output_file=inputs["output_file"],
        )

    return json.dumps({"error": f"Unknown tool: {name}"})


# ══════════════════════════════════════════════════════════════════════════════
# CLAUDE TOOLS SCHEMA
# ══════════════════════════════════════════════════════════════════════════════

TOOLS_SCHEMA = [
    {
        "name": "get_vendor_categories",
        "description": (
            "Look up a vendor in the Status Tracker Excel file and return all their "
            "product categories with scraping URLs. Call this FIRST before anything else."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "vendor_name": {
                    "type": "string",
                    "description": "The vendor / brand name to search for",
                }
            },
            "required": ["vendor_name"],
        },
    },
    {
        "name": "fetch_page_html",
        "description": (
            "Fetch a webpage and return cleaned HTML for you to read and extract data from. "
            "Use use_selenium=true for sites that load products with JavaScript. "
            "The returned HTML is pre-cleaned (scripts/styles removed). "
            "You will read this HTML and extract product information."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {"type": "string", "description": "The URL to fetch"},
                "use_selenium": {
                    "type": "boolean",
                    "description": "Use a real Chrome browser for JS-heavy pages (default false)",
                },
                "wait_seconds": {
                    "type": "number",
                    "description": "Seconds to wait for JS rendering (Selenium only, default 3)",
                },
            },
            "required": ["url"],
        },
    },
    {
        "name": "save_products",
        "description": (
            "Save collected products for one category to the Excel output file. "
            "Call this after you have scraped all products in a category. "
            "Each category gets its own sheet in the workbook."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "vendor_name": {"type": "string"},
                "category_name": {"type": "string", "description": "e.g. 'Dining Tables'"},
                "category_url": {"type": "string", "description": "The list page URL"},
                "products": {
                    "type": "array",
                    "description": "List of product objects",
                    "items": {
                        "type": "object",
                        "properties": {
                            "Product Name": {"type": "string"},
                            "Source": {
                                "type": "string",
                                "description": "Product detail page URL",
                            },
                            "Image URL": {"type": "string"},
                            "Description": {"type": "string"},
                            "Width": {
                                "type": "string",
                                "description": "Width in INCHES (2 decimal places)",
                            },
                            "Depth": {"type": "string", "description": "Depth in INCHES"},
                            "Height": {"type": "string", "description": "Height in INCHES"},
                            "Diameter": {
                                "type": "string",
                                "description": "Diameter in INCHES (for round items)",
                            },
                            "Finish": {"type": "string"},
                        },
                    },
                },
                "output_file": {
                    "type": "string",
                    "description": "Excel filename, e.g. 'Julian Chichester.xlsx'",
                },
            },
            "required": [
                "vendor_name",
                "category_name",
                "category_url",
                "products",
                "output_file",
            ],
        },
    },
]


# ══════════════════════════════════════════════════════════════════════════════
# SYSTEM PROMPT
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """You are a professional web-scraping AI agent. You collect furniture/decor product data from vendor websites and save it to Excel.

## Your Workflow

### Step 1 — Discover categories
Call `get_vendor_categories(vendor_name)`.
This returns a list of {category, url} pairs from the Status Tracker spreadsheet.

### Step 2 — Scrape each category
For every category URL:
1. Call `fetch_page_html(url)` to get the list page HTML.
2. Read the HTML and extract:
   - All product detail page links (href values pointing to individual products)
   - Product names, thumbnail image URLs (if visible on list page)
3. If there is a "next page" / pagination link, fetch those pages too until all products are collected.
4. For **each product detail URL** you found, call `fetch_page_html(product_url)` and extract:
   - **Product Name** — the main product title
   - **Source** — the product URL (same URL you fetched)
   - **Image URL** — main hero/product image src
   - **Description** — the full product description text (clean, no HTML tags)
   - **Width**, **Depth**, **Height**, **Diameter** — dimensions in **INCHES** (divide cm by 2.54, round to 2 dp)
   - **Finish** — finish, material, or color options
   - Any extra fields relevant to the category (e.g. Shape, Base Finish, Seat Height)
5. After collecting ALL products for a category, call `save_products(...)`.

### Step 3 — Repeat for all categories

## Rules

**Dimensions must be in INCHES.**
  - If you see cm: divide by 2.54 and round to 2 decimal places.
  - If you see mm: divide by 25.4 and round to 2 decimal places.
  - If a range is given (e.g. "60–72 inches"), use the smaller value.

**SKU generation:** [First 3 letters of vendor (UPPER)] + [First letter of each word in category (UPPER)] + [2-digit sequential number]
  - Julian Chichester + "Dining Tables"  → JULDI01, JULDI02 …
  - Brownstone + "Nightstands"          → BRONI01 …
  - Wesley Hall + "Lounge Chairs"       → WESLO01 …

**Output file name:** "{Vendor Name}.xlsx"  (e.g. "Julian Chichester.xlsx")

**JavaScript sites:** If `fetch_page_html` returns HTML with no product data (empty product area), retry with `use_selenium=true`.

**Pagination:** Look for "Next", ">", page numbers, or "Load More" links. Fetch ALL pages before moving to detail pages.

**Image URLs:** Always use absolute URLs. If you see a relative path like `/images/foo.jpg`, prepend the site base URL.

**Clean descriptions:** Strip HTML tags. Collapse whitespace. Keep meaningful text only.

**If a detail page fails:** Include the product with whatever data you have (name + URL at minimum), set other fields to empty string.

**Proceed category by category.** Save after each category. Do not wait until all categories are done to save.
"""


# ══════════════════════════════════════════════════════════════════════════════
# AGENT LOOP
# ══════════════════════════════════════════════════════════════════════════════

def run_agent(vendor_name: str) -> None:
    output_file = f"{vendor_name}.xlsx"

    print()
    print("=" * 62)
    print(f"  SKU Scraping Agent")
    print(f"  Vendor  : {vendor_name}")
    print(f"  Output  : {output_file}")
    print(f"  Tracker : {STATUS_TRACKER.name}")
    print("=" * 62)

    messages: list[dict] = [
        {
            "role": "user",
            "content": (
                f"Please scrape all product data for the vendor: '{vendor_name}'\n"
                f"Output file: '{output_file}'\n\n"
                f"Start by calling get_vendor_categories to see what categories exist, "
                f"then scrape each one systematically."
            ),
        }
    ]

    iteration = 0

    while iteration < MAX_ITERATIONS:
        iteration += 1
        print(f"\n[Iteration {iteration}]", end=" ", flush=True)

        response = client.messages.create(
            model=MODEL,
            max_tokens=8096,
            system=SYSTEM_PROMPT,
            tools=TOOLS_SCHEMA,
            messages=messages,
        )

        print(f"stop_reason={response.stop_reason}")

        # Print any text Claude outputs
        for block in response.content:
            if hasattr(block, "text") and block.text:
                preview = block.text.strip()
                if len(preview) > 400:
                    preview = preview[:400] + "…"
                print(f"\n[Claude] {preview}")

        # Done?
        if response.stop_reason == "end_turn":
            print("\n[Agent] Finished.")
            break

        # Handle tool calls
        if response.stop_reason == "tool_use":
            tool_results: list[dict] = []

            for block in response.content:
                if block.type != "tool_use":
                    continue

                # Summarise inputs for logging
                safe_inputs = {
                    k: (str(v)[:120] + "…" if len(str(v)) > 120 else str(v))
                    for k, v in block.input.items()
                    if k != "products"  # skip large product lists in log
                }
                print(f"\n  >> Tool: {block.name}({safe_inputs})")

                result = execute_tool(block.name, block.input)

                # Log brief result
                try:
                    r_obj = json.loads(result)
                    if "error" in r_obj:
                        print(f"  << ERROR: {r_obj['error']}")
                    elif "html" in r_obj:
                        print(f"  << HTML fetched ({r_obj.get('char_count', '?')} chars) from {r_obj.get('url','')}")
                    elif "message" in r_obj:
                        print(f"  << {r_obj['message']}")
                    else:
                        preview = result[:200] + ("…" if len(result) > 200 else "")
                        print(f"  << {preview}")
                except Exception:
                    print(f"  << {result[:200]}")

                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result,
                })

            messages.append({"role": "assistant", "content": response.content})
            messages.append({"role": "user", "content": tool_results})

        else:
            print(f"[Agent] Unexpected stop_reason: {response.stop_reason}")
            break

    if iteration >= MAX_ITERATIONS:
        print("\n[WARNING] Max iterations reached — agent stopped.")

    out_path = SCRIPT_DIR / output_file
    print()
    print("=" * 62)
    if out_path.exists():
        print(f"  Output saved → {out_path}")
    else:
        print(f"  Output file was NOT created: {out_path}")
    print("=" * 62)


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    print()
    print("╔══════════════════════════════════════════════════════════╗")
    print("║           SKU Scraping Agent  v1.0                      ║")
    print("║   Reads Status Tracker → Scrapes → Saves to Excel       ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print()

    if len(sys.argv) > 1:
        vendor_name = " ".join(sys.argv[1:]).strip()
        print(f"Vendor from CLI args: {vendor_name}")
    else:
        vendor_name = input("Enter vendor name: ").strip()

    if not vendor_name:
        print("No vendor name provided. Exiting.")
        sys.exit(1)

    run_agent(vendor_name)


if __name__ == "__main__":
    main()
