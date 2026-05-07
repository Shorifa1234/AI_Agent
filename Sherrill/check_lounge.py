import openpyxl
wb = openpyxl.load_workbook("Sherrill_demo.xlsx")
ws = wb["Lounge Chairs"]
headers = [c.value for c in ws[4]]
print("=== Lounge Chairs ===")
for i in range(5, 10):
    row = [c.value for c in ws[i]]
    d = dict(zip(headers, row))
    print(f"Row {i-4}: {d.get('Product Name')}")
    for k in ["Width","Depth","Height","Inside Length","Inside Depth","Seat Height","Arm Height","Finish","Body Fabric","Style","Product Type","Tags"]:
        v = d.get(k)
        if v:
            print(f"  {k}: {v}")
    print()
