"""
Sherrill Furniture -- Step 1
============================
Reads 'Sherrill' sheet from Status Tracker.
Scrapes all category list pages with requests + BeautifulSoup.
Saves to Sherrill/Sherrill_step1.xlsx -- one sheet per category.

Usage:
    python sherrill_step1.py
"""

import re
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

SCRIPT_DIR   = Path(__file__).parent
TRACKER_FILE = SCRIPT_DIR.parent / "SD_Web Scraping - Status Tracker.xlsx"
OUT_FILE     = SCRIPT_DIR / "Sherrill_step1.xlsx"
BASE_URL     = "https://www.sherrillfurniture.com"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
}
SLEEP = 0.6


# ── Status Tracker ─────────────────────────────────────────────────────────────

def read_categories() -> list[dict]:
    """
    Parse 'Sherrill' sheet.
    Handles Link / Link 2 rows for the same category.
    Returns list of {category, urls, first_url}.
    """
    wb = openpyxl.load_workbook(str(TRACKER_FILE))
    ws = wb["Sherrill"]

    categories: list[dict] = []
    current_cat = None
    current_urls: list[str] = []

    for r in range(1, ws.max_row + 1):
        c3 = str(ws.cell(r, 3).value or "").strip()
        c4 = str(ws.cell(r, 4).value or "").strip()

        if c3 == "Category" and c4:
            if current_cat and current_urls:
                categories.append({
                    "category":  current_cat,
                    "urls":      current_urls,
                    "first_url": current_urls[0],
                })
            current_cat  = c4
            current_urls = []

        elif c3.startswith("Link") and c4.startswith("http"):
            if c4 not in current_urls:
                current_urls.append(c4)

    if current_cat and current_urls:
        categories.append({
            "category":  current_cat,
            "urls":      current_urls,
            "first_url": current_urls[0],
        })

    wb.close()
    return categories


# ── Image upgrade ───────────────────────────────────────────────────────────────

def upgrade_image(src: str) -> str:
    """Convert medium thumbnail to full-size catalog image."""
    if not src:
        return ""
    # /styles/medium/public/catalog/DC333.jpg -> /files/catalog/DC333.jpg
    src = re.sub(r"/styles/[^/]+/public/", "/files/", src)
    return src.split("?")[0]


# ── List page scraper ──────────────────────────────────────────────────────────

def scrape_list_page(base_url: str) -> list[dict]:
    """Scrape one category list page (all pages via ?page=N)."""
    products: list[dict] = []
    seen: set[str] = set()
    page = 0

    while True:
        url = base_url if page == 0 else f"{base_url}?page={page}"
        print(f"    Page {page}: {url}")

        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
        except Exception as e:
            print(f"    Fetch error: {e}")
            break

        soup = BeautifulSoup(r.text, "html.parser")
        tiles = soup.select("a.product-results-tile")

        # Fallback: landing pages use dropdown links to catalog items
        if not tiles and page == 0:
            fallback = [
                a for a in soup.find_all("a", href=True)
                if "/catalog/" in a.get("href", "")
                and "dropdown-links" in " ".join(a.get("class", []))
            ]
            if fallback:
                print(f"    No tiles — using {len(fallback)} dropdown catalog link(s).")
                for a in fallback:
                    href = a.get("href", "").strip()
                    if not href or href in seen:
                        continue
                    seen.add(href)
                    product_url = BASE_URL + href if href.startswith("/") else href
                    model = a.get_text(strip=True).replace("Model ", "").strip() or href.split("/")[-1].upper()
                    products.append({
                        "Product Name": model,
                        "SKU":          model,
                        "Product URL":  product_url,
                        "Image URL":    "",
                    })
                break
            print(f"    No tiles on page {page} -- stopping.")
            break

        if not tiles:
            print(f"    No tiles on page {page} -- stopping.")
            break

        new_count = 0
        for tile in tiles:
            href = tile.get("href", "").strip()
            if not href or href in seen:
                continue
            seen.add(href)

            product_url = BASE_URL + href if href.startswith("/") else href

            img_el  = tile.find("img")
            img_src = upgrade_image(img_el.get("src", "") if img_el else "")

            h3 = tile.find("h3")
            model = h3.get_text(strip=True).replace("Model ", "").strip() if h3 else href.split("/")[-1].upper()

            products.append({
                "Product Name": model,
                "SKU":          model,
                "Product URL":  product_url,
                "Image URL":    img_src,
            })
            new_count += 1

        print(f"    +{new_count} products (total: {len(products)})")

        if new_count == 0:
            break

        # Check for next page link
        next_link = soup.select_one("a[title='Go to next page'], li.pager-next a")
        if not next_link:
            break

        page += 1
        time.sleep(SLEEP)

    return products


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if not TRACKER_FILE.exists():
        print(f"ERROR: Status Tracker not found: {TRACKER_FILE}")
        return

    print("Reading Sherrill categories from Status Tracker...")
    categories = read_categories()
    print(f"Found {len(categories)} categories with URLs:\n")
    for c in categories:
        print(f"  {c['category']} ({len(c['urls'])} URL(s))")

    wb = openpyxl.Workbook()
    for s in list(wb.sheetnames):
        del wb[s]

    bold    = Font(bold=True)
    headers = ["#", "Category", "Product Name", "SKU", "Product URL", "Image URL", "Category URL"]

    for cat in categories:
        category  = cat["category"]
        urls      = cat["urls"]
        first_url = cat["first_url"]

        print(f"\n{'='*60}")
        print(f"[{category}]  ({len(urls)} URL(s))")
        print(f"{'='*60}")

        all_products: list[dict] = []
        seen: set[str] = set()

        for url in urls:
            print(f"  URL: {url}")
            prods = scrape_list_page(url)
            for p in prods:
                if p["Product URL"] not in seen:
                    all_products.append(p)
                    seen.add(p["Product URL"])
            time.sleep(SLEEP)

        print(f"  Total unique products: {len(all_products)}")

        sheet_name = category[:31]
        ws = wb.create_sheet(sheet_name)
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h).font = bold

        for i, p in enumerate(all_products, 1):
            ws.cell(i + 1, 1, i)
            ws.cell(i + 1, 2, category)
            ws.cell(i + 1, 3, p["Product Name"])
            ws.cell(i + 1, 4, p["SKU"])
            ws.cell(i + 1, 5, p["Product URL"])
            ws.cell(i + 1, 6, p["Image URL"])
            ws.cell(i + 1, 7, first_url)

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"

        wb.save(str(OUT_FILE))
        print(f"  Saved sheet '{sheet_name}'")

    wb.save(str(OUT_FILE))
    print(f"\nDone -> {OUT_FILE}")


if __name__ == "__main__":
    main()
