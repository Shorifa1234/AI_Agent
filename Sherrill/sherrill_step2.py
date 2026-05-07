"""
Sherrill Furniture -- Step 2
============================
Reads Sherrill_step1.xlsx (one sheet per category).
Visits each product detail page with requests + BeautifulSoup.
Extracts: dimensions, finish, fabric, full-size image.
Saves to Sherrill/Sherrill.xlsx in Julian Chichester format.

Usage:
    python sherrill_step2.py
    python sherrill_step2.py --demo      # 5 products per category (test)
"""

import re
import sys
import json
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

SCRIPT_DIR = Path(__file__).parent
STEP1_FILE = SCRIPT_DIR / "Sherrill_step1.xlsx"
BASE_URL   = "https://www.sherrillfurniture.com"
VENDOR     = "Sherrill"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
}
SLEEP = 0.5

FIXED_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Base SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Length", "Weight",
    "Extension", "Canopy", "Maximum Adjustable Height",
    "Outside Length", "Outside Depth", "Outside Height",
    "Inside Length", "Inside Depth", "Inside Height", "Seat Height", "Arm Height",
    "Finish", "Finish Sample Code", "Color", "Collection",
    "Materials", "Material", "Origin", "Country of Origin",
    "Body Fabric", "Welt Fabric",
    "Price", "List Price", "Availability", "Shipping", "Shipping Method",
    "Style", "Product Type", "Features", "Tags",
    "Total Bulbs", "Wattage", "Dimmable", "Lamping Type",
    "Socket", "Voltage", "Shape", "Glass Features",
    "Install Position", "UL Ratings", "Prop 65", "Title 20", "Warranty", "UPC",
    "All SKUs",
]


# ── SKU ────────────────────────────────────────────────────────────────────────

def generate_sku(category: str, index: int) -> str:
    words = re.sub(r"[^A-Za-z\s]", "", category).strip().split()
    if len(words) >= 2:
        code = (words[0][0] + words[1][0]).upper()
    elif words:
        code = words[0][:2].upper()
    else:
        code = "XX"
    return f"SHE{code}{index:02d}"


# ── Dimension + detail parser ───────────────────────────────────────────────────

def parse_meta_description(text: str) -> dict[str, str]:
    """
    Parse Sherrill meta description content.
    Handles three formats:
      Compact:  'H41 W21 D16 in.' or 'W28 D36 H37'
      Labeled:  'Width/Dia: 91 Depth: 39 Height: 37'
      Mixed:    'W28 D36 H37 Inside Width: 19 Inside Depth: 22'
    """
    result = {
        "Width": "", "Depth": "", "Height": "", "Diameter": "",
        "Inside Length": "", "Inside Depth": "",
        "Seat Height": "", "Arm Height": "",
        "Finish": "", "Body Fabric": "",
    }
    if not text:
        return result

    # ── Overall dimensions: compact first (H W D or W D H)
    m = re.search(r"H([\d.]+)\s+W([\d.]+)\s+D([\d.]+)", text, re.I)
    if m:
        result["Height"] = m.group(1)
        result["Width"]  = m.group(2)
        result["Depth"]  = m.group(3)
    else:
        m2 = re.search(r"W([\d.]+)\s+D([\d.]+)\s+H([\d.]+)", text, re.I)
        if m2:
            result["Width"]  = m2.group(1)
            result["Depth"]  = m2.group(2)
            result["Height"] = m2.group(3)

    # ── If no compact, try labeled overall: "Width/Dia: 91 Depth: 39 Height: 37"
    if not result["Width"] and not result["Height"]:
        m_w = re.search(r"(?<!\w)Width(?:/Dia)?:\s*([\d.]+)", text, re.I)
        m_d = re.search(r"(?<!Inside\s)Depth:\s*([\d.]+)", text, re.I)
        m_h = re.search(r"(?<!Seat\s)(?<!Arm\s)Height:\s*([\d.]+)", text, re.I)
        if m_w: result["Width"]  = m_w.group(1)
        if m_d: result["Depth"]  = m_d.group(1)
        if m_h: result["Height"] = m_h.group(1)

    # ── "Overall Depth/Width/Height" overrides compact if explicitly stated
    m_od = re.search(r"Overall\s+Depth:\s*([\d.]+)", text, re.I)
    m_ow = re.search(r"Overall\s+Width:\s*([\d.]+)", text, re.I)
    m_oh = re.search(r"Overall\s+Height:\s*([\d.]+)", text, re.I)
    if m_od: result["Depth"]  = m_od.group(1)
    if m_ow: result["Width"]  = m_ow.group(1)
    if m_oh: result["Height"] = m_oh.group(1)

    # ── Inside dimensions: labeled "Inside Width: 19 Inside Depth: 22"
    m_iw = re.search(r"Inside\s+Width:\s*([\d.]+)", text, re.I)
    m_id = re.search(r"Inside\s+Depth:\s*([\d.]+)", text, re.I)
    if m_iw: result["Inside Length"] = m_iw.group(1)
    if m_id: result["Inside Depth"]  = m_id.group(1)

    # ── Inside compact: "Inside: W18 D18"
    if not result["Inside Length"]:
        m_in = re.search(r"Inside[:\s]+W([\d.]+)\s+D([\d.]+)", text, re.I)
        if m_in:
            result["Inside Length"] = m_in.group(1)
            result["Inside Depth"]  = m_in.group(2)

    # ── Diameter
    m_dia = re.search(r"(?:Dia(?:meter)?|Width/Dia):\s*([\d.]+)", text, re.I)
    if m_dia and not result["Width"]:
        result["Diameter"] = m_dia.group(1)

    # ── Arm Height
    m_arm = re.search(r"Arm Height[:\s]*([\d.]+)", text, re.I)
    if m_arm:
        result["Arm Height"] = m_arm.group(1)

    # ── Seat Height (handles typo "Seat Heght" and variations of "Approx")
    m_seat = re.search(r"(?:Approx(?:imate)?\.?\s*)?Seat He[i]?ght[:\s]*([\d.]+)", text, re.I)
    if m_seat:
        result["Seat Height"] = m_seat.group(1)

    # ── Fabric ("As Shown Fabric:", "As Shown: Fabric:", or plain "Fabric:")
    m_fabric = re.search(r"(?:As Shown\s*:?\s+)?Fabric:\s*([^\n]+?)(?:\s+Finish:|$)", text, re.I)
    if m_fabric:
        result["Body Fabric"] = m_fabric.group(1).strip()

    # ── Finish: split on all occurrences, take last non-empty, strip trailing fabric noise
    parts = re.split(r"(?:Standard\s+)?Finish:\s*", text, flags=re.I)
    for candidate in reversed(parts[1:]):
        clean = candidate.strip()
        if clean:
            # trim "As Shown..." or "Fabric:..." that may follow the finish value
            clean = re.sub(r"\s+As Shown.*", "", clean, flags=re.I).strip()
            clean = re.sub(r"\s+Fabric:.*",  "", clean, flags=re.I).strip()
            if clean:
                result["Finish"] = clean
            break

    return result


def parse_datalayer(html_text: str) -> dict[str, str]:
    """Extract Style, Product Type, Tags from datalayer JSON in page script."""
    result = {"Style": "", "Product Type": "", "Tags": ""}
    m = re.search(r"dataLayer\s*=\s*(\[.*?\]);", html_text, re.DOTALL)
    if not m:
        return result
    try:
        data = json.loads(m.group(1))
        taxonomy = data[0].get("entityTaxonomy", {})

        styles = list(taxonomy.get("room_style", {}).values())
        if styles:
            result["Style"] = ", ".join(styles)

        ftypes = list(taxonomy.get("furniture_type", {}).values())
        if ftypes:
            result["Product Type"] = ", ".join(ftypes)

        rooms = list(taxonomy.get("room_type", {}).values())
        if rooms:
            result["Tags"] = ", ".join(rooms)
    except Exception:
        pass
    return result


# ── Detail page scraper ─────────────────────────────────────────────────────────

def scrape_detail(url: str, fallback_name: str = "") -> dict:
    data: dict = {}
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        data["_error"] = str(e)
        return data

    soup = BeautifulSoup(r.text, "html.parser")

    # Product Name from h1
    h1 = soup.find("h1")
    data["Product Name"] = h1.get_text(strip=True) if h1 else fallback_name

    # Full-size image -- first catalog image in HTML
    for img in soup.find_all("img"):
        src = img.get("src", "")
        if "/files/catalog/" in src:
            data["Image URL"] = src.split("?")[0]
            break

    # All dimensions, finish, fabric from meta description (most reliable)
    meta_desc = soup.find("meta", attrs={"name": "description"})
    if meta_desc:
        parsed = parse_meta_description(meta_desc.get("content", ""))
        data.update({k: v for k, v in parsed.items() if v})

    # Style, Product Type, Tags from datalayer JSON
    dl = parse_datalayer(r.text)
    data.update({k: v for k, v in dl.items() if v})

    return data


# ── Excel helpers ───────────────────────────────────────────────────────────────

def save_sheet(wb, category: str, category_url: str, rows: list[dict]) -> None:
    sheet_name = category[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    bold = Font(bold=True)

    ws.cell(1, 1, "Brand").font = bold
    ws.cell(1, 2, VENDOR)
    ws.cell(2, 1, "Category Link").font = bold
    ws.cell(2, 2, category_url)

    extra: list[str] = []
    for row in rows:
        for k in row:
            if k not in FIXED_COLS and not k.startswith("_") and k not in extra:
                extra.append(k)
    all_cols = FIXED_COLS + extra

    for col, h in enumerate(all_cols, 1):
        ws.cell(4, col, h).font = bold

    for i, row in enumerate(rows, 1):
        row.setdefault("Index",           i)
        row.setdefault("Category",        category)
        row.setdefault("Manufacturer",    VENDOR)
        row.setdefault("Product Family Id", row.get("Product Name", ""))
        if not row.get("SKU"):
            row["SKU"] = generate_sku(category, i)

        for col, key in enumerate(all_cols, 1):
            ws.cell(4 + i, col, row.get(key, ""))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A5"


# ── Main ───────────────────────────────────────────────────────────────────────

def main(demo: bool = False) -> None:
    if not STEP1_FILE.exists():
        print(f"ERROR: {STEP1_FILE} not found. Run sherrill_step1.py first.")
        return

    demo_limit = 5 if demo else None
    out_path   = SCRIPT_DIR / ("Sherrill_demo.xlsx" if demo else "Sherrill.xlsx")

    if demo:
        print("[DEMO MODE] 5 products per category -> Sherrill_demo.xlsx\n")

    wb_in = openpyxl.load_workbook(str(STEP1_FILE))

    try:
        wb_out = openpyxl.load_workbook(str(out_path))
    except FileNotFoundError:
        wb_out = openpyxl.Workbook()
        for s in list(wb_out.sheetnames):
            del wb_out[s]

    # Step1 columns: #, Category, Product Name, SKU, Product URL, Image URL, Category URL
    for sheet_name in wb_in.sheetnames:
        ws_in    = wb_in[sheet_name]
        category = sheet_name

        products_in: list[dict] = []
        category_url = ""

        for row in ws_in.iter_rows(min_row=2, values_only=True):
            url = str(row[4] or "").strip()
            if not url:
                continue
            if not category_url:
                category_url = str(row[6] or "")
            products_in.append({
                "name":  str(row[2] or ""),
                "sku":   str(row[3] or ""),
                "url":   url,
                "image": str(row[5] or ""),
            })

        if not products_in:
            print(f"\nSkipping '{category}' -- no products.")
            continue

        if demo_limit:
            products_in = products_in[:demo_limit]

        print(f"\n{'='*60}")
        print(f"[{category}]  {len(products_in)} products")
        print(f"{'='*60}")

        output_rows: list[dict] = []

        for i, p in enumerate(products_in, 1):
            safe = (p["name"][:55] or p["url"]).encode("ascii", "replace").decode()
            print(f"  [{i}/{len(products_in)}] {safe} ...", end=" ", flush=True)

            detail = scrape_detail(p["url"], fallback_name=p["name"])

            if "_error" in detail:
                print(f"ERROR: {detail['_error']}")
            else:
                print("OK")

            row: dict = {
                "Product Name":  detail.get("Product Name")  or p["name"],
                "Source":        p["url"],
                "Image URL":     detail.get("Image URL")     or p["image"],
                "Width":         detail.get("Width",         ""),
                "Depth":         detail.get("Depth",         ""),
                "Height":        detail.get("Height",        ""),
                "Diameter":      detail.get("Diameter",      ""),
                "Inside Length": detail.get("Inside Length", ""),
                "Inside Depth":  detail.get("Inside Depth",  ""),
                "Seat Height":   detail.get("Seat Height",   ""),
                "Arm Height":    detail.get("Arm Height",    ""),
                "Finish":        detail.get("Finish",        ""),
                "Body Fabric":   detail.get("Body Fabric",   ""),
                "Style":         detail.get("Style",         ""),
                "Product Type":  detail.get("Product Type",  ""),
                "Tags":          detail.get("Tags",          ""),
                "SKU":           p["sku"],
            }
            output_rows.append(row)
            time.sleep(SLEEP)

        save_sheet(wb_out, category, category_url, output_rows)
        wb_out.save(str(out_path))
        print(f"  Saved '{category}' ({len(output_rows)} products)")

    wb_out.save(str(out_path))
    print(f"\nDone -> {out_path}")


if __name__ == "__main__":
    main(demo="--demo" in sys.argv)
