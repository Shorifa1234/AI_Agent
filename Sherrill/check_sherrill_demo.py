import openpyxl
wb = openpyxl.load_workbook("Sherrill_demo.xlsx")
ws = wb.worksheets[0]
headers = [c.value for c in ws[4]]
print("Sheet:", ws.title)
print("Headers:", headers)
print()
for i in range(5, 8):
    row = [c.value for c in ws[i]]
    d = dict(zip(headers, row))
    print(f"Row {i-4}: {d.get('Product Name')}")
    for k in ["Width","Depth","Height","Inside Length","Inside Depth","Seat Height","Arm Height","Finish","Body Fabric","Style","Product Type","Tags","Source"]:
        v = d.get(k)
        if v:
            print(f"  {k}: {v}")
    print()
