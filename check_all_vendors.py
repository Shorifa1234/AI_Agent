import openpyxl
from pathlib import Path

FILES = {
    "Worlds Away":       r"d:\Automation\SKU_AGENT\WorldsAway\Worlds Away.xlsx",
    "Vanguard":          r"d:\Automation\SKU_AGENT\Vanguard\Vanguard.xlsx",
    "Blackman Cruz":     r"d:\Automation\SKU_AGENT\BlackmanCruz\Blackman Cruz.xlsx",
    "Crystorama":        r"d:\Automation\SKU_AGENT\Crystorama\Crystorama.xlsx",
    "Julian Chichester": r"d:\Automation\SKU_AGENT\Julian Chichester.xlsx",
}

for vendor, path in FILES.items():
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[0]

        # Try row 4 first (Julian Chichester format), then row 1
        headers_r4 = [ws.cell(4, c).value for c in range(1, ws.max_column + 1) if ws.cell(4, c).value]
        headers_r1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]

        headers = headers_r4 if len(headers_r4) > 3 else headers_r1
        print(f"\n{vendor} ({len(headers)} cols):")
        print("  " + " | ".join(str(h) for h in headers))

        # First data row
        data_row = 5 if len(headers_r4) > 3 else 2
        sample = {ws.cell(4 if data_row==5 else 1, c).value: ws.cell(data_row, c).value
                  for c in range(1, len(headers) + 1)}
        for k in ["Product Name", "SKU", "Availability", "Shipping", "Finish Sample Code"]:
            if k in sample:
                print(f"    {k}: {sample[k]}")
    except Exception as e:
        print(f"\n{vendor}: ERROR - {e}")
