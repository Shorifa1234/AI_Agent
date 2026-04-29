"""
STEP 2 — Wesley Hall Detail Page Scraper
Reads step1_products.xlsx, visits each Detail URL,
extracts product details, and saves to step2_products.xlsx
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
import re
import time

INPUT_FILE  = "step1_Bed_products.xlsx"
OUTPUT_FILE = "step2_products.xlsx"
BASE_URL    = "https://wesleyhall.com"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

COLUMNS = [
    "Index", "Category", "Manufacturer", "Source URL", "Image URL",
    "Product Name", "SKU", "Base SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Length", "Weight",
    "Extension", "Canopy", "Maximum Adjustable Height",
    "Outside Length", "Outside Depth", "Outside Height",
    "Inside Length", "Inside Depth", "Inside Height", "Seat Height", "Arm Height",
    "Finish", "Finish Sample Code", "Color", "Collection",
    "Materials", "Material", "Origin", "Country of Origin",
    "Body Fabric", "Welt Fabric",
    "Price", "List Price", "Availability", "Shipping", "Shipping Method",
    "Style", "Product Type", "Features", "Tags",
    "Total Bulbs", "Wattage", "Dimmable", "Lamping Type",
    "Socket", "Voltage", "Shape", "Glass Features",
    "Install Position", "UL Ratings", "Prop 65", "Title 20", "Warranty", "UPC",
    "All SKUs",
    # Wesley Hall-specific
    "Outside Dimension", "Inside Dimension", "Seat Dimension", "Comments", "Thumbnail URL",
]

def fetch_page(url):
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser")

def parse_dim(text, part):
    patterns = {
        "L":    r'L\s*([\d.]+)',
        "D":    r'D\s*([\d.]+)',
        "H":    r'H\s*([\d.]+)',
        "ARM":  r'ARM\s*H\s*([\d.]+)',
    }
    match = re.search(patterns.get(part, ""), text, re.IGNORECASE)
    return match.group(1) if match else ""

def parse_detail(soup, source_url, step1_row):
    data = {col: "" for col in COLUMNS}

    # Carry over from step1
    data["Manufacturer"]  = step1_row.get("Manufacturer", "Wesley Hall")
    data["Thumbnail URL"] = step1_row.get("Thumbnail URL", "")
    data["Source URL"]    = source_url

    # Product Name + SKU from detail page
    prod_div = soup.find("div", class_="prod-details")
    if prod_div:
        full_name = prod_div.get_text(strip=True)
        parts = full_name.split(" ", 1)
        data["SKU"]          = parts[0] if parts else step1_row.get("SKU", "")
        data["Product Name"] = parts[1] if len(parts) > 1 else full_name
    else:
        data["SKU"]          = step1_row.get("SKU", "")
        data["Product Name"] = step1_row.get("Product Name", "")

    data["Product Family Id"] = data["Product Name"]
    data["Image URL"] = f"{BASE_URL}/assets/images/products/thumbnail/{data['SKU']}.jpg"

    # Parse dimension tables + "SHOWN IN" fields
    outside_raw = inside_raw = seat_raw = ""
    for table in soup.find_all("table", class_="style_details"):
        for row in table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) < 2:
                continue
            label = cells[0].get_text(strip=True).upper().replace(":", "").strip()
            value = cells[1].get_text(strip=True)
            if label == "OUTSIDE":
                outside_raw = value
            elif label == "INSIDE":
                inside_raw = value
            elif label == "SEAT":
                seat_raw = value
            elif label == "BODY FABRIC":
                data["Body Fabric"] = value
            elif label == "WELT FABRIC":
                data["Welt Fabric"] = value
            elif label == "FINISH":
                data["Finish"] = value
            elif label == "COMMENTS":
                data["Comments"] = value

    data["Outside Dimension"] = outside_raw
    data["Outside Length"]    = parse_dim(outside_raw, "L")
    data["Outside Depth"]     = parse_dim(outside_raw, "D")
    data["Outside Height"]    = parse_dim(outside_raw, "H")

    data["Inside Dimension"]  = inside_raw
    data["Inside Length"]     = parse_dim(inside_raw, "L")
    data["Inside Depth"]      = parse_dim(inside_raw, "D")
    data["Inside Height"]     = parse_dim(inside_raw, "H")

    data["Seat Dimension"]    = seat_raw
    seat_h = re.search(r'H\s*([\d.]+)', seat_raw)
    data["Seat Height"]       = seat_h.group(1) if seat_h else ""
    data["Arm Height"]        = parse_dim(seat_raw, "ARM")

    return data

def load_step1(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows

def save_xlsx(records, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Details"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    for row_idx, rec in enumerate(records, start=2):
        for col_idx, key in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(key, ""))

    wb.save(filepath)
    print(f"\nSaved {len(records)} rows -> {filepath}")

def main():
    step1_rows = load_step1(INPUT_FILE)
    print(f"Loaded {len(step1_rows)} products from {INPUT_FILE}")

    records = []
    for i, row in enumerate(step1_rows, start=1):
        url = row.get("Detail URL", "")
        if not url:
            print(f"  [{i}] No URL, skipping")
            continue
        print(f"  [{i}/{len(step1_rows)}] {url}")
        try:
            soup   = fetch_page(url)
            detail = parse_detail(soup, url, row)
            records.append(detail)
        except Exception as e:
            print(f"    ERROR: {e}")
            rec = {col: "" for col in COLUMNS}
            rec["Manufacturer"]  = row.get("Manufacturer", "Wesley Hall")
            rec["SKU"]           = row.get("SKU", "")
            rec["Product Name"]  = row.get("Product Name", "")
            rec["Thumbnail URL"] = row.get("Thumbnail URL", "")
            rec["Source URL"]    = url
            records.append(rec)

        time.sleep(0.5)

    save_xlsx(records, OUTPUT_FILE)

if __name__ == "__main__":
    main()