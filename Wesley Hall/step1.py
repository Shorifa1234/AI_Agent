"""
STEP 1 — Wesley Hall List Page Scraper
Category: Lounge Chairs (multiple URLs)
Output: step1_Lounge_products.xlsx
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font

BASE_URL     = "https://wesleyhall.com"
MANUFACTURER = "Wesley Hall"
OUTPUT_FILE  = "step1_Benches_products.xlsx"

SEARCH_URLS = [
    "https://wesleyhall.com/styles/func/cat/BCH",
    "https://wesleyhall.com/styles/func/cat/LBC",
    
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

def fetch_page(url):
    print(f"[GET] {url}")
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    print(f"  -> {resp.status_code}  ({len(resp.content):,} bytes)")
    return BeautifulSoup(resp.text, "html.parser")

def parse_products(soup):
    cards = soup.find_all("div", class_="style_thumbs")
    print(f"  -> Found {len(cards)} product card(s)")
    products = []
    for card in cards:
        a_tag      = card.find("a", href=True)
        detail_url = (BASE_URL + a_tag["href"]) if a_tag else ""

        img_tag       = card.find("img")
        thumbnail_url = ""
        if img_tag:
            thumbnail_url = img_tag.get("src") or img_tag.get("lazyload", "")
            if thumbnail_url and not thumbnail_url.startswith("http"):
                thumbnail_url = BASE_URL + thumbnail_url

        bold = card.find("b")
        sku  = bold.get_text(strip=True) if bold else ""

        desc_span    = card.find("span", class_="desc")
        product_name = desc_span.get_text(strip=True) if desc_span else ""

        products.append({
            "Manufacturer":  MANUFACTURER,
            "SKU":           sku,
            "Product Name":  product_name,
            "Detail URL":    detail_url,
            "Thumbnail URL": thumbnail_url,
        })
    return products

def save_xlsx(products, filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    columns = ["Manufacturer", "SKU", "Product Name", "Detail URL", "Thumbnail URL"]

    # Bold header only
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, product in enumerate(products, start=2):
        for col_idx, key in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=product[key])

    wb.save(filepath)
    print(f"\nSaved {len(products)} rows -> {filepath}")

def main():
    all_products = []
    seen_skus    = set()

    for url in SEARCH_URLS:
        soup     = fetch_page(url)
        products = parse_products(soup)
        for p in products:
            if p["SKU"] not in seen_skus:
                seen_skus.add(p["SKU"])
                all_products.append(p)
            else:
                print(f"  -> Duplicate SKU {p['SKU']} skipped")

    print(f"\nTotal unique products: {len(all_products)}")

    if not all_products:
        print("No products found.")
        return
    save_xlsx(all_products, OUTPUT_FILE)

if __name__ == "__main__":
    main()