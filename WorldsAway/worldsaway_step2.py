"""
Worlds Away -- Step 2
=====================
Reads WorldsAway_step1.xlsx.
Visits each product detail page with requests.
Extracts better image (og:image), full description, structured dimensions.
Saves to Worlds Away.xlsx in Julian Chichester format.

Usage:
    python worldsaway_step2.py
    python worldsaway_step2.py --demo      # 5 products per category (test)
"""

import re
import sys
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

SCRIPT_DIR = Path(__file__).parent
STEP1_FILE = SCRIPT_DIR / "WorldsAway_step1.xlsx"
OUT_FILE   = SCRIPT_DIR / "Worlds Away.xlsx"
VENDOR     = "Worlds Away"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
}
SLEEP = 0.5

FIXED_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Base SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Length", "Weight",
    "Extension", "Canopy", "Maximum Adjustable Height",
    "Outside Length", "Outside Depth", "Outside Height",
    "Inside Length", "Inside Depth", "Inside Height", "Seat Height", "Arm Height",
    "Finish", "Finish Sample Code", "Color", "Collection",
    "Materials", "Material", "Origin", "Country of Origin",
    "Body Fabric", "Welt Fabric",
    "Price", "List Price", "Availability", "Shipping", "Shipping Method",
    "Style", "Product Type", "Features", "Tags",
    "Total Bulbs", "Wattage", "Dimmable", "Lamping Type",
    "Socket", "Voltage", "Shape", "Glass Features",
    "Install Position", "UL Ratings", "Prop 65", "Title 20", "Warranty", "UPC",
    "All SKUs",
]


# ── SKU ────────────────────────────────────────────────────────────────────────

def generate_sku(category: str, index: int) -> str:
    """WOR + 2-letter category code + 2-digit index."""
    words = re.sub(r"[^A-Za-z\s]", "", category).strip().split()
    if len(words) >= 2:
        code = (words[0][0] + words[1][0]).upper()
    elif words:
        code = words[0][:2].upper()
    else:
        code = "XX"
    return f"WOR{code}{index:02d}"


# ── Dimension parser ────────────────────────────────────────────────────────────

def parse_dimensions(text: str) -> dict[str, str]:
    """
    Parse dimensions from detail page description (all in inches).

    Tries structured format first:  Width: 47.50"  Height: 18.38"  Depth: 23.75"
    Falls back to inline format:    18.375" H X 47.5" W X 23.75" D
    """
    result = {"Width": "", "Depth": "", "Height": "", "Diameter": ""}
    if not text:
        return result

    # Structured format
    m_dia = re.search(r'Diameter:\s*([\d.]+)"?', text, re.IGNORECASE)
    if m_dia:
        result["Diameter"] = m_dia.group(1)
        m_h = re.search(r'Height:\s*([\d.]+)"?', text, re.IGNORECASE)
        if m_h:
            result["Height"] = m_h.group(1)
        return result

    m_w  = re.search(r'Width:\s*([\d.]+)"?',  text, re.IGNORECASE)
    m_h  = re.search(r'Height:\s*([\d.]+)"?', text, re.IGNORECASE)
    m_d  = re.search(r'Depth:\s*([\d.]+)"?',  text, re.IGNORECASE)

    if m_w: result["Width"]  = m_w.group(1)
    if m_h: result["Height"] = m_h.group(1)
    if m_d: result["Depth"]  = m_d.group(1)

    if any(result.values()):
        return result

    # Inline format: '18.375" H X 47.5" W X 23.75" D'
    m_dia2 = re.search(r'([\d.]+)"\s*Dia', text, re.IGNORECASE)
    if m_dia2:
        result["Diameter"] = m_dia2.group(1)
        m_h2 = re.search(r'([\d.]+)"\s*H', text, re.IGNORECASE)
        if m_h2:
            result["Height"] = m_h2.group(1)
        return result

    m_w2 = re.search(r'([\d.]+)"\s*W', text, re.IGNORECASE)
    m_h2 = re.search(r'([\d.]+)"\s*H', text, re.IGNORECASE)
    m_d2 = re.search(r'([\d.]+)"\s*D(?!ia)', text, re.IGNORECASE)

    if m_w2: result["Width"]  = m_w2.group(1)
    if m_h2: result["Height"] = m_h2.group(1)
    if m_d2: result["Depth"]  = m_d2.group(1)

    return result


def extract_finish_sample_code(desc: str) -> str:
    """Extract 'Finish Sample Code: XXXX' from description text."""
    m = re.search(r'(?:Finish Sample Code|FINISH SAMPLE CODE):\s*([A-Z0-9]+)', desc, re.IGNORECASE)
    return m.group(1).strip() if m else ""


def extract_finish(desc: str) -> str:
    """
    Extract finish/color description.
    Remove dimension block and finish sample code line, return clean finish text.
    """
    if not desc:
        return ""
    clean = re.sub(r'\s*(Width|Height|Depth|Diameter):\s*[\d.]+["\s]*', "", desc, flags=re.IGNORECASE)
    clean = re.sub(r'(?:Finish Sample Code|FINISH SAMPLE CODE):\s*[A-Z0-9]+', "", clean, flags=re.IGNORECASE)
    clean = re.sub(r'\d+\.?\d*"\s*[HWDX\s]+(?:X\s*\d+\.?\d*"\s*[HWDX]+)*', "", clean)
    clean = re.sub(r'\s{2,}', " ", clean).strip()
    return clean[:300]


# ── Detail page scraper ─────────────────────────────────────────────────────────

def scrape_detail(url: str, fallback: dict) -> dict:
    """
    Visit product detail page and extract all fields.
    """
    data: dict = {}
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        data["_error"] = str(e)
        return data

    soup = BeautifulSoup(r.text, "html.parser")

    # Product Name
    h1 = soup.select_one("h1")
    if h1:
        data["Product Name"] = h1.get_text(strip=True)

    # High-quality image from og:image
    og = soup.find("meta", property="og:image")
    if og:
        data["Image URL"] = og.get("content", "").split("?")[0]

    # Full description (includes structured dimensions at end)
    desc_el = soup.find(attrs={"itemprop": "description"})
    if desc_el:
        full_desc = desc_el.get_text(" ", strip=True)

        # Parse dimensions
        dims = parse_dimensions(full_desc)
        data.update(dims)

        # Finish Sample Code
        data["Finish Sample Code"] = extract_finish_sample_code(full_desc)

        # Finish text
        data["Finish"] = extract_finish(full_desc)

        # Clean description (remove trailing dimension block)
        clean_desc = re.sub(
            r'\s*(Width|Height|Depth|Diameter):\s*[\d.]+["\s]*',
            "", full_desc, flags=re.IGNORECASE
        ).strip()
        data["Description"] = clean_desc

    # Availability: value is in dt.productView-info-value (not dd)
    stock_div = soup.select_one("div.prod-stock-level")
    if stock_div:
        val_el = stock_div.select_one("dt.productView-info-value")
        if val_el:
            data["Availability"] = val_el.get_text(strip=True)

    # Shipping
    shipping_val = soup.select_one("span.shipping-type-value")
    if shipping_val:
        data["Shipping"] = shipping_val.get_text(strip=True)

    return data


# ── Excel helpers ───────────────────────────────────────────────────────────────

def save_sheet(wb, category: str, category_url: str, rows: list[dict]) -> None:
    sheet_name = category[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    bold = Font(bold=True)

    ws.cell(1, 1, "Brand").font = bold
    ws.cell(1, 2, VENDOR)
    ws.cell(2, 1, "Category Link").font = bold
    ws.cell(2, 2, category_url)

    extra: list[str] = []
    for row in rows:
        for k in row:
            if k not in FIXED_COLS and not k.startswith("_") and k not in extra:
                extra.append(k)
    all_cols = FIXED_COLS + extra

    for col, h in enumerate(all_cols, 1):
        ws.cell(4, col, h).font = bold

    for i, row in enumerate(rows, 1):
        row.setdefault("Index",        i)
        row.setdefault("Category",     category)
        row.setdefault("Manufacturer", VENDOR)
        if not row.get("SKU"):
            row["SKU"] = generate_sku(category, i)
        if not row.get("Product Family Id"):
            row["Product Family Id"] = row.get("Product Name", "")

        for col, key in enumerate(all_cols, 1):
            ws.cell(4 + i, col, row.get(key, ""))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A5"


# ── Main ───────────────────────────────────────────────────────────────────────

def main(demo: bool = False) -> None:
    if not STEP1_FILE.exists():
        print(f"ERROR: {STEP1_FILE} not found. Run worldsaway_step1.py first.")
        return

    demo_limit = 5 if demo else None
    out_path   = SCRIPT_DIR / ("WorldsAway_demo.xlsx" if demo else "Worlds Away.xlsx")

    if demo:
        print("[DEMO MODE] 5 products per category -> WorldsAway_demo.xlsx\n")

    wb_in = openpyxl.load_workbook(str(STEP1_FILE))

    try:
        wb_out = openpyxl.load_workbook(str(out_path))
    except FileNotFoundError:
        wb_out = openpyxl.Workbook()
        for s in list(wb_out.sheetnames):
            del wb_out[s]

    # Step1 columns: #, Category, Product Name, SKU, Product URL, Image URL,
    #                Width, Depth, Height, Diameter, Description, Category URL
    COL = {"idx":0,"cat":1,"name":2,"sku":3,"url":4,"img":5,
           "w":6,"d":7,"h":8,"dia":9,"desc":10,"cat_url":11}

    for sheet_name in wb_in.sheetnames:
        ws_in    = wb_in[sheet_name]
        category = sheet_name

        products_in: list[dict] = []
        category_url = ""

        for row in ws_in.iter_rows(min_row=2, values_only=True):
            url = str(row[COL["url"]] or "").strip()
            if not url:
                continue
            if not category_url:
                category_url = str(row[COL["cat_url"]] or "")
            products_in.append({
                "name":    str(row[COL["name"]] or ""),
                "sku":     str(row[COL["sku"]]  or ""),
                "url":     url,
                "image":   str(row[COL["img"]]  or ""),
                "width":   str(row[COL["w"]]    or ""),
                "depth":   str(row[COL["d"]]    or ""),
                "height":  str(row[COL["h"]]    or ""),
                "dia":     str(row[COL["dia"]]  or ""),
                "desc":    str(row[COL["desc"]] or ""),
            })

        if not products_in:
            print(f"\nSkipping '{category}' -- no products.")
            continue

        if demo_limit:
            products_in = products_in[:demo_limit]

        print(f"\n{'='*60}")
        print(f"[{category}]  {len(products_in)} products")
        print(f"{'='*60}")

        output_rows: list[dict] = []

        for i, p in enumerate(products_in, 1):
            safe = (p["name"][:55] or p["url"]).encode("ascii", "replace").decode()
            print(f"  [{i}/{len(products_in)}] {safe} ...", end=" ", flush=True)

            detail = scrape_detail(p["url"], p)

            if "_error" in detail:
                print(f"ERROR: {detail['_error']}")
            else:
                print("OK")

            # Use detail page data; fall back to step1 data
            row: dict = {
                "Product Name": detail.get("Product Name") or p["name"],
                "Source":       p["url"],
                "Image URL":    detail.get("Image URL")    or p["image"],
                "Description":  detail.get("Description")  or p["desc"],
                "Width":        detail.get("Width")         or p["width"],
                "Depth":        detail.get("Depth")         or p["depth"],
                "Height":       detail.get("Height")        or p["height"],
                "Diameter":     detail.get("Diameter")      or p["dia"],
                "Finish":             detail.get("Finish",             ""),
                "Finish Sample Code": detail.get("Finish Sample Code", ""),
                "Availability":       detail.get("Availability",       ""),
                "Shipping":           detail.get("Shipping",           ""),
                "SKU":                p["sku"],
            }
            output_rows.append(row)
            time.sleep(SLEEP)

        save_sheet(wb_out, category, category_url, output_rows)
        wb_out.save(str(out_path))
        print(f"  Saved '{category}' ({len(output_rows)} products)")

    wb_out.save(str(out_path))
    print(f"\nDone -> {out_path}")


if __name__ == "__main__":
    main(demo="--demo" in sys.argv)
