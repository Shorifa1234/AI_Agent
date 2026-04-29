import openpyxl
wb = openpyxl.load_workbook("WorldsAway_demo.xlsx")
ws = wb.worksheets[0]
headers = [c.value for c in ws[4]]
print("Headers:", headers)
print()
for i in range(5, 9):
    row = [c.value for c in ws[i]]
    d = dict(zip(headers, row))
    print("Row", i - 4)
    print("  Product Name  :", d.get("Product Name"))
    print("  Availability  :", d.get("Availability"))
    print("  Finish Sample :", d.get("Finish Sample Code"))
    print("  Shipping      :", d.get("Shipping"))
    print("  Family Id     :", d.get("Product Family Id"))
    print("  Source        :", d.get("Source"))
    print()
