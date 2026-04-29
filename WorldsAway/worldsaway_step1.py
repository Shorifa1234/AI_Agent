"""
Worlds Away -- Step 1
=====================
Reads 'Worlds Away' sheet from Status Tracker.
Scrapes all category list pages using requests + BeautifulSoup.
No Selenium needed -- BigCommerce site renders products in static HTML.

Extracts from each card:
  - Product URL, Name, SKU, Image URL
  - Dimensions (card-dimensions-standard)
  - Description (card-desc)

Saves to WorldsAway/WorldsAway_step1.xlsx -- one sheet per category.

Usage:
    python worldsaway_step1.py
"""

import re
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from bs4 import BeautifulSoup

SCRIPT_DIR   = Path(__file__).parent
TRACKER_FILE = SCRIPT_DIR.parent / "SD_Web Scraping - Status Tracker.xlsx"
OUT_FILE     = SCRIPT_DIR / "WorldsAway_step1.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}
SLEEP = 0.6   # polite delay between requests


# ── Status Tracker ─────────────────────────────────────────────────────────────

def read_categories() -> list[dict]:
    """
    Parse 'Worlds Away' sheet.
    Returns list of {category, url} for rows that have a URL.
    """
    wb = openpyxl.load_workbook(str(TRACKER_FILE))
    ws = wb["Worlds Away"]

    categories = []
    current_cat = None

    for r in range(1, ws.max_row + 1):
        c_val = str(ws.cell(r, 3).value or "").strip()
        d_val = str(ws.cell(r, 4).value or "").strip()

        if c_val == "Category" and d_val:
            current_cat = d_val

        elif c_val == "Link" and current_cat:
            if d_val.startswith("http"):
                categories.append({"category": current_cat, "url": d_val})
            # If no URL (None), skip this category

    wb.close()
    return categories


# ── Helpers ────────────────────────────────────────────────────────────────────

def upgrade_image(data_src: str) -> str:
    """Upgrade BigCommerce stencil image from 265x265 to 1280x1280."""
    if not data_src:
        return ""
    url = re.sub(r"/\d+x\d+/", "/1280x1280/", data_src)
    return url.split("?")[0]   # strip cache-buster param


def parse_card_dims(dim_text: str) -> dict[str, str]:
    """
    Parse card-dimensions-standard text.
    Format: '23.75"D x 47.5"W x 18.375"H'  or  '18"Dia x 18"H'
    All dimensions already in inches.
    """
    result = {"Width": "", "Depth": "", "Height": "", "Diameter": ""}
    if not dim_text:
        return result

    m_dia = re.search(r'([\d.]+)"Dia', dim_text, re.IGNORECASE)
    if m_dia:
        result["Diameter"] = m_dia.group(1)
        m_h = re.search(r'([\d.]+)"H', dim_text, re.IGNORECASE)
        if m_h:
            result["Height"] = m_h.group(1)
        return result

    m_w = re.search(r'([\d.]+)"W', dim_text, re.IGNORECASE)
    m_h = re.search(r'([\d.]+)"H', dim_text, re.IGNORECASE)
    m_d = re.search(r'([\d.]+)"D', dim_text, re.IGNORECASE)

    if m_w: result["Width"]  = m_w.group(1)
    if m_h: result["Height"] = m_h.group(1)
    if m_d: result["Depth"]  = m_d.group(1)
    return result


# ── List page scraper ──────────────────────────────────────────────────────────

def scrape_list_pages(base_url: str) -> list[dict]:
    """
    Scrape all paginated list pages for a category.
    Returns list of product dicts.
    """
    products: list[dict] = []
    seen: set[str] = set()
    page = 1

    while True:
        url = base_url if page == 1 else f"{base_url}?page={page}"
        print(f"    Page {page}: {url}")

        try:
            r = requests.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
        except Exception as e:
            print(f"    Fetch error: {e}")
            break

        soup = BeautifulSoup(r.text, "html.parser")
        cards = soup.select("article.card")

        if not cards:
            print(f"    No cards on page {page} -- stopping.")
            break

        new_count = 0
        for card in cards:
            try:
                # URL
                a_tag = card.select_one("figure.card-figure a")
                product_url = (a_tag.get("href") or "").strip()
                if not product_url or product_url in seen:
                    continue
                seen.add(product_url)

                # Image
                img = card.select_one("img.card-image")
                img_src   = upgrade_image(img.get("data-src") or "") if img else ""
                name_from_alt = (img.get("alt") or "").strip() if img else ""

                # Name (from card-title; fallback to img alt)
                title_el = card.select_one("h4.card-title a")
                name = title_el.get_text(strip=True) if title_el else name_from_alt

                # Dimensions from card
                dim_el = card.select_one("div.card-dimensions-standard")
                dim_text = dim_el.get_text(" ", strip=True) if dim_el else ""
                dims = parse_card_dims(dim_text)

                # Description from card
                desc_el = card.select_one("div.card-desc")
                desc = desc_el.get_text(" ", strip=True) if desc_el else ""

                products.append({
                    "Product Name": name,
                    "SKU":          name,           # SKU = product name for Worlds Away
                    "Product URL":  product_url,
                    "Image URL":    img_src,
                    "Width":        dims["Width"],
                    "Depth":        dims["Depth"],
                    "Height":       dims["Height"],
                    "Diameter":     dims["Diameter"],
                    "Description":  desc,
                })
                new_count += 1

            except Exception:
                pass

        print(f"    +{new_count} products  (total: {len(products)})")

        if new_count == 0:
            break

        # Check for next page
        next_link = soup.select_one("a[rel='next'], .pagination a:-soup-contains('Next')")
        if not next_link:
            # Try checking if page+1 exists via pagination numbers
            pag_nums = []
            for a in soup.select(".pagination a"):
                try:
                    pag_nums.append(int(a.text.strip()))
                except Exception:
                    pass
            if page >= (max(pag_nums) if pag_nums else 1):
                break

        page += 1
        time.sleep(SLEEP)

    return products


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if not TRACKER_FILE.exists():
        print(f"ERROR: Status Tracker not found: {TRACKER_FILE}")
        return

    print("Reading categories from Status Tracker ('Worlds Away' sheet)...")
    categories = read_categories()
    print(f"Found {len(categories)} categories with URLs:\n")
    for c in categories:
        print(f"  {c['category']}")

    wb = openpyxl.Workbook()
    for s in list(wb.sheetnames):
        del wb[s]

    bold    = Font(bold=True)
    headers = [
        "#", "Category", "Product Name", "SKU",
        "Product URL", "Image URL",
        "Width", "Depth", "Height", "Diameter",
        "Description", "Category URL",
    ]

    for cat in categories:
        category = cat["category"]
        cat_url  = cat["url"]

        print(f"\n{'='*60}")
        print(f"[{category}]")
        print(f"  URL: {cat_url}")
        print(f"{'='*60}")

        products = scrape_list_pages(cat_url)
        print(f"  Total: {len(products)} products")

        sheet_name = category[:31]
        ws = wb.create_sheet(sheet_name)

        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h).font = bold

        for i, p in enumerate(products, 1):
            ws.cell(i + 1,  1, i)
            ws.cell(i + 1,  2, category)
            ws.cell(i + 1,  3, p["Product Name"])
            ws.cell(i + 1,  4, p["SKU"])
            ws.cell(i + 1,  5, p["Product URL"])
            ws.cell(i + 1,  6, p["Image URL"])
            ws.cell(i + 1,  7, p["Width"])
            ws.cell(i + 1,  8, p["Depth"])
            ws.cell(i + 1,  9, p["Height"])
            ws.cell(i + 1, 10, p["Diameter"])
            ws.cell(i + 1, 11, p["Description"])
            ws.cell(i + 1, 12, cat_url)

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"

        wb.save(str(OUT_FILE))
        print(f"  Saved sheet '{sheet_name}'")
        time.sleep(SLEEP)

    wb.save(str(OUT_FILE))
    print(f"\nDone -> {OUT_FILE}")


if __name__ == "__main__":
    main()
