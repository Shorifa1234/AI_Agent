"""
Gabby Step 2 — Detail Page Scraper
Reads: Gabby_step1.xlsx
Writes: Gabby.xlsx  (Julian Chichester format, one sheet per category)

Dimension pattern on gabby.com:
  Product depth 18.5"  Product height 28"  Product width 34"
"""

import requests
import openpyxl
import re
import time
import os
from bs4 import BeautifulSoup

INPUT_FILE  = os.path.join(os.path.dirname(__file__), "Gabby_step1.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "Gabby.xlsx")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

# 2-letter codes for SKU generation (GAB + code + sequence)
CATEGORY_CODES = {
    "Sofas & Loveseats":      "SO",
    "Sectionals":             "SE",
    "Lounge & Swivel Chairs": "LO",
    "Accent Chairs":          "AC",
    "Benches & Banquettes":   "BE",
    "Ottomans & Stools":      "OT",
    "Coffee Tables":          "CO",
    "Side & End Tables":      "SI",
    "Console Tables":         "CN",
    "Accent Tables":          "AT",
    "Cabinets":               "CA",
    "Bookcases":              "BO",
    "Credenzas":              "CR",
    "Sideboards & Buffets":   "SB",
    "Dining Tables":          "DI",
    "Dining Chairs":          "DC",
    "Bar & Counter Stools":   "BA",
    "Beds":                   "BD",
    "Headboards":             "HB",
    "Dressers":               "DR",
    "Nightstands":            "NI",
    "Desks":                  "DE",
    "Mirrors":                "MI",
    "Ceiling Lights":         "CL",
    "Lamps":                  "LA",
    "Wall Lights":            "WL",
}


def parse_dimensions(text):
    """Extract Width, Depth, Height, Diameter from Gabby page text."""
    dims = {"Width": "", "Depth": "", "Height": "", "Diameter": ""}
    patterns = {
        "Width":    r'[Pp]roduct\s+width\s+([\d.]+)',
        "Depth":    r'[Pp]roduct\s+depth\s+([\d.]+)',
        "Height":   r'[Pp]roduct\s+height\s+([\d.]+)',
        "Diameter": r'[Pp]roduct\s+diameter\s+([\d.]+)',
    }
    for key, pat in patterns.items():
        m = re.search(pat, text)
        if m:
            dims[key] = m.group(1)
    return dims


def fetch_product(url):
    """Scrape description, dimensions, and material from a Gabby product page."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.get_text(" ", strip=True)

        # Description — first paragraph(s) from product description section
        desc = ""
        desc_el = soup.select_one(
            ".product__description, .product-description, .rte, [class*='description']"
        )
        if desc_el:
            paras = [p.get_text(strip=True) for p in desc_el.find_all("p") if p.get_text(strip=True)]
            desc = " ".join(paras[:3])  # first 3 paragraphs max
        if not desc:
            # fallback: grab first long sentence from page text
            m = re.search(r'([A-Z][^.]{60,}\.)', text)
            if m:
                desc = m.group(1).strip()

        # Dimensions
        dims = parse_dimensions(text)

        # Material / Finish — from "Material X, Y, Z" in FEATURES block
        material = ""
        mat_m = re.search(r'\bMaterial\s+([A-Za-z][^\n]{3,80}?)(?=\s{2,}|\bCountry|\bContract|\bBrand|\bcom\b|$)', text)
        if mat_m:
            material = mat_m.group(1).strip()

        return desc, dims, material

    except Exception as e:
        print(f"    Error: {e}")
        return "", {"Width": "", "Depth": "", "Height": "", "Diameter": ""}, ""


def setup_sheet(ws, category_name, category_url):
    """Write standard header rows (Julian Chichester format)."""
    ws["A1"] = "Brand";         ws["B1"] = "Gabby"
    ws["A2"] = "Category Link"; ws["B2"] = category_url
    # row 3 empty
    headers = ["Index", "Category", "Manufacturer", "Source", "Image URL",
               "Product Name", "SKU", "Product Family Id", "Description",
               "Width", "Depth", "Height", "Diameter", "Finish"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)


def main():
    wb_in  = openpyxl.load_workbook(INPUT_FILE)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in wb_in.sheetnames:
        ws_in = wb_in[sheet_name]
        rows  = list(ws_in.iter_rows(min_row=2, values_only=True))
        if not rows:
            continue

        category_name = sheet_name
        cat_code      = CATEGORY_CODES.get(category_name, category_name[:2].upper())

        # Get collection slug from first data row (col 7)
        slug = rows[0][6] if len(rows[0]) > 6 else ""
        category_url = f"https://gabby.com/collections/{slug}" if slug else ""

        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        setup_sheet(ws_out, category_name, category_url)

        print(f"\n[{category_name}]  {len(rows)} products")

        index = 1
        for row in rows:
            category, name, mfr_sku, product_url, image_url = row[0], row[1], row[2], row[3], row[4]

            print(f"  [{index:>3}/{len(rows)}] {name}")
            desc, dims, material = fetch_product(product_url)

            sku = f"GAB{cat_code}{index:02d}"

            ws_out.append([
                index,
                category_name,
                "Gabby",
                product_url,
                image_url,
                name,
                sku,
                "",               # Product Family Id
                desc,
                dims["Width"],
                dims["Depth"],
                dims["Height"],
                dims["Diameter"],
                material,
            ])

            index += 1
            time.sleep(0.8)

        print(f"  Done: {index - 1} products")
        wb_out.save(OUTPUT_FILE)  # save after every category

    print(f"\nFinal output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
