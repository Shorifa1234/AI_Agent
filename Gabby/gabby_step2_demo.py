"""
Gabby Step 2 DEMO — first 5 products from first 3 categories only
"""

import requests
import openpyxl
import re
import time
import os
from bs4 import BeautifulSoup

INPUT_FILE  = os.path.join(os.path.dirname(__file__), "Gabby_step1.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "Gabby_demo.xlsx")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

CATEGORY_CODES = {
    "Sofas & Loveseats":      "SO",
    "Sectionals":             "SE",
    "Lounge & Swivel Chairs": "LO",
    "Accent Chairs":          "AC",
    "Benches & Banquettes":   "BE",
    "Ottomans & Stools":      "OT",
    "Coffee Tables":          "CO",
    "Side & End Tables":      "SI",
    "Console Tables":         "CN",
    "Accent Tables":          "AT",
    "Cabinets":               "CA",
    "Bookcases":              "BO",
    "Credenzas":              "CR",
    "Sideboards & Buffets":   "SB",
    "Dining Tables":          "DI",
    "Dining Chairs":          "DC",
    "Bar & Counter Stools":   "BA",
    "Beds":                   "BD",
    "Headboards":             "HB",
    "Dressers":               "DR",
    "Nightstands":            "NI",
    "Desks":                  "DE",
    "Mirrors":                "MI",
    "Ceiling Lights":         "CL",
    "Lamps":                  "LA",
    "Wall Lights":            "WL",
}

DEMO_LIMIT = 5        # products per category
DEMO_SHEETS = 3       # number of categories


def parse_dimensions(text):
    dims = {"Width": "", "Depth": "", "Height": "", "Diameter": ""}
    patterns = {
        "Width":    r'[Pp]roduct\s+width\s+([\d.]+)',
        "Depth":    r'[Pp]roduct\s+depth\s+([\d.]+)',
        "Height":   r'[Pp]roduct\s+height\s+([\d.]+)',
        "Diameter": r'[Pp]roduct\s+diameter\s+([\d.]+)',
    }
    for key, pat in patterns.items():
        m = re.search(pat, text)
        if m:
            dims[key] = m.group(1)
    return dims


def fetch_product(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        text = soup.get_text(" ", strip=True)

        # Description
        desc = ""
        desc_el = soup.select_one(
            ".product__description, .product-description, .rte, [class*='description']"
        )
        if desc_el:
            paras = [p.get_text(strip=True) for p in desc_el.find_all("p") if p.get_text(strip=True)]
            desc = " ".join(paras[:3])
        if not desc:
            m = re.search(r'([A-Z][^.]{60,}\.)', text)
            if m:
                desc = m.group(1).strip()

        # Dimensions
        dims = parse_dimensions(text)

        # Material
        material = ""
        mat_m = re.search(
            r'\bMaterial\s+([A-Za-z][^\n]{3,80}?)(?=\s{2,}|\bCountry|\bContract|\bBrand|\bcom\b|$)',
            text
        )
        if mat_m:
            material = mat_m.group(1).strip()

        return desc, dims, material

    except Exception as e:
        print(f"    Error: {e}")
        return "", {"Width": "", "Depth": "", "Height": "", "Diameter": ""}, ""


def setup_sheet(ws, category_name, category_url):
    ws["A1"] = "Brand";         ws["B1"] = "Gabby"
    ws["A2"] = "Category Link"; ws["B2"] = category_url
    headers = ["Index", "Category", "Manufacturer", "Source", "Image URL",
               "Product Name", "SKU", "Product Family Id", "Description",
               "Width", "Depth", "Height", "Diameter", "Finish"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)


def main():
    wb_in  = openpyxl.load_workbook(INPUT_FILE)
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for sheet_idx, sheet_name in enumerate(wb_in.sheetnames[:DEMO_SHEETS]):
        ws_in = wb_in[sheet_name]
        rows  = list(ws_in.iter_rows(min_row=2, values_only=True))[:DEMO_LIMIT]
        if not rows:
            continue

        category_name = sheet_name
        cat_code      = CATEGORY_CODES.get(category_name, category_name[:2].upper())
        slug          = rows[0][6] if len(rows[0]) > 6 else ""
        category_url  = f"https://gabby.com/collections/{slug}" if slug else ""

        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        setup_sheet(ws_out, category_name, category_url)

        print(f"\n[{category_name}]  (demo: {len(rows)} products)")

        for index, row in enumerate(rows, start=1):
            category, name, mfr_sku, product_url, image_url = row[0], row[1], row[2], row[3], row[4]
            print(f"  [{index}/{len(rows)}] {name}")
            desc, dims, material = fetch_product(product_url)
            sku = f"GAB{cat_code}{index:02d}"
            ws_out.append([
                index, category_name, "Gabby", product_url, image_url,
                name, sku, "", desc,
                dims["Width"], dims["Depth"], dims["Height"], dims["Diameter"],
                material,
            ])
            time.sleep(0.8)

    wb_out.save(OUTPUT_FILE)
    print(f"\nDemo output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
