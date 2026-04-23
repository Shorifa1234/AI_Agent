# crystorama_step1.py
# Crystorama product list collector — uses Shopify products.json API (no Selenium needed)
# deps: pip install requests openpyxl beautifulsoup4
#
# Output: crystorama_step1.xlsx
# Collected per product: Category, Product Name, Handle, SKU (base), All SKUs,
#                        Image URL, Description, Width, Finish, Product Type, Style, Tags

from __future__ import annotations
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font
from bs4 import BeautifulSoup

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
BASE_URL   = "https://www.crystorama.com"
OUTPUT_FILE = "crystorama_step1.xlsx"

CATEGORIES = [
    ("Chandeliers",       "chandeliers"),
    ("Pendants",          "pendants"),
    ("Flush & Semi-Flush","flush-semi-flush"),
    ("Wall Sconces",      "sconces"),
    ("Bath Vanities",     "bath-vanities"),
    ("Linear Fixtures",   "linear"),
    ("Task Lights",       "task-light"),
    ("Picture Lights",    "picture-light"),
    ("ADA Sconces",       "ada-sconces"),
    ("Hanging Shades",    "hanging-shades"),
    ("Outdoor Ceiling",   "outdoor-ceiling"),
    ("Outdoor Wall",      "outdoor-wall"),
    ("Post Lights",       "outdoor-post"),
    ("Mirrors",           "mirrors"),
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

DELAY_SEC = 0.3   # polite delay between API calls
# ──────────────────────────────────────────────────────────────────────────────


def _tag_value(tags: list[str], prefix: str) -> str:
    """Extract value from a tag like 'WIDTH:13.75' → '13.75'."""
    for t in tags:
        if t.upper().startswith(prefix.upper() + ":"):
            return t.split(":", 1)[1].strip()
    return ""


def _all_tag_values(tags: list[str], prefix: str) -> list[str]:
    """Extract all values for a given prefix, e.g. all FINISH: tags."""
    return [t.split(":", 1)[1].strip() for t in tags if t.upper().startswith(prefix.upper() + ":")]


def _base_sku(sku: str) -> str:
    """
    ABB-3003-VG  →  ABB-3003
    531-GA       →  531-GA   (already 2-part)
    Strips the last dash-segment if it looks like a finish code (2–3 letters/digits).
    """
    parts = sku.rsplit("-", 1)
    if len(parts) == 2 and re.fullmatch(r"[A-Z0-9]{1,4}", parts[1]):
        return parts[0]
    return sku


def _clean_desc(body_html: str) -> str:
    soup = BeautifulSoup(body_html or "", "html.parser")
    text = soup.get_text(" ", strip=True)
    return re.sub(r"\s+", " ", text).strip()


def fetch_collection(category_name: str, slug: str) -> list[dict]:
    """Paginate through products.json and return all products for a collection."""
    all_products: list[dict] = []
    page = 1
    seen_handles: set[str] = set()

    while True:
        url = f"{BASE_URL}/collections/{slug}/products.json?limit=250&page={page}"
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            products = resp.json().get("products", [])
        except Exception as e:
            print(f"    [ERROR] {url}: {e}")
            break

        if not products:
            break

        for p in products:
            handle = p.get("handle", "")
            if handle in seen_handles:
                continue
            seen_handles.add(handle)

            tags       = p.get("tags", [])
            variants   = p.get("variants", [])
            images     = p.get("images", [])

            # SKU: first variant is the "hero" finish; derive base SKU
            first_sku  = variants[0].get("sku", "") if variants else ""
            base        = _base_sku(first_sku)
            all_skus    = " | ".join(v.get("sku", "") for v in variants)

            # Image URL
            image_url = ""
            if images:
                img0 = images[0]
                if isinstance(img0, dict):
                    image_url = img0.get("src", "")
                elif isinstance(img0, str):
                    image_url = img0

            # From tags
            width      = _tag_value(tags, "WIDTH")
            finishes   = _all_tag_values(tags, "FINISH")
            finish     = " | ".join(finishes)
            style      = _tag_value(tags, "STYLE")
            family     = _tag_value(tags, "FAMILY")
            collection = _tag_value(tags, "COLLECTION")

            # Price from first variant
            price = variants[0].get("price", "") if variants else ""

            all_products.append({
                "Category":      category_name,
                "Category Slug": slug,
                "Product Name":  p.get("title", ""),
                "Handle":        handle,
                "Product URL":   f"{BASE_URL}/products/{handle}",
                "SKU":           first_sku,
                "Base SKU":      base,
                "All SKUs":      all_skus,
                "Image URL":     image_url,
                "Description":   _clean_desc(p.get("body_html", "")),
                "Width":         width,
                "Finish":        finish,
                "Style":         style,
                "Collection":    collection,
                "Product Type":  p.get("product_type", ""),
                "Family":        family,
                "Price":         price,
                "Tags":          ", ".join(tags),
            })

        print(f"    Page {page}: +{len(products)} products  (total {len(all_products)})")

        if len(products) < 250:
            break

        page += 1
        time.sleep(DELAY_SEC)

    return all_products


def save_to_excel(all_rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Products"

    if not all_rows:
        wb.save(OUTPUT_FILE)
        print("No products found.")
        return

    bold = Font(bold=True)
    cols = list(all_rows[0].keys())

    for c, header in enumerate(cols, 1):
        ws.cell(1, c, header).font = bold

    for r, row in enumerate(all_rows, 2):
        for c, key in enumerate(cols, 1):
            ws.cell(r, c, row.get(key, ""))

    # Auto width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"\nSaved {len(all_rows)} products to {OUTPUT_FILE}")


def main() -> None:
    print("=" * 60)
    print("Crystorama Step 1 — Product List Collection")
    print("=" * 60)

    all_rows: list[dict] = []

    for category_name, slug in CATEGORIES:
        print(f"\n[{category_name}] /collections/{slug}")
        rows = fetch_collection(category_name, slug)
        all_rows.extend(rows)
        print(f"  -> {len(rows)} products")
        time.sleep(DELAY_SEC)

    print(f"\nTotal: {len(all_rows)} products across {len(CATEGORIES)} categories")
    save_to_excel(all_rows)


if __name__ == "__main__":
    main()
