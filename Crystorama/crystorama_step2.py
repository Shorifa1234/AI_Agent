# crystorama_step2.py
# Crystorama Step 2 -- Detail page scraper + final Excel output
# Reads crystorama_step1.xlsx, visits each product page to get dimensions,
# then saves Crystorama.xlsx in Julian Chichester format (one sheet per category).
#
# deps: pip install requests openpyxl beautifulsoup4

from __future__ import annotations
import os
import re
import time
import requests
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# ------------------------------------------------------------------------------
# CONFIG
# ------------------------------------------------------------------------------
INPUT_FILE   = "crystorama_step1.xlsx"
OUTPUT_FILE  = "Crystorama.xlsx"
VENDOR_NAME  = "Crystorama"

DELAY_SEC    = 0.4    # polite delay between product page requests
SAVE_EVERY_N = 50     # auto-save progress every N products

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# Dimension field labels to extract from detail page
DIM_FIELDS = ["Width", "Height", "Depth", "Diameter", "Extension", "Canopy"]

# Output columns in Julian Chichester format
OUTPUT_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Base SKU", "Product Family Id",
    "Description",
    "Width", "Depth", "Height", "Diameter", "Extension", "Canopy",
    "Finish", "Collection", "Color", "Style", "Product Type",
    "Total Bulbs", "Wattage", "Dimmable",
    "Price",
    "All SKUs",
]
# ------------------------------------------------------------------------------


# Dimension patterns: (regex label on page, output column name)
# More specific patterns first so they take priority
DIM_PATTERNS = [
    (r"Product\s+Width",              "Width"),
    (r"Product\s+Height",             "Height"),
    (r"Product\s+Depth",              "Depth"),
    (r"Product\s+Diameter",           "Diameter"),
    (r"Maximum\s+Adjustable\s+Height","Extension"),
    (r"Width",                        "Width"),
    (r"Height",                       "Height"),
    (r"Depth",                        "Depth"),
    (r"Diameter",                     "Diameter"),
    (r"Extension",                    "Extension"),
]


def extract_detail_data(html: str) -> dict:
    # Parse all useful fields from the crystorama.com product detail page
    data = {}
    soup = BeautifulSoup(html, "html.parser")

    # Collect text from ALL tab-content sections and the details accordion
    parts = []
    for tc in soup.select(".tab-content"):
        parts.append(tc.get_text(" ", strip=True))
    for sel in (".enable-lumens-details", ".disable-lumens-details"):
        el = soup.select_one(sel)
        if el:
            parts.append(el.get_text(" ", strip=True))
    text = " ".join(parts)

    # Dimensions (numeric inch values)
    for pattern, col in DIM_PATTERNS:
        if col in data:
            continue
        m = re.search(rf"{pattern}\s*:\s*([\d.]+)", text, re.I)
        if m:
            data[col] = m.group(1)

    # Canopy/Backplate: 5"W x 0.75"H  ->  first number is canopy width
    m = re.search(r"Canopy[^\w].*?:\s*([\d.]+)", text, re.I)
    if m:
        data["Canopy"] = m.group(1)

    # Finish (text value, not a number)
    m = re.search(r"Finish\s*:\s*([A-Za-z /&,]+?)(?:\s+(?:Total|Wattage|Product|Max|Dimmable|Model)\b|$)", text, re.I)
    if m:
        finish = m.group(1).strip().rstrip(",")
        if finish:
            data["Finish"] = finish

    # Extra useful fields
    m = re.search(r"Total\s+Bulbs\s*:\s*(\d+)", text, re.I)
    if m:
        data["Total Bulbs"] = m.group(1)
    m = re.search(r"(?<!\w)Wattage\s*:\s*([\d.]+)\s*W", text, re.I)
    if m:
        data["Wattage"] = m.group(1) + "W"
    m = re.search(r"Dimmable\s*:\s*(Yes|No)", text, re.I)
    if m:
        data["Dimmable"] = m.group(1)

    return data


def fetch_detail(url: str) -> dict:
    # Fetch a product page and return all extracted fields
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        return extract_detail_data(resp.text)
    except Exception as e:
        print(f"    [WARN] fetch_detail failed ({url}): {e}")
        return {}


def load_step1(filepath: str) -> list:
    if not os.path.exists(filepath):
        raise FileNotFoundError(
            f"Step 1 file not found: '{filepath}'\n"
            "Run crystorama_step1.py first."
        )
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        if any(v for v in row.values()):
            rows.append(row)
    wb.close()
    return rows


def _generate_sku(category: str, index: int) -> str:
    words = re.sub(r"[^A-Za-z\s]", "", category).strip().split()
    if len(words) >= 2:
        c = (words[0][0] + words[1][0]).upper()
    elif words:
        c = words[0][:2].upper()
    else:
        c = "XX"
    return f"CRY{c}{index:02d}"


# ------------------------------------------------------------------------------
# Save helpers
# ------------------------------------------------------------------------------

def _write_excel(enriched_by_cat: dict, out_path: str) -> None:
    wb = openpyxl.Workbook()
    for name in list(wb.sheetnames):
        del wb[name]

    bold = Font(bold=True)

    for category, products in enriched_by_cat.items():
        sheet_name = category[:31]
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        ws.cell(1, 1, "Brand").font = bold
        ws.cell(1, 2, VENDOR_NAME)

        cat_slug = products[0].get("Category Slug", "") if products else ""
        cat_url  = f"https://www.crystorama.com/collections/{cat_slug}" if cat_slug else ""
        ws.cell(2, 1, "Category Link").font = bold
        ws.cell(2, 2, cat_url)

        extra_cols = []
        for p in products:
            for k in p:
                if k not in OUTPUT_COLS and k not in extra_cols and k != "Category Slug":
                    extra_cols.append(k)
        all_cols = OUTPUT_COLS + extra_cols

        for col, header in enumerate(all_cols, 1):
            ws.cell(4, col, header).font = bold

        for idx, prod in enumerate(products, 1):
            prod["Index"]            = idx
            prod["Manufacturer"]     = VENDOR_NAME
            prod["Category"]         = category
            prod["Source"]           = prod.get("Product URL", "")
            if not prod.get("SKU"):
                prod["SKU"]          = _generate_sku(category, idx)
            prod.setdefault("Base SKU",          prod.get("Base SKU", ""))
            prod.setdefault("Product Family Id", prod.get("Product Name", ""))

            for col, key in enumerate(all_cols, 1):
                ws.cell(4 + idx, col, prod.get(key, ""))

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max_len + 2, 55)

        ws.freeze_panes = "A5"

    wb.save(out_path)


def save_progress(enriched_by_cat: dict) -> None:
    tmp = OUTPUT_FILE.replace(".xlsx", "_tmp.xlsx")
    _write_excel(enriched_by_cat, tmp)
    try:
        if os.path.exists(OUTPUT_FILE):
            os.remove(OUTPUT_FILE)
        os.rename(tmp, OUTPUT_FILE)
    except PermissionError:
        alt = OUTPUT_FILE.replace(".xlsx", "_saved.xlsx")
        if os.path.exists(alt):
            os.remove(alt)
        os.rename(tmp, alt)
        print(f"  [WARN] '{OUTPUT_FILE}' is open in Excel -- saved to '{alt}' instead")


# ------------------------------------------------------------------------------
# Resume helper
# ------------------------------------------------------------------------------

def load_done_urls(out_path: str) -> set:
    done = set()
    if not os.path.exists(out_path):
        return done
    try:
        wb = openpyxl.load_workbook(out_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            src_col = None
            for c in range(1, ws.max_column + 1):
                if ws.cell(4, c).value == "Source":
                    src_col = c
                    break
            if src_col is None:
                continue
            for r in range(5, ws.max_row + 1):
                v = ws.cell(r, src_col).value
                if v:
                    done.add(str(v).strip())
        wb.close()
    except Exception:
        pass
    return done


# ------------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------------

def main() -> None:
    print("=" * 60)
    print("Crystorama Step 2 -- Detail Scraper")
    print(f"Input : {INPUT_FILE}")
    print(f"Output: {OUTPUT_FILE}")
    print("=" * 60)

    rows = load_step1(INPUT_FILE)
    print(f"Loaded {len(rows)} products from {INPUT_FILE}")

    done_urls = load_done_urls(OUTPUT_FILE)
    if done_urls:
        print(f"Resume: {len(done_urls)} products already in output -- skipping them")

    categories_order = []
    by_category = {}
    for row in rows:
        cat = row.get("Category", "Unknown")
        if cat not in by_category:
            by_category[cat] = []
            categories_order.append(cat)
        by_category[cat].append(row)

    print(f"Categories: {len(categories_order)}")

    enriched_by_cat = {cat: [] for cat in categories_order}
    processed = 0
    skipped   = 0
    total     = len(rows)

    for cat in categories_order:
        print(f"\n[{cat}] -- {len(by_category[cat])} products")

        for row in by_category[cat]:
            url  = row.get("Product URL", "")
            name = row.get("Product Name", url)

            if url in done_urls:
                skipped += 1
                continue

            processed += 1
            print(f"  [{processed}/{total - skipped}] {name[:65]}")

            dims = fetch_detail(url) if url else {}
            row.update(dims)

            enriched_by_cat[cat].append(row)

            if processed % SAVE_EVERY_N == 0:
                save_progress(enriched_by_cat)
                print(f"  [auto-save] {processed} products written to {OUTPUT_FILE}")

            time.sleep(DELAY_SEC)

    save_progress(enriched_by_cat)

    total_saved = sum(len(v) for v in enriched_by_cat.values())
    print(f"\n{'=' * 60}")
    print(f"Done! {total_saved} products saved to {OUTPUT_FILE}")
    print(f"Sheets: {', '.join(categories_order)}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
