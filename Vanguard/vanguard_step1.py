"""
Vanguard Designs — Step 1
=========================
Reads 'Vanguard Designs' sheet from Status Tracker.
Scrapes all category list pages with Selenium (lazy-load + pagination).
Saves to Vanguard/Vanguard_step1.xlsx — one sheet per category.

Usage:
    python vanguard_step1.py
"""

import time
import re
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

SCRIPT_DIR   = Path(__file__).parent
TRACKER_FILE = SCRIPT_DIR.parent / "SD_Web Scraping - Status Tracker.xlsx"
OUT_FILE     = SCRIPT_DIR / "Vanguard_step1.xlsx"
BASE_URL     = "https://www.vanguardfurniture.com"
CHROMEDRIVER = "C:/chromedriver.exe"


# ── Driver ─────────────────────────────────────────────────────────────────────

def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    try:
        import chromedriver_autoinstaller
        chromedriver_autoinstaller.install()
        return webdriver.Chrome(options=opts)
    except Exception:
        return webdriver.Chrome(service=Service(CHROMEDRIVER), options=opts)


# ── Read Status Tracker ────────────────────────────────────────────────────────

def _expand_pipe_url(url: str) -> list[str]:
    """
    Expand URLs with | in ProdType value:
    ?ProdType=037|016  →  [?ProdType=037, ?ProdType=016]
    """
    if "|" not in url:
        return [url]

    base = url.split("?")[0]
    params = url.split("?", 1)[1].split("&") if "?" in url else []
    result = []

    for p in params:
        if p.startswith("ProdType=") and "|" in p:
            key, values = p.split("=", 1)
            others = [x for x in params if x != p]
            for val in values.split("|"):
                new_params = others + [f"{key}={val}"]
                result.append(f"{base}?{'&'.join(new_params)}")
            return result

    return [url]  # fallback


def read_categories() -> list[dict]:
    """
    Parse 'Vanguard Designs' sheet.
    Returns list of {category, urls: [str], first_url: str}
    Merges duplicate category names.
    """
    wb = openpyxl.load_workbook(str(TRACKER_FILE))
    ws = wb["Vanguard Designs"]

    raw: list[dict] = []
    current_cat = None
    current_urls: list[str] = []

    for r in range(1, ws.max_row + 1):
        c_val = str(ws.cell(r, 3).value or "").strip()
        d_val = str(ws.cell(r, 4).value or "").strip()

        if c_val == "Category" and d_val:
            if current_cat and current_urls:
                raw.append({"category": current_cat, "urls": current_urls})
            current_cat = d_val
            current_urls = []

        elif c_val == "Link" and d_val.startswith("http"):
            for expanded in _expand_pipe_url(d_val):
                if expanded not in current_urls:
                    current_urls.append(expanded)

    if current_cat and current_urls:
        raw.append({"category": current_cat, "urls": current_urls})

    wb.close()

    # Merge categories with the same name
    merged: dict[str, list[str]] = {}
    for item in raw:
        name = item["category"]
        if name not in merged:
            merged[name] = []
        for u in item["urls"]:
            if u not in merged[name]:
                merged[name].append(u)

    return [{"category": k, "urls": v, "first_url": v[0]} for k, v in merged.items()]


# ── List page scraper ──────────────────────────────────────────────────────────

def scrape_list_page(driver: webdriver.Chrome, url: str) -> list[dict]:
    """
    Scrape a category list page. All products load on one page (no pagination).
    Selector: .SearchResults_Container
    URL format: /styles/sku/{SKU}
    Image: data-src-150 attribute (lazy-loaded)
    """
    products: list[dict] = []
    seen: set[str] = set()

    print(f"      Fetching: {url}")
    driver.get(url)
    time.sleep(4)

    # Scroll to trigger lazy-load images
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(3)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(0.5)

    items = driver.find_elements(By.CSS_SELECTOR, ".SearchResults_Container")
    print(f"      Found {len(items)} products")

    for item in items:
        try:
            a_tag = item.find_element(By.TAG_NAME, "a")
            href  = (a_tag.get_attribute("href") or "").strip()
            if not href:
                continue

            # Normalize to absolute URL
            if href.startswith("/"):
                href = BASE_URL + href
            product_url = href.split("?")[0]

            if product_url in seen:
                continue
            seen.add(product_url)

            # Image — data-src-150 is lazy-loaded; use 600 version for higher quality
            try:
                img     = a_tag.find_element(By.TAG_NAME, "img")
                img_src = (
                    img.get_attribute("data-src-600") or
                    img.get_attribute("data-src-150") or
                    img.get_attribute("src") or ""
                ).strip()
                sku = (img.get_attribute("alt") or "").strip()
            except Exception:
                img_src = ""
                sku = ""

            # SKU also from URL last segment
            if not sku:
                sku = href.rstrip("/").split("/")[-1]

            # Name: full text is "SKU Name [In Stock]"; strip SKU prefix
            raw_text = item.get_attribute("innerText") or item.text or ""
            raw_text = " ".join(raw_text.split())
            name = raw_text
            if name.startswith(sku):
                name = name[len(sku):].strip()
            # Remove " In Stock" suffix
            name = re.sub(r"\s*In Stock\s*$", "", name, flags=re.IGNORECASE).strip()

            products.append({
                "Product Name": name,
                "SKU":          sku,
                "Product URL":  product_url,
                "Image URL":    img_src,
            })

        except Exception:
            pass

    return products


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    if not TRACKER_FILE.exists():
        print(f"ERROR: Status Tracker not found at {TRACKER_FILE}")
        return

    print("Reading categories from Status Tracker...")
    categories = read_categories()
    print(f"Found {len(categories)} categories:\n")
    for c in categories:
        print(f"  {c['category']:30s} ({len(c['urls'])} URL(s))")

    driver = make_driver()

    wb = openpyxl.Workbook()
    for s in list(wb.sheetnames):
        del wb[s]

    bold    = Font(bold=True)
    headers = ["#", "Category", "Product Name", "SKU", "Product URL", "Image URL", "Category URL"]

    try:
        for cat in categories:
            category  = cat["category"]
            urls      = cat["urls"]
            first_url = cat["first_url"]

            print(f"\n{'='*60}")
            print(f"[{category}]  ({len(urls)} URL(s))")
            print(f"{'='*60}")

            all_products: list[dict] = []
            seen: set[str] = set()

            for url in urls:
                print(f"  URL: {url}")
                prods = scrape_list_page(driver, url)
                for p in prods:
                    if p["Product URL"] not in seen:
                        all_products.append(p)
                        seen.add(p["Product URL"])
                time.sleep(0.5)

            print(f"  Total unique products: {len(all_products)}")

            # Create sheet (handle duplicate names)
            sheet_name = category[:31]
            base_name  = sheet_name
            n = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name[:28]}_{n}"
                n += 1

            ws = wb.create_sheet(sheet_name)
            for col, h in enumerate(headers, 1):
                ws.cell(1, col, h).font = bold

            for i, p in enumerate(all_products, 1):
                ws.cell(i + 1, 1, i)
                ws.cell(i + 1, 2, category)
                ws.cell(i + 1, 3, p["Product Name"])
                ws.cell(i + 1, 4, p["SKU"])
                ws.cell(i + 1, 5, p["Product URL"])
                ws.cell(i + 1, 6, p["Image URL"])
                ws.cell(i + 1, 7, first_url)  # category URL for step2

            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)
            ws.freeze_panes = "A2"

            wb.save(str(OUT_FILE))
            print(f"  Saved sheet '{sheet_name}' ({len(all_products)} products)")

    finally:
        driver.quit()

    wb.save(str(OUT_FILE))
    print(f"\nDone -> {OUT_FILE}")


if __name__ == "__main__":
    main()
