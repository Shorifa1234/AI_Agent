"""
Vanguard Designs — Step 2
=========================
Reads Vanguard_step1.xlsx (one sheet per category).
Visits each product detail page with Selenium.
Extracts: description, dimensions, finish/features.
Saves to Vanguard.xlsx in Julian Chichester format.

Usage:
    python vanguard_step2.py
    python vanguard_step2.py --demo      # 3 products per category (test run)
"""

import re
import sys
import time
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup

SCRIPT_DIR   = Path(__file__).parent
STEP1_FILE   = SCRIPT_DIR / "Vanguard_step1.xlsx"
OUT_FILE     = SCRIPT_DIR / "Vanguard.xlsx"
BASE_URL     = "https://www.vanguardfurniture.com"
CHROMEDRIVER = "C:/chromedriver.exe"
VENDOR       = "Vanguard Designs"

FIXED_COLS = [
    "Index", "Category", "Manufacturer", "Source", "Image URL",
    "Product Name", "SKU", "Product Family Id", "Description",
    "Width", "Depth", "Height", "Diameter", "Finish", "Features",
]


# ── Driver ─────────────────────────────────────────────────────────────────────

def make_driver() -> webdriver.Chrome:
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    try:
        import chromedriver_autoinstaller
        chromedriver_autoinstaller.install()
        return webdriver.Chrome(options=opts)
    except Exception:
        return webdriver.Chrome(service=Service(CHROMEDRIVER), options=opts)


# ── SKU ────────────────────────────────────────────────────────────────────────

def generate_sku(category: str, index: int) -> str:
    """VAN + 2-letter category code + 2-digit index."""
    words = re.sub(r"[^A-Za-z\s]", "", category).strip().split()
    if len(words) >= 2:
        code = (words[0][0] + words[1][0]).upper()
    elif words:
        code = words[0][:2].upper()
    else:
        code = "XX"
    return f"VAN{code}{index:02d}"


# ── Selenium helpers ────────────────────────────────────────────────────────────

def expand_section(driver: webdriver.Chrome, panel_name: str) -> bool:
    """Click a collapsed CCP section to reveal its content."""
    xpath = f"//*[contains(@class,'CCP_Header') and @panel-name='{panel_name}']"
    try:
        header = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        # Check body visibility
        body_xpath = f"//*[contains(@class,'CCP_Body') and @panel-name='{panel_name}']"
        try:
            body  = driver.find_element(By.XPATH, body_xpath)
            style = body.get_attribute("style") or ""
            if "display: none" not in style.lower():
                return True  # Already visible
        except NoSuchElementException:
            pass

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", header)
        time.sleep(0.2)
        driver.execute_script("arguments[0].click();", header)
        time.sleep(0.6)
        return True
    except TimeoutException:
        return False


# ── Dimension parser ────────────────────────────────────────────────────────────

def parse_dimensions(dim_text: str) -> dict[str, str]:
    """
    Parse Vanguard dimension strings (already in inches).
    Handles: "Overall: W 31.75 D 20 H 30", "Dia 18", etc.
    """
    result = {"Width": "", "Depth": "", "Height": "", "Diameter": ""}
    if not dim_text:
        return result

    text = " ".join(dim_text.split())

    # Diameter first — round items have no W/D
    m_dia = re.search(r'\bDia(?:meter)?\.?\s*:?\s*([\d.]+)', text, re.IGNORECASE)
    if m_dia:
        result["Diameter"] = m_dia.group(1)
        return result

    m_w = re.search(r'\bW\s+([\d.]+)', text, re.IGNORECASE)
    m_d = re.search(r'\bD\s+([\d.]+)', text, re.IGNORECASE)
    m_h = re.search(r'\bH\s+([\d.]+)', text, re.IGNORECASE)

    if m_w: result["Width"]  = m_w.group(1)
    if m_d: result["Depth"]  = m_d.group(1)
    if m_h: result["Height"] = m_h.group(1)

    return result


# ── Detail page scraper ─────────────────────────────────────────────────────────

def scrape_detail(driver: webdriver.Chrome, url: str, fallback_name: str = "") -> dict:
    """Fetch product detail page and extract all fields."""
    data: dict[str, str] = {}

    try:
        driver.get(url)
        try:
            WebDriverWait(driver, 12).until(
                EC.presence_of_element_located((By.ID, "divRightSide"))
            )
        except TimeoutException:
            time.sleep(5)
    except Exception as e:
        data["_error"] = str(e)
        return data

    soup = BeautifulSoup(driver.page_source, "html.parser")

    # Product Name — parse from divRightSide text:
    # "Return to Search {Name} SKU: {SKU} ..."
    try:
        right_div = soup.find(id="divRightSide")
        right_text = right_div.get_text(" ", strip=True) if right_div else ""
        m = re.search(r"Return to Search\s+(.+?)\s+SKU\s*:", right_text, re.IGNORECASE)
        if m:
            data["Product Name"] = m.group(1).strip()
    except Exception:
        pass
    if not data.get("Product Name"):
        data["Product Name"] = fallback_name

    # Image — find largest available (1200x1200 preferred, then 600x600)
    try:
        for img in soup.find_all("img"):
            src = img.get("src") or ""
            if "1200x1200" in src and "Styles/" in src:
                data["Image URL"] = src.split("?")[0]  # strip cache-buster
                break
        if not data.get("Image URL"):
            for img in soup.find_all("img"):
                src = img.get("src") or ""
                if "600x600" in src and "Styles/" in src:
                    data["Image URL"] = src.split("?")[0]
                    break
    except Exception:
        pass

    # Description
    try:
        desc_el = soup.find(class_="Romance")
        if desc_el:
            data["Description"] = desc_el.get_text(" ", strip=True)
    except Exception:
        pass

    # Expand Dimensions section (JavaScript click needed)
    expand_section(driver, "Dimensions")

    # Re-parse after JS expansion
    soup2 = BeautifulSoup(driver.page_source, "html.parser")

    # Dimensions
    try:
        dim_grid = soup2.find(class_="DimensionsGrid")
        if dim_grid:
            dims = parse_dimensions(dim_grid.get_text(" ", strip=True))
            data.update(dims)
    except Exception:
        pass

    # Features
    try:
        feat_id = "ctl00_ctl00_ChildBodyContent_ContentPlaceHolderFullWidth_divFeatures"
        feat_el = soup2.find(id=feat_id)
        if feat_el:
            feat_text = feat_el.get_text(" ", strip=True)
            if feat_text:
                data["Features"] = feat_text
    except Exception:
        pass

    # Finish — "As Shown" divs
    try:
        as_shown_els = soup2.find_all(id="divAsShown")
        parts = [el.get_text(" ", strip=True) for el in as_shown_els if el.get_text(strip=True)]
        if parts:
            data["Finish"] = " | ".join(parts)
    except Exception:
        pass

    return data


# ── Excel helpers ───────────────────────────────────────────────────────────────

def save_sheet(
    wb: openpyxl.Workbook,
    category: str,
    category_url: str,
    rows: list[dict],
) -> None:
    sheet_name = category[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    bold = Font(bold=True)

    ws.cell(1, 1, "Brand").font = bold
    ws.cell(1, 2, VENDOR)
    ws.cell(2, 1, "Category Link").font = bold
    ws.cell(2, 2, category_url)

    # Collect any extra columns
    extra: list[str] = []
    for row in rows:
        for k in row:
            if k not in FIXED_COLS and not k.startswith("_") and k not in extra:
                extra.append(k)
    all_cols = FIXED_COLS + extra

    for col, h in enumerate(all_cols, 1):
        ws.cell(4, col, h).font = bold

    for i, row in enumerate(rows, 1):
        row.setdefault("Index",       i)
        row.setdefault("Category",    category)
        row.setdefault("Manufacturer", VENDOR)
        if not row.get("SKU"):
            row["SKU"] = generate_sku(category, i)
        if not row.get("Product Family Id"):
            row["Product Family Id"] = row.get("Product Name", "")

        for col, key in enumerate(all_cols, 1):
            ws.cell(4 + i, col, row.get(key, ""))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A5"


# ── Main ───────────────────────────────────────────────────────────────────────

def main(demo: bool = False) -> None:
    if not STEP1_FILE.exists():
        print(f"ERROR: {STEP1_FILE} not found. Run vanguard_step1.py first.")
        return

    demo_limit = 3 if demo else None
    out_path   = SCRIPT_DIR / ("Vanguard_demo.xlsx" if demo else "Vanguard.xlsx")

    if demo:
        print("[DEMO MODE] 3 products per category -> Vanguard_demo.xlsx\n")

    wb_in = openpyxl.load_workbook(str(STEP1_FILE))

    try:
        wb_out = openpyxl.load_workbook(str(out_path))
    except FileNotFoundError:
        wb_out = openpyxl.Workbook()
        for s in list(wb_out.sheetnames):
            del wb_out[s]

    driver = make_driver()

    try:
        for sheet_name in wb_in.sheetnames:
            ws_in    = wb_in[sheet_name]
            category = sheet_name

            # Step1 columns: #, Category, Product Name, SKU, Product URL, Image URL, Category URL
            products_in: list[dict] = []
            category_url = ""

            for row in ws_in.iter_rows(min_row=2, values_only=True):
                url = str(row[4] or "").strip()
                if not url:
                    continue
                cat_url = str(row[6] or "").strip()
                if cat_url and not category_url:
                    category_url = cat_url
                products_in.append({
                    "name":  str(row[2] or ""),
                    "sku":   str(row[3] or ""),
                    "url":   url,
                    "image": str(row[5] or ""),
                })

            if not products_in:
                print(f"\nSkipping '{category}' — no products in step1 file.")
                continue

            if demo_limit:
                products_in = products_in[:demo_limit]

            print(f"\n{'='*60}")
            print(f"[{category}]  {len(products_in)} products")
            print(f"{'='*60}")

            output_rows: list[dict] = []

            for i, p in enumerate(products_in, 1):
                safe_name = (p["name"][:55] or p["url"]).encode("ascii", "replace").decode("ascii")
                print(f"  [{i}/{len(products_in)}] {safe_name} ...", end=" ", flush=True)

                detail = scrape_detail(driver, p["url"], fallback_name=p["name"])

                if "_error" in detail:
                    print(f"ERROR: {detail['_error']}")
                else:
                    print("OK")

                row: dict = {
                    "Product Name": detail.get("Product Name") or p["name"],
                    "Source":       p["url"],
                    "Image URL":    detail.get("Image URL")  or p["image"],
                    "Description":  detail.get("Description", ""),
                    "Width":        detail.get("Width",       ""),
                    "Depth":        detail.get("Depth",       ""),
                    "Height":       detail.get("Height",      ""),
                    "Diameter":     detail.get("Diameter",    ""),
                    "Finish":       detail.get("Finish",      ""),
                    "Features":     detail.get("Features",    ""),
                    "SKU":          p["sku"],
                }
                output_rows.append(row)
                time.sleep(0.8)

            save_sheet(wb_out, category, category_url, output_rows)
            wb_out.save(str(out_path))
            print(f"  Saved '{category}' ({len(output_rows)} products)")

    finally:
        driver.quit()

    wb_out.save(str(out_path))
    print(f"\nDone -> {out_path}")


if __name__ == "__main__":
    main(demo="--demo" in sys.argv)
